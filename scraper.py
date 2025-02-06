import time
from playwright.async_api import async_playwright
import asyncio
import pandas as pd
import random
from io import StringIO
import re
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from bs4 import BeautifulSoup
import yfinance as yf
import aiohttp
import json


class StockScraper:
    def __init__(self, stocks, template_path, headless=True, max_concurrent=3):
        """
        初始化爬蟲類別。
        stocks: 股票代碼的列表
        headless: 是否使用無頭模式
        max_concurrent: 同時執行的股票數量（控制併發數）
        """
        self.stocks = stocks
        self.template_path = template_path
        self.headless = headless
        self.max_concurrent = max_concurrent  # 限制同時執行的併發數量
        self.browser = None
        self.playwright = None

    async def setup_browser(self):
        """設定瀏覽器環境。"""
        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.launch(
            headless=self.headless,
            args=[
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",
                "--disable-accelerated-2d-canvas",
                "--disable-gpu",
                "--disable-features=IsolateOrigins,site-per-process",
                "--disable-background-timer-throttling",
            ],
        )

    async def cleanup(self):
        """清理資源。"""
        if self.browser:
            await self.browser.close()  # 確保瀏覽器關閉
        if self.playwright:
            await self.playwright.stop()  # 停止 Playwright

    async def fetch_summary_data(self, stock, semaphore):
        """
        抓取單一股票的數據（summary）。
        """
        async with semaphore:  # 控制併發數量
            try:
                # 建立新的瀏覽器上下文
                context = await self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
                    viewport={"width": 800, "height": 600},
                )
                try:
                    # 為每個數據抓取建立頁面
                    page_summary = await context.new_page()
                    # 使用 asyncio.gather 同時執行抓取任務
                    summary = await asyncio.gather(self.get_summary(stock, page_summary))

                    # return ratios
                    return {stock: summary}

                finally:
                    await context.close()  # 確保上下文被正確關閉
            except Exception as e:
                return {"stock": stock, "error": str(e)}

    async def get_summary(self, stock, page, retries=1):
        """抓取特定股票的摘要資料並回傳 DataFrame。"""
        URL = f'https://www.roic.ai/quote/{stock}'
        attempt = 0

        while attempt < retries:
            try:
                await asyncio.sleep(random.uniform(1, 3))  # 模擬人工操作
                '''
                依狀況選擇
                'load'：等到整個頁面載入完成（包括圖片、JS等）。
                'domcontentloaded'：只等到 HTML 的基本結構載入完成。
                'networkidle'：等到沒有額外的網路請求在進行時。
                'timeout'：等多久（以毫秒為單位）才判定超時，預設是 30 秒（30000 毫秒）
                '''
                await page.goto(URL, wait_until='load', timeout=50000)
                # await page.wait_for_selector('table[style="width: 1199.8px;"]', timeout=50000)
                await page.wait_for_selector('table.w-full.caption-bottom.text-sm.table-fixed', timeout=100000)
                content = await page.content()
                dfs = pd.read_html(StringIO(content))
                return dfs
            except Exception as e:
                attempt += 1
                if attempt == retries:
                    return f"Error for {stock}: {e}"

        return f"Failed to retrieve data for {stock}"

    async def run_summary(self):
        await self.setup_browser()
        semaphore = asyncio.Semaphore(self.max_concurrent)  # 限制同時執行的任務數量
        try:
            tasks = [self.fetch_summary_data(stock, semaphore) for stock in self.stocks]
            result = await asyncio.gather(*tasks)
        finally:
            await self.cleanup()
        return result

    async def fetch_financials_data(self, stock, semaphore):
        """
        抓取單一股票的數據（financials）。
        """
        async with semaphore:  # 控制併發數量
            try:
                # 建立新的瀏覽器上下文
                context = await self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
                    viewport={"width": 800, "height": 600},
                )
                try:
                    # 為每個數據抓取建立頁面
                    page_financials = await context.new_page()
                    # 使用 asyncio.gather 同時執行抓取任務
                    financials = await asyncio.gather(self.get_financials(stock, page_financials))

                    # return financials
                    return {stock: financials}

                finally:
                    await context.close()  # 確保上下文被正確關閉
            except Exception as e:
                return {"stock": stock, "error": str(e)}

    async def get_financials(self, stock, page, retries=1):
        """抓取特定股票的摘要資料並回傳 DataFrame。"""
        URL = f'https://www.roic.ai/quote/{stock}/financials'
        attempt = 0

        while attempt < retries:
            try:
                await asyncio.sleep(random.uniform(1, 3))  # 模擬人工操作
                '''
                依狀況選擇
                'load'：等到整個頁面載入完成（包括圖片、JS等）。
                'domcontentloaded'：只等到 HTML 的基本結構載入完成。
                'networkidle'：等到沒有額外的網路請求在進行時。
                'timeout'：等多久（以毫秒為單位）才判定超時，預設是 30 秒（30000 毫秒）
                '''
                await page.goto(URL, wait_until='load', timeout=50000)
                # await page.wait_for_selector('table[style="width: 1199.8px;"]', timeout=50000)
                # await page.wait_for_selector('table.w-full.caption-bottom.text-sm.table-fixed', timeout=100000)
                if await page.query_selector(
                        'div.rounded-lg.bg-card.text-card-foreground.shadow-sm.mx-auto.flex.w-\\[500px\\].flex-col.items-center.border.drop-shadow-lg'):
                    return f'{stock}是非美國企業，此頁面須付費！'
                else:
                    await page.wait_for_selector('table.w-full.caption-bottom.text-sm.table-fixed', timeout=100000)
                    content = await page.content()
                    dfs = pd.read_html(StringIO(content))
                    return dfs

            except Exception as e:
                attempt += 1
                if attempt == retries:
                    return f"Error for {stock}: {e}"

        return f"Failed to retrieve data for {stock}"

    async def run_financial(self):
        await self.setup_browser()
        semaphore = asyncio.Semaphore(self.max_concurrent)  # 限制同時執行的任務數量
        try:
            tasks = [self.fetch_financials_data(stock, semaphore) for stock in self.stocks]
            result = await asyncio.gather(*tasks)
        finally:
            await self.cleanup()
        return result

    async def fetch_ratios_data(self, stock, semaphore):
        """
        抓取單一股票的數據（Ratios）。
        """
        async with semaphore:  # 控制併發數量
            try:
                # 建立新的瀏覽器上下文
                context = await self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
                    viewport={"width": 800, "height": 600},
                )
                try:
                    # 為每個數據抓取建立頁面
                    page_ratios = await context.new_page()
                    # 使用 asyncio.gather 同時執行抓取任務
                    ratios = await asyncio.gather(self.get_ratios(stock, page_ratios))

                    # return ratios
                    return {stock: ratios}

                finally:
                    await context.close()  # 確保上下文被正確關閉
            except Exception as e:
                return {"stock": stock, "error": str(e)}

    async def get_ratios(self, stock, page, retries=1):
        """抓取特定股票的摘要資料並回傳 DataFrame。"""
        URL = f'https://www.roic.ai/quote/{stock}/ratios'
        attempt = 0

        while attempt < retries:
            try:
                await asyncio.sleep(random.uniform(1, 3))  # 模擬人工操作
                '''
                依狀況選擇
                'load'：等到整個頁面載入完成（包括圖片、JS等）。
                'domcontentloaded'：只等到 HTML 的基本結構載入完成。
                'networkidle'：等到沒有額外的網路請求在進行時。
                'timeout'：等多久（以毫秒為單位）才判定超時，預設是 30 秒（30000 毫秒）
                '''
                await page.goto(URL, wait_until='load', timeout=50000)
                # await page.wait_for_selector('table[style="width: 1199.8px;"]', timeout=50000)
                if await page.query_selector(
                        'div.rounded-lg.bg-card.text-card-foreground.shadow-sm.mx-auto.flex.w-\\[500px\\].flex-col.items-center.border.drop-shadow-lg'):
                    return f'{stock}是非美國企業，此頁面須付費！'
                else:
                    await page.wait_for_selector('table.w-full.caption-bottom.text-sm.table-fixed', timeout=100000)
                    content = await page.content()
                    dfs = pd.read_html(StringIO(content))
                    return dfs
            except Exception as e:
                attempt += 1
                if attempt == retries:
                    return f"Error for {stock}: {e}"

        return f"Failed to retrieve data for {stock}"

    async def run_ratios(self):
        await self.setup_browser()
        semaphore = asyncio.Semaphore(self.max_concurrent)  # 限制同時執行的任務數量
        try:
            tasks = [self.fetch_ratios_data(stock, semaphore) for stock in self.stocks]
            result = await asyncio.gather(*tasks)
        finally:
            await self.cleanup()
        return result

    async def fetch_EPS_PE_MarketCap_data(self, stock, semaphore):
        """
        抓取單一股票的數據（EPS_PE_MarketCap）。
        """
        async with semaphore:  # 控制併發數量
            try:
                # 建立新的瀏覽器上下文
                context = await self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
                    viewport={"width": 800, "height": 600},
                )
                try:
                    # 為每個數據抓取建立頁面
                    page_EPS_PE_MarketCap = await context.new_page()
                    # 使用 asyncio.gather 同時執行抓取任務
                    EPS_PE_MarketCap = await asyncio.gather(self.get_EPS_PE_MarketCap(stock, page_EPS_PE_MarketCap))

                    # return EPS_PE_MarketCap
                    return {stock: EPS_PE_MarketCap}

                finally:
                    await context.close()  # 確保上下文被正確關閉
            except Exception as e:
                return {"stock": stock, "error": str(e)}

    async def get_EPS_PE_MarketCap(self, stock, page, retries=3):
        url = f'https://www.roic.ai/quote/{stock}'
        attempt = 0
        while attempt < retries:
            try:
                await asyncio.sleep(random.uniform(1, 3))
                await page.goto(url, wait_until='load', timeout=30000)
                # await page.goto(url, wait_until='networkidle', timeout=50000)
                await page.wait_for_selector('table', timeout=30000)
                content = await page.content()
                soup = BeautifulSoup(content, 'html.parser')
                span_string = soup.find_all('span', attrs={'class': 'flex text-sm uppercase text-muted-foreground'})
                span_int_value = soup.find_all('span', attrs={'class': 'flex text-lg text-foreground'})

                dic_data = {span_string[0].text: float(span_int_value[0].text),
                            span_string[1].text: float(span_int_value[1].text),
                            span_string[2].text: span_int_value[2].text, span_string[3].text: span_int_value[3].text}
                return dic_data
            except Exception as e:
                return f'error message:{e}'

    async def run_EPS_PE_MarketCap(self):
        await self.setup_browser()
        semaphore = asyncio.Semaphore(self.max_concurrent)  # 限制同時執行的任務數量
        try:
            tasks = [self.fetch_EPS_PE_MarketCap_data(stock, semaphore) for stock in self.stocks]
            result = await asyncio.gather(*tasks)
        finally:
            await self.cleanup()
        return result

    async def EPS_Growth_Rate_and_write_to_excel(self, stock):
        file_path = fr'{self.template_path}/STOCK_{stock}.xlsx'

        if '-' in stock:
            stock = ''.join(['.' if char == '-' else char for char in stock])

        async with aiohttp.ClientSession() as session:
            async with session.get(fr'https://api.stockboss.io/api/symbol?symbol={stock}') as response:
                content = await response.text()
                dic = json.loads(content)
                l_eps_growth5y = []
                try:
                    EPS_Growth_Rate_3_Year = \
                        dic['symbol']['keyratio']['keyratio']['annuals']['3-Year EPS Growth Rate %'][
                            -1]
                    EPS_Growth_Rate_5_Year = \
                        dic['symbol']['keyratio']['keyratio']['annuals']['5-Year EPS Growth Rate %'][
                            -1]
                    EPS_Growth_Rate_10_Year = \
                        dic['symbol']['keyratio']['keyratio']['annuals']['10-Year EPS Growth Rate %'][
                            -1]

                    EPS_Growth_Rate_3_Year = 0 if EPS_Growth_Rate_3_Year == '-' else EPS_Growth_Rate_3_Year
                    EPS_Growth_Rate_5_Year = 0 if EPS_Growth_Rate_5_Year == '-' else EPS_Growth_Rate_5_Year
                    EPS_Growth_Rate_10_Year = 0 if EPS_Growth_Rate_10_Year == '-' else EPS_Growth_Rate_10_Year

                    l_eps_growth5y = l_eps_growth5y + [EPS_Growth_Rate_3_Year, EPS_Growth_Rate_5_Year,
                                                       EPS_Growth_Rate_10_Year]
                    # print(f'{stock}: {l_eps_growth5y}')
                except KeyError as e:
                    return f"EPS_Growth_Rate的dictionary鍵錯誤：{stock}"

                # 選擇成長率：如果最小值大於 0，則取最小值，否則取最大值
                selected_growth_rate = min(l_eps_growth5y) / 100 if min(l_eps_growth5y) > 0 else max(
                    l_eps_growth5y) / 100
                # print(selected_growth_rate)
                # 寫入 Excel
                try:
                    wb = load_workbook(file_path)
                    ws = wb.worksheets[3]  # 假設需要寫入的工作表是第四個
                    ws['C4'] = selected_growth_rate
                    wb.save(file_path)

                    return f"{stock}的EPS成長率成功寫入及儲存成功"

                except FileNotFoundError:
                    return f"找不到此檔案，請確認路徑！您目前提供的路徑如下：{file_path}"
                except PermissionError:
                    return f"存檔失敗：請先關閉Excel再執行程式！請先關閉Excel檔案路徑：{file_path}"


class StockProcess:
    def __init__(self, template_path):
        self.template_path = template_path

    async def process_df_summary(self, raw_df_summary, stock):
        # print(raw_df_summary)

        if raw_df_summary:
            # print(raw_df_summary)
            file_path = fr'{self.template_path}/STOCK_{stock}.xlsx'
            # print(file_path)
            try:
                wb = load_workbook(file_path)
                ws = wb.worksheets[0]
            except FileNotFoundError:
                return f'找不到此檔案，請確認路徑！您目前提供的路徑如下：{file_path}'

            # 清除舊資料 & 初始化寫入緩存
            try:
                for row in ws.iter_rows(min_row=1, min_col=1, max_row=30, max_col=12):
                    for cell in row:
                        cell.value = None

                # print(f'{stock}成功清除舊資料')
            except Exception as e:
                return f'清除資料時發生錯誤：{e}'

            d_1_raw_df_summary = [y for x in raw_df_summary.get(stock, pd.DataFrame({})) for y in x]
            # print(d_1_raw_df_summary)
            for df_amount, df in enumerate(d_1_raw_df_summary):
                df_column_list = df.columns.tolist()
                # print(df_column_list)
                # print(df)
                # 篩選需要的年份
                years_list = []
                pattern = r'\d{4}\sY'
                for years in df_column_list:
                    if re.match(pattern, years):
                        years_list.append(int(years.split()[0]))

                if years_list:
                    end_year = max(years_list)
                    start_year = end_year - 10
                    drop_column = [
                        x for x in df_column_list
                        if not (re.match(pattern, x) and start_year <= int(x.split()[0]) <= end_year)
                    ]
                    drop_column.pop(0)  # 移除列名的第一欄
                    df = df.drop(columns=drop_column)

                    # 資料轉型為數值型
                    years_data = df.columns[1:]
                    df[years_data] = df[years_data].apply(pd.to_numeric, errors='coerce')
                    # print(df)
                    # 將資料寫入 Excel，並設置欄位格式
                    start_row, start_column = 1, 1
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True),
                                                start=start_row):
                        for c_idx, value in enumerate(row, start=start_column):
                            cell = ws.cell(row=r_idx, column=c_idx, value=value)
                            cell.font = Font(size=12, bold=(r_idx == start_row))

                    # 自動調整欄寬
                    for col in ws.columns:
                        max_length = max(len(str(cell.value or '')) for cell in col)
                        ws.column_dimensions[col[0].column_letter].width = max_length + 2
                    # print(f"{stock}的Summary成功寫入")
            try:
                wb.save(file_path)  # 一次性儲存所有修改
                return f"{stock}的Summary成功寫入及儲存成功"
            except PermissionError:
                return f"存檔失敗：請先關閉Excel再執行程式！請先關閉Excel檔案路徑：{file_path}"

        else:
            return '無原始資料'

    async def process_df_financial(self, raw_df_financial, stock):
        if raw_df_financial:
            file_path = fr'{self.template_path}/STOCK_{stock}.xlsx'
            # print(file_path)
            try:
                wb = load_workbook(file_path)
                ws = wb.worksheets[0]
            except FileNotFoundError:
                return f'找不到此檔案，請確認路徑！您目前提供的路徑如下：{file_path}'

            # 定義各類財務數據的起始位置
            starting_cell = [("IncomeStatement", 1, 14),  # N1
                             ("BalanceSheet", 1, 27),  # AA1
                             ("CashFlowStatement", 1, 40)]  # AN1
            # 清除舊資料 & 初始化寫入緩存
            try:
                # 刪除col_N到col_AN
                for row in ws.iter_rows(min_row=1, min_col=14, max_row=100, max_col=25):
                    # print(row)
                    for cell in row:
                        cell.value = None

                for row in ws.iter_rows(min_row=1, min_col=27, max_row=100, max_col=38):
                    # print(row)
                    for cell in row:
                        cell.value = None

                for row in ws.iter_rows(min_row=1, min_col=40, max_row=100, max_col=51):
                    # print(row)
                    for cell in row:
                        cell.value = None

                # print(f'{stock}成功清除舊資料')
            except Exception as e:
                return f'清除資料時發生錯誤：{e}'
            # print(raw_df_financial)
            if raw_df_financial.get(stock) == [f'{stock}是非美國企業，此頁面須付費！']:
                return f'{stock}是非美國企業，此頁面須付費！'

            else:
                d_1_raw_df_financial = [y for x in raw_df_financial.get(stock, pd.DataFrame({})) for y in x]
                # print(d_1_raw_df_financial)
                for df_amount, df in enumerate(d_1_raw_df_financial):
                    df_column_list = df.columns.tolist()
                    # print(df_column_list)
                    # print(df)
                    # 篩選需要的年份
                    years_list = []
                    pattern = r'\d{4}\sY'
                    for years in df_column_list:
                        if re.match(pattern, years):
                            years_list.append(int(years.split()[0]))

                    if years_list:
                        end_year = max(years_list)
                        start_year = end_year - 10
                        drop_column = [
                            x for x in df_column_list
                            if not (re.match(pattern, x) and start_year <= int(x.split()[0]) <= end_year)
                        ]
                        drop_column.pop(0)  # 移除列名的第一欄
                        df = df.drop(columns=drop_column)

                        # 資料轉型為數值型
                        years_data = df.columns[1:]
                        df[years_data] = df[years_data].apply(pd.to_numeric, errors='coerce')
                        # print(df)

                        # 將表格資料寫入指定位置並調整格式
                        start_row, start_col = starting_cell[df_amount][1], starting_cell[df_amount][2]
                        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True),
                                                    start=start_row):
                            for c_idx, value in enumerate(row, start=start_col):
                                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                                cell.font = Font(size=12, bold=(r_idx == start_row))

                        # 自動調整欄寬
                        for col in ws.columns:
                            max_length = max(len(str(cell.value or '')) for cell in col)
                            ws.column_dimensions[col[0].column_letter].width = max_length + 2
                        # print(f"{stock}的Financial成功寫入")
            try:
                wb.save(file_path)  # 一次性儲存所有修改
                return f"{stock}的Financial成功寫入及儲存成功"
            except PermissionError:
                return f"存檔失敗：請先關閉Excel再執行程式！請先關閉Excel檔案路徑：{file_path}"

        else:
            return '無原始資料'

    async def process_df_ratios(self, raw_df_ratios, stock):
        if raw_df_ratios:
            # print(raw_df_ratios)
            file_path = fr'{self.template_path}/STOCK_{stock}.xlsx'
            # print(file_path)
            try:
                wb = load_workbook(file_path)
                ws = wb.worksheets[0]
            except FileNotFoundError:
                return f'找不到此檔案，請確認路徑！您目前提供的路徑如下：{file_path}'

            # 定義各類財務數據的起始位置
            starting_cell = [
                ('Profitability', 1, 53), ('Credit', 1, 66),
                ('Liquidity', 1, 79), ('Working Capital', 1, 92),
                ('Enterprise Value', 1, 105), ('Multiples', 1, 118),
                ('Per Share Data Items', 1, 131)
            ]
            # 清除舊資料 & 初始化寫入緩存
            try:
                for row in ws.iter_rows(min_row=1, min_col=53, max_row=100, max_col=64):
                    for cell in row:
                        cell.value = None

                for row in ws.iter_rows(min_row=1, min_col=66, max_row=100, max_col=77):
                    for cell in row:
                        cell.value = None

                for row in ws.iter_rows(min_row=1, min_col=79, max_row=100, max_col=90):
                    for cell in row:
                        cell.value = None

                for row in ws.iter_rows(min_row=1, min_col=92, max_row=100, max_col=103):
                    for cell in row:
                        cell.value = None

                for row in ws.iter_rows(min_row=1, min_col=105, max_row=100, max_col=116):
                    for cell in row:
                        cell.value = None

                for row in ws.iter_rows(min_row=1, min_col=118, max_row=100, max_col=129):
                    for cell in row:
                        cell.value = None

                for row in ws.iter_rows(min_row=1, min_col=131, max_row=100, max_col=142):
                    for cell in row:
                        cell.value = None

                # print(f'{stock}成功清除舊資料')
            except Exception as e:
                return f'清除資料時發生錯誤：{e}'
            if raw_df_ratios.get(stock) == [f'{stock}是非美國企業，此頁面須付費！']:
                return f'{stock}是非美國企業，此頁面須付費！'
            else:
                d_1_raw_df_ratios = [y for x in raw_df_ratios.get(stock, pd.DataFrame({})) for y in x]
                # print(d_1_raw_df_financial)
                for df_amount, df in enumerate(d_1_raw_df_ratios):
                    df_column_list = df.columns.tolist()
                    # print(df_column_list)
                    # print(df)
                    # 篩選需要的年份
                    years_list = []
                    pattern = r'\d{4}\sY'
                    for years in df_column_list:
                        if re.match(pattern, years):
                            years_list.append(int(years.split()[0]))

                    if years_list:
                        end_year = max(years_list)
                        start_year = end_year - 10
                        drop_column = [
                            x for x in df_column_list
                            if not (re.match(pattern, x) and start_year <= int(x.split()[0]) <= end_year)
                        ]
                        drop_column.pop(0)  # 移除列名的第一欄
                        df = df.drop(columns=drop_column)

                        # 資料轉型為數值型
                        years_data = df.columns[1:]
                        df[years_data] = df[years_data].apply(pd.to_numeric, errors='coerce')
                        # print(df)

                        # 將表格資料寫入指定位置並調整格式
                        start_row, start_col = starting_cell[df_amount][1], starting_cell[df_amount][2]
                        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True),
                                                    start=start_row):
                            for c_idx, value in enumerate(row, start=start_col):
                                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                                cell.font = Font(size=12, bold=(r_idx == start_row))

                        # 自動調整欄寬
                        for col in ws.columns:
                            max_length = max(len(str(cell.value or '')) for cell in col)
                            ws.column_dimensions[col[0].column_letter].width = max_length + 2
                        # print(f"{stock}的Ratios成功寫入")

            try:
                wb.save(file_path)  # 一次性儲存所有修改
                return f"{stock}的Ratios成功寫入及儲存成功"
            except PermissionError:
                return f"存檔失敗：請先關閉Excel再執行程式！請先關閉Excel檔案路徑：{file_path}"

        else:
            return '無原始資料'

    async def EPS_PE_MarketCap_data_write_to_excel(self, EPS_PE_MarketCap_content, stock):
        file_path = fr'{self.template_path}/STOCK_{stock}.xlsx'
        # print(file_path)
        try:
            wb = load_workbook(file_path)
            ws = wb.worksheets[0]
        except FileNotFoundError:
            return f'找不到此檔案，請確認路徑！您目前提供的路徑如下：{file_path}'
        # print(EPS_PE_MarketCap_content)
        # 處理資料
        for data in EPS_PE_MarketCap_content.get(stock, dict):
            # print(f'{stock}:{data}')
            start_cell = "EN1"
            start_row = int(start_cell[2:])  # 提取行號，例如 "1"

            for i, (key, value) in enumerate(data.items()):
                row = start_row + i  # 從起始行開始逐行寫入
                ws[f"EN{row}"] = key  # 寫入鍵到 EN 列
                ws[f"EO{row}"] = value  # 寫入值到 EO 列

        # 嘗試保存檔案
        try:
            wb.save(file_path)
            return f'{stock}的EPS_PE_MarketCap成功寫入及儲存成功'
        except PermissionError:
            return f"存檔失敗：請先關閉Excel再執行程式！請先關閉Excel檔案路徑：{file_path}"

    async def others_data(self, stock):
        # 查詢 10 年期美國國債收益率
        tnx = yf.Ticker("^TNX")
        # print(tnx.info)
        rf_rate = tnx.info['previousClose'] / 100  # 將 TNX 的值轉換為百分比（例如 4.575 轉為 0.4575）
        # rf_rate = tnx.info['fiftyDayAverage'] / 100  # 使用 50 天平均值作為無風險利率

        Stock = yf.Ticker(stock)
        # print(Stock.info)
        beta = Stock.info['beta']
        currentPrice = Stock.info['currentPrice']
        symbol = Stock.info['symbol']
        # dividendYield = Stock.info['dividendYield']
        # marketCap = Stock.info['marketCap']
        # PE_ratio = f"{Stock.info['trailingPE']:.2f}"
        dic_data = {'Stock': symbol, 'CurrentPrice': currentPrice, 'beta': beta,
                    'rf_rate': rf_rate}
        # print(f'{stock}:{dic_data}')

        # Step 2: 打開 Excel 檔案
        file_path = fr'{self.template_path}/STOCK_{stock}.xlsx'

        try:
            wb = load_workbook(file_path)  # 使用 openpyxl 開啟檔案

        except FileNotFoundError:
            return f'找不到此檔案，請確認路徑！您目前提供的路徑如下：{file_path}'

        ws = wb.worksheets[0]  # 選擇第一個工作表
        # Step 3: 寫入資料到 EQ2 和 ER2
        ws['EQ2'] = dic_data['Stock']  # 將股票代號寫入 EQ2
        ws['ER2'] = dic_data['CurrentPrice']  # 將目前價格寫入 ER2

        ws = wb.worksheets[3]
        ws['C31'] = dic_data['beta']
        ws['C36'] = dic_data['rf_rate']
        # Step 4: 儲存檔案
        try:
            wb.save(file_path)  # 儲存修改後的檔案
            return f'{stock}的其他資料成功寫入及儲存成功'

        except PermissionError:
            return f"存檔失敗：請先關閉Excel再執行程式！請先關閉Excel檔案路徑：{file_path}"


class StockManager:
    def __init__(self, scraper, processor, template_path):
        self.scraper = scraper
        self.processor = processor
        self.template_path = template_path
        # 抓取stock代碼pattern
        self.pattern1 = r'^[a-zA-Z\-\.]{1,5}'
        # 抓取非美國stock代碼pattern
        self.pattern2 = r'是非美國企業，此頁面須付費！$'
        # 抓取FileNotFound以及path
        self.pattern3 = fr'找不到此檔案，請確認路徑！您目前提供的路徑如下：(.*)'
        # 抓取FilePermissionError以及path
        self.pattern4 = fr'存檔失敗：請先關閉Excel再執行程式！請先關閉Excel檔案路徑：(.*)'

    async def process_summary(self, stocks):
        raw_df_summary = await self.scraper.run_summary()
        tasks = [self.processor.process_df_summary(raw_df_summary=raw_df_summary[index], stock=stock) for index, stock
                 in enumerate(stocks)]
        results = await asyncio.gather(*tasks)  # 並行處理所有股票的任務
        # print(results)
        match_list, file_not_found_list, file_permission_list = [], [], []
        for result in results:
            match1 = re.match(self.pattern1, result)
            file_not_found_match = re.match(self.pattern3, result)
            file_permission_match = re.match(self.pattern4, result)
            if match1:
                match_list.append(match1.group())
            elif file_not_found_match:
                file_not_found_list.append(file_not_found_match.group(1))
            elif file_permission_match:
                file_permission_list.append(file_permission_match.group(1))

        if match_list:
            print(f"{', '.join([re.match(self.pattern1, m).group() for m in match_list])}的Summary成功寫入及儲存成功")
        if file_not_found_list:
            file_path = '\n'.join(file_not_found_list)
            print(f"找不到此檔案，請確認路徑！您目前提供的路徑如下：\n{file_path}'")
        if file_permission_list:
            file_path = '\n'.join(file_permission_list)
            print(f"""存檔失敗：請先關閉Excel再執行程式！請先關閉Excel檔案路徑：\n{file_path}""")

    async def process_financial(self, stocks):
        raw_df_financial = await self.scraper.run_financial()
        # print(raw_df_financial)
        tasks = [self.processor.process_df_financial(raw_df_financial[index], stock=stock) for index, stock in
                 enumerate(stocks)]
        results = await asyncio.gather(*tasks)  # 並行處理所有股票的任務
        # print(results)
        match_list, unmatch_list_1, file_not_found_list, file_permission_list = [], [], [], []
        for index, result in enumerate(results):
            match1 = re.match(self.pattern1, result)
            match2 = re.search(self.pattern2, result)
            file_not_found_match = re.search(self.pattern3, result)
            file_permission_match = re.match(self.pattern4, result)
            if match1 and not match2:
                match_list.append(match1.group())
            elif match2:
                unmatch_list_1.append(result)
            elif file_not_found_match:
                file_not_found_list.append(file_not_found_match.group(1))
            elif file_permission_match:
                file_permission_list.append(file_permission_match.group(1))

        # print(match_list)
        # print(unmatch_list_1)
        # print(unmatch_list_2)
        if match_list:
            print(f"{', '.join([re.match(self.pattern1, m).group() for m in match_list])}的Financial成功寫入及儲存成功")
        if unmatch_list_1:
            print(
                f"{', '.join([re.match(self.pattern1, u).group() for u in unmatch_list_1])}是非美國企業，此頁面須付費！")
        if file_not_found_list:
            file_path = '\n'.join(file_not_found_list)
            print(f"找不到此檔案，請確認路徑！您目前提供的路徑如下：\n{file_path}'")
        if file_permission_list:
            file_path = '\n'.join(file_permission_list)
            print(f'存檔失敗：請先關閉Excel再執行程式！請先關閉Excel檔案路徑：\n{file_path}')

    async def process_ratios(self, stocks):
        raw_df_ratios = await self.scraper.run_ratios()
        tasks = [self.processor.process_df_ratios(raw_df_ratios=raw_df_ratios[index], stock=stock) for index, stock in
                 enumerate(stocks)]
        results = await asyncio.gather(*tasks)  # 並行處理所有股票的任務
        match_list, unmatch_list, file_not_found_list, file_permission_list = [], [], [], []
        for result in results:
            match1 = re.match(self.pattern1, result)
            match2 = re.search(self.pattern2, result)
            file_not_found_match = re.search(self.pattern3, result)
            file_permission_match = re.match(self.pattern4, result)
            if match1 and not match2:
                match_list.append(match1.group())
            elif match2:
                unmatch_list.append(result)
            elif file_not_found_match:
                file_not_found_list.append(file_not_found_match.group(1))
            elif file_permission_match:
                file_permission_list.append(file_permission_match.group(1))

        if match_list:
            print(f"{', '.join([re.match(self.pattern1, m).group() for m in match_list])}的Ratios成功寫入及儲存成功")
        if unmatch_list:
            print(f"{', '.join([re.match(self.pattern1, u).group() for u in unmatch_list])}是非美國企業，此頁面須付費！")
        if file_not_found_list:
            file_path = '\n'.join(file_not_found_list)
            print(f"找不到此檔案，請確認路徑！您目前提供的路徑如下：\n{file_path}'")
        if file_permission_list:
            file_path = '\n'.join(file_permission_list)
            print(f'存檔失敗：請先關閉Excel再執行程式！請先關閉Excel檔案路徑：\n{file_path}')

    async def process_EPS_PE_MarketCap(self, stocks):
        raw_df_EPS_PE_MarketCap = await self.scraper.run_EPS_PE_MarketCap()
        tasks = [
            self.processor.EPS_PE_MarketCap_data_write_to_excel(EPS_PE_MarketCap_content=raw_df_EPS_PE_MarketCap[index],
                                                                stock=stock) for index, stock in enumerate(stocks)]
        results = await asyncio.gather(*tasks)  # 並行處理所有股票的任務
        match_list, file_not_found_list, file_permission_list = [], [], []

        for result in results:
            match1 = re.match(self.pattern1, result)
            file_not_found_match = re.search(self.pattern3, result)
            file_permission_match = re.match(self.pattern4, result)
            if match1:
                match_list.append(match1.group())
            elif file_not_found_match:
                file_not_found_list.append(file_not_found_match.group(1))
            elif file_permission_match:
                file_permission_list.append(file_permission_match.group(1))

        if match_list:
            print(
                f"{', '.join([re.match(self.pattern1, m).group() for m in match_list])}的EPS_PE_MarketCap成功寫入及儲存成功")
        if file_not_found_list:
            file_path = '\n'.join(file_not_found_list)
            print(f"找不到此檔案，請確認路徑！您目前提供的路徑如下：\n{file_path}'")
        if file_permission_list:
            file_path = '\n'.join(file_permission_list)
            print(f'存檔失敗：請先關閉Excel再執行程式！請先關閉Excel檔案路徑：\n{file_path}')

    async def process_others_data(self, stocks):
        tasks = [self.processor.others_data(stock) for stock in stocks]
        results = await asyncio.gather(*tasks)  # 並行處理所有股票的任務
        match_list, file_not_found_list, file_permission_list = [], [], []

        for result in results:
            match1 = re.match(self.pattern1, result)
            file_not_found_match = re.search(self.pattern3, result)
            file_permission_match = re.match(self.pattern4, result)
            if match1:
                match_list.append(match1.group())
            elif file_not_found_match:
                file_not_found_list.append(file_not_found_match.group(1))
            elif file_permission_match:
                file_permission_list.append(file_permission_match.group(1))

        if match_list:
            print(f"{', '.join([re.match(self.pattern1, m).group() for m in match_list])}的其他資料成功寫入及儲存成功")
        if file_not_found_list:
            file_path = '\n'.join(file_not_found_list)
            print(f"找不到此檔案，請確認路徑！您目前提供的路徑如下：\n{file_path}'")
        if file_permission_list:
            file_path = '\n'.join(file_permission_list)
            print(f'存檔失敗：請先關閉Excel再執行程式！請先關閉Excel檔案路徑：\n{file_path}')

    async def process_EPS_Growth_Rate(self, stocks):
        tasks = [self.scraper.EPS_Growth_Rate_and_write_to_excel(stock) for stock in stocks]
        results = await asyncio.gather(*tasks)  # 並行處理所有股票的任務
        match_list, file_not_found_list, file_permission_list = [], [], []

        for result in results:
            match1 = re.match(self.pattern1, result)
            file_not_found_match = re.search(self.pattern3, result)
            file_permission_match = re.match(self.pattern4, result)
            if match1:
                match_list.append(match1.group())
            elif file_not_found_match:
                file_not_found_list.append(file_not_found_match.group(1))
            elif file_permission_match:
                file_permission_list.append(file_permission_match.group(1))

        if match_list:
            print(f"{', '.join([re.match(self.pattern1, m).group() for m in match_list])}的EPS成長率成功寫入及儲存成功")
        if file_not_found_list:
            file_path = '\n'.join(file_not_found_list)
            print(f"找不到此檔案，請確認路徑！您目前提供的路徑如下：\n{file_path}'")
        if file_permission_list:
            file_path = '\n'.join(file_permission_list)
            print(f'存檔失敗：請先關閉Excel再執行程式！請先關閉Excel檔案路徑：\n{file_path}')

if __name__ == '__main__':
    start = time.time()

    # 注意事項
    print(
        f'###注意事項###\n股票代碼請勿輸入非美國本土企業！！！\nROIC網站在Financial及Ratios頁面時，查看國外企業時需額外付費才能觀看！！！\n如"ASML、"TSM"等等...非美國本土企業！！！')
    print()
    # Excel檔案路徑
    template_path = r'C:\tmp\stock_template'

    # 股票代碼(可自行任意輸入，只需更改您要查詢的股票代碼即可)
    stocks = ['MSFT', 'DELL', 'NVDA', 'META', 'PLTR', 'AAPL', 'MMM', 'TSLA', 'AMAT', 'GOOG']
    # stocks = ['MSFT', 'DELL', 'NVDA', 'META', 'PLTR']
    # stocks = ['AAPL', 'MMM', 'TSLA', 'AMAT', 'GOOG']
    # stocks = ['MSFT', 'MU', 'MMM', 'GOOGL', 'C', 'BRK-B', 'TSM']
    # stocks = ['MSFT', 'MU', 'MMM', 'TSM', 'ASML']
    # stocks = ['TSM', 'ASML']
    # stocks = ['NVDA', 'META']
    scraper = StockScraper(stocks=stocks, template_path=template_path)

    # 創建物件
    processor = StockProcess(template_path)
    Manager = StockManager(scraper, processor, template_path)

    # 執行資料爬蟲及處理並回傳結果
    print(f'###程式執行訊息回報###\n')
    asyncio.run(Manager.process_summary(stocks))
    print('-'*100)
    asyncio.run(Manager.process_financial(stocks))
    print('-' * 100)
    asyncio.run(Manager.process_ratios(stocks))
    print('-' * 100)
    asyncio.run(Manager.process_EPS_PE_MarketCap(stocks))
    print('-' * 100)
    asyncio.run(Manager.process_others_data(stocks))
    print('-' * 100)
    asyncio.run(Manager.process_EPS_Growth_Rate(stocks))

    # 計時
    print()
    print(f'抓取ROIC及寫入至Excel所花費時間：{time.time() - start:.2f}(單位：秒)')