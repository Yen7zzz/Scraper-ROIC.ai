# 在您的 stock_analyzer_gui.py 檔案最上方添加這段程式碼

import os
import sys


def setup_playwright_path():
    """設定 Playwright 瀏覽器路徑"""

    # 如果是打包後的執行檔
    if hasattr(sys, '_MEIPASS'):
        # PyInstaller 打包後的臨時資料夾
        base_path = sys._MEIPASS

        # 設定瀏覽器路徑到打包後的位置
        browser_path = os.path.join(base_path, 'ms-playwright')

        if os.path.exists(browser_path):
            os.environ['PLAYWRIGHT_BROWSERS_PATH'] = browser_path
            print(f"設定瀏覽器路徑: {browser_path}")
        else:
            # 嘗試原始路徑
            original_path = r'C:\Users\2993\AppData\Local\ms-playwright'
            if os.path.exists(original_path):
                os.environ['PLAYWRIGHT_BROWSERS_PATH'] = original_path
                print(f"使用原始瀏覽器路徑: {original_path}")
    else:
        # 開發環境，使用原始路徑
        original_path = r'C:\Users\2993\AppData\Local\ms-playwright'
        if os.path.exists(original_path):
            os.environ['PLAYWRIGHT_BROWSERS_PATH'] = original_path


# 在導入 playwright 之前呼叫這個函數
setup_playwright_path()

# 然後才導入 playwright
from playwright.sync_api import sync_playwright

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
import asyncio
import sys
import os
from datetime import datetime
import time

# 匯入您原本的模組
import base64
import io
from playwright.async_api import async_playwright
import pandas as pd
import random
from io import StringIO
import re
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from bs4 import BeautifulSoup
import yfinance as yf
import aiohttp
import json
from concurrent.futures import ThreadPoolExecutor

# 您需要將您的 Excel 模板轉換後的 base64 字串放在這裡
EXCEL_TEMPLATE_BASE64 = """
我的模板
"""


# ===== 將您原本的 StockScraper, StockProcess, StockManager 類別複製到這裡 =====
# [這裡應該包含您原本的所有類別，由於篇幅限制我省略了，您需要將它們完整複製過來]

class StockScraper:
    def __init__(self, stocks, headless=True, max_concurrent=3):
        """
        初始化爬蟲類別。
        stocks: 股票代碼的列表
        headless: 是否使用無頭模式
        max_concurrent: 同時執行的股票數量（控制併發數）
        """
        self.stocks = stocks
        self.headless = headless
        self.max_concurrent = max_concurrent
        self.browser = None
        self.playwright = None

    async def setup_browser(self):
        """設定瀏覽器環境。"""
        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.launch(
            headless=self.headless,
            args=[
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",
                "--disable-accelerated-2d-canvas",
                "--disable-gpu",
                "--disable-features=IsolateOrigins,site-per-process",
                "--disable-background-timer-throttling",
            ],
        )

    async def cleanup(self):
        """清理資源。"""
        if self.browser:
            await self.browser.close()
        if self.playwright:
            await self.playwright.stop()

    async def fetch_summary_data(self, stock, semaphore):
        """抓取單一股票的數據（summary）。"""
        async with semaphore:
            try:
                context = await self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
                    viewport={"width": 800, "height": 600},
                )
                try:
                    page_summary = await context.new_page()
                    summary = await asyncio.gather(self.get_summary(stock, page_summary))
                    return {stock: summary}
                finally:
                    await context.close()
            except Exception as e:
                return {"stock": stock, "error": str(e)}

    async def get_summary(self, stock, page, retries=1):
        """抓取特定股票的摘要資料並回傳 DataFrame。"""
        URL = f'https://www.roic.ai/quote/{stock}'
        attempt = 0

        while attempt < retries:
            try:
                await asyncio.sleep(random.uniform(1, 3))
                await page.goto(URL, wait_until='load', timeout=50000)
                await page.wait_for_selector('table.w-full.caption-bottom.text-sm.table-fixed', timeout=100000)
                content = await page.content()
                dfs = pd.read_html(StringIO(content))
                return dfs
            except Exception as e:
                attempt += 1
                if attempt == retries:
                    return f"Error for {stock}: {e}"

        return f"Failed to retrieve data for {stock}"

    async def run_summary(self):
        await self.setup_browser()
        semaphore = asyncio.Semaphore(self.max_concurrent)
        try:
            tasks = [self.fetch_summary_data(stock, semaphore) for stock in self.stocks]
            result = await asyncio.gather(*tasks)
        finally:
            await self.cleanup()
        return result

    async def fetch_financials_data(self, stock, semaphore):
        """抓取單一股票的數據（financials）。"""
        async with semaphore:
            try:
                context = await self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
                    viewport={"width": 800, "height": 600},
                )
                try:
                    page_financials = await context.new_page()
                    financials = await asyncio.gather(self.get_financials(stock, page_financials))
                    return {stock: financials}
                finally:
                    await context.close()
            except Exception as e:
                return {"stock": stock, "error": str(e)}

    async def get_financials(self, stock, page, retries=1):
        """抓取特定股票的財務資料並回傳 DataFrame。"""
        URL = f'https://www.roic.ai/quote/{stock}/financials'
        attempt = 0

        while attempt < retries:
            try:
                await asyncio.sleep(random.uniform(1, 3))
                await page.goto(URL, wait_until='load', timeout=50000)

                if await page.query_selector(
                        'div.rounded-lg.bg-card.text-card-foreground.shadow-sm.mx-auto.flex.w-\\[500px\\].flex-col.items-center.border.drop-shadow-lg'):
                    return f'{stock}是非美國企業，此頁面須付費！'
                else:
                    await page.wait_for_selector('table.w-full.caption-bottom.text-sm.table-fixed', timeout=100000)
                    content = await page.content()
                    dfs = pd.read_html(StringIO(content))
                    return dfs

            except Exception as e:
                attempt += 1
                if attempt == retries:
                    return f"Error for {stock}: {e}"

        return f"Failed to retrieve data for {stock}"

    async def run_financial(self):
        await self.setup_browser()
        semaphore = asyncio.Semaphore(self.max_concurrent)
        try:
            tasks = [self.fetch_financials_data(stock, semaphore) for stock in self.stocks]
            result = await asyncio.gather(*tasks)
        finally:
            await self.cleanup()
        return result

    async def fetch_ratios_data(self, stock, semaphore):
        """抓取單一股票的數據（Ratios）。"""
        async with semaphore:
            try:
                context = await self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
                    viewport={"width": 800, "height": 600},
                )
                try:
                    page_ratios = await context.new_page()
                    ratios = await asyncio.gather(self.get_ratios(stock, page_ratios))
                    return {stock: ratios}
                finally:
                    await context.close()
            except Exception as e:
                return {"stock": stock, "error": str(e)}

    async def get_ratios(self, stock, page, retries=1):
        """抓取特定股票的比率資料並回傳 DataFrame。"""
        URL = f'https://www.roic.ai/quote/{stock}/ratios'
        attempt = 0

        while attempt < retries:
            try:
                await asyncio.sleep(random.uniform(1, 3))
                await page.goto(URL, wait_until='load', timeout=50000)

                if await page.query_selector(
                        'div.rounded-lg.bg-card.text-card-foreground.shadow-sm.mx-auto.flex.w-\\[500px\\].flex-col.items-center.border.drop-shadow-lg'):
                    return f'{stock}是非美國企業，此頁面須付費！'
                else:
                    await page.wait_for_selector('table.w-full.caption-bottom.text-sm.table-fixed', timeout=100000)
                    content = await page.content()
                    dfs = pd.read_html(StringIO(content))
                    return dfs

            except Exception as e:
                attempt += 1
                if attempt == retries:
                    return f"Error for {stock}: {e}"

        return f"Failed to retrieve data for {stock}"

    async def run_ratios(self):
        await self.setup_browser()
        semaphore = asyncio.Semaphore(self.max_concurrent)
        try:
            tasks = [self.fetch_ratios_data(stock, semaphore) for stock in self.stocks]
            result = await asyncio.gather(*tasks)
        finally:
            await self.cleanup()
        return result

    async def fetch_EPS_PE_MarketCap_data(self, stock, semaphore):
        """抓取單一股票的數據（EPS_PE_MarketCap）。"""
        async with semaphore:
            try:
                context = await self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
                    viewport={"width": 800, "height": 600},
                )
                try:
                    page_EPS_PE_MarketCap = await context.new_page()
                    EPS_PE_MarketCap = await asyncio.gather(self.get_EPS_PE_MarketCap(stock, page_EPS_PE_MarketCap))
                    return {stock: EPS_PE_MarketCap}
                finally:
                    await context.close()
            except Exception as e:
                return {"stock": stock, "error": str(e)}

    async def get_EPS_PE_MarketCap(self, stock, page, retries=3):
        url = f'https://www.roic.ai/quote/{stock}'
        attempt = 0
        while attempt < retries:
            try:
                await asyncio.sleep(random.uniform(1, 3))
                await page.goto(url, wait_until='load', timeout=30000)
                await page.wait_for_selector('table', timeout=30000)
                content = await page.content()
                soup = BeautifulSoup(content, 'html.parser')
                span_string = soup.find_all('span', attrs={'stock_class': 'flex text-sm uppercase text-muted-foreground'})
                span_int_value = soup.find_all('span', attrs={'stock_class': 'flex text-lg text-foreground'})

                dic_data = {
                    span_string[0].text: float(span_int_value[0].text),
                    span_string[1].text: float(span_int_value[1].text),
                    span_string[2].text: span_int_value[2].text,
                    span_string[3].text: span_int_value[3].text
                }
                return dic_data
            except Exception as e:
                return f'error message:{e}'

    async def run_EPS_PE_MarketCap(self):
        await self.setup_browser()
        semaphore = asyncio.Semaphore(self.max_concurrent)
        try:
            tasks = [self.fetch_EPS_PE_MarketCap_data(stock, semaphore) for stock in self.stocks]
            result = await asyncio.gather(*tasks)
        finally:
            await self.cleanup()
        return result

    async def EPS_Growth_Rate_and_write_to_excel(self, stock, excel_base64):
        """抓取EPS成長率並寫入Excel"""
        if '-' in stock:
            stock = ''.join(['.' if char == '-' else char for char in stock])

        async with aiohttp.ClientSession() as session:
            async with session.get(f'https://api.stockboss.io/api/symbol?symbol={stock}') as response:
                content = await response.text()
                dic = json.loads(content)
                # print(dic['symbol']['guru_summary']['summary']['summary']['company_data']['wacc'])
                wacc = float(dic['symbol']['guru_summary']['summary']['summary']['company_data']['wacc'])/100
                l_eps_growth5y = []
                try:
                    EPS_Growth_Rate_3_Year = \
                        dic['symbol']['keyratio']['keyratio']['annuals']['3-Year EPS Growth Rate %'][-1]
                    EPS_Growth_Rate_5_Year = \
                        dic['symbol']['keyratio']['keyratio']['annuals']['5-Year EPS Growth Rate %'][-1]
                    EPS_Growth_Rate_10_Year = \
                        dic['symbol']['keyratio']['keyratio']['annuals']['10-Year EPS Growth Rate %'][-1]

                    EPS_Growth_Rate_3_Year = 0 if EPS_Growth_Rate_3_Year == '-' else EPS_Growth_Rate_3_Year
                    EPS_Growth_Rate_5_Year = 0 if EPS_Growth_Rate_5_Year == '-' else EPS_Growth_Rate_5_Year
                    EPS_Growth_Rate_10_Year = 0 if EPS_Growth_Rate_10_Year == '-' else EPS_Growth_Rate_10_Year

                    l_eps_growth5y = l_eps_growth5y + [EPS_Growth_Rate_3_Year, EPS_Growth_Rate_5_Year,
                                                       EPS_Growth_Rate_10_Year]

                except KeyError as e:
                    return f"EPS_Growth_Rate的dictionary錯誤：{stock}", excel_base64

                # 選擇成長率：如果最小值大於 0，則取最小值，否則取最大值
                selected_growth_rate = min(l_eps_growth5y) / 100 if min(l_eps_growth5y) > 0 else max(
                    l_eps_growth5y) / 100
                # print(selected_growth_rate)
                # print(wacc)
                # 寫入 Excel
                try:
                    excel_binary = base64.b64decode(excel_base64)
                    excel_buffer = io.BytesIO(excel_binary)
                    wb = load_workbook(excel_buffer)
                    ws = wb.worksheets[3]  # 假設需要寫入的工作表是第四個
                    ws['C4'] = selected_growth_rate
                    ws['C6'] = wacc

                    output_buffer = io.BytesIO()
                    wb.save(output_buffer)
                    output_buffer.seek(0)
                    modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

                    return f"{stock}的EPS成長率及WACC成功寫入", modified_base64

                except Exception as e:
                    return f"寫入Excel時發生錯誤：{e}", excel_base64


class StockProcess:
    def __init__(self, max_concurrent=2, request_delay=2.0):
        # 將 semaphore 移到類別層級，確保全域限制
        self.semaphore = asyncio.Semaphore(max_concurrent)
        self.request_delay = request_delay  # 請求之間的延遲（秒）
        self.last_request_time = {}  # 記錄每個API的上次請求時間

    def create_excel_from_base64(self, stock):
        """從base64模板創建Excel文件的base64"""
        try:
            if EXCEL_TEMPLATE_BASE64.strip() == "" or "請將您從轉換工具得到的" in EXCEL_TEMPLATE_BASE64:
                return "", "❌ 錯誤：請先設定 EXCEL_TEMPLATE_BASE64 變數"

            excel_binary = base64.b64decode(EXCEL_TEMPLATE_BASE64.strip())
            excel_buffer = io.BytesIO(excel_binary)
            workbook = load_workbook(excel_buffer)

            # 儲存修改後的檔案到記憶體
            output_buffer = io.BytesIO()
            workbook.save(output_buffer)
            output_buffer.seek(0)
            excel_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            return excel_base64, f"成功為 {stock} 創建Excel檔案"

        except Exception as e:
            return "", f"創建Excel檔案時發生錯誤: {e}"

    def write_wacc_to_excel_base64(self, stock_code: str, wacc_value: str, excel_base64: str) -> tuple:
        """將WACC值寫入Excel的base64"""
        try:
            excel_binary = base64.b64decode(excel_base64)
            excel_buffer = io.BytesIO(excel_binary)
            workbook = load_workbook(excel_buffer)

            if "現金流折現法" in workbook.sheetnames:
                worksheet = workbook["現金流折現法"]
            else:
                worksheet = workbook.active

            # 寫入 WACC 值到 C6 儲存格
            if wacc_value != "未知":
                if wacc_value.endswith('%'):
                    numeric_value = float(wacc_value.rstrip('%')) / 100
                    worksheet['C6'] = numeric_value
                    worksheet['C6'].number_format = '0.00%'
                else:
                    worksheet['C6'] = wacc_value
            else:
                worksheet['C6'] = "無法取得"

            # 儲存修改後的檔案到記憶體
            output_buffer = io.BytesIO()
            workbook.save(output_buffer)
            output_buffer.seek(0)
            modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            return modified_base64, f"已將 {stock_code} 的 WACC 值 ({wacc_value}) 寫入 C6 儲存格"

        except Exception as e:
            return excel_base64, f"操作 Excel 檔案時發生錯誤: {e}"

    async def process_df_summary(self, raw_df_summary, stock, excel_base64):
        """處理summary數據並寫入Excel base64"""
        if raw_df_summary:
            try:
                excel_binary = base64.b64decode(excel_base64)
                excel_buffer = io.BytesIO(excel_binary)
                wb = load_workbook(excel_buffer)
                ws = wb.worksheets[0]

                # 清除舊資料
                for row in ws.iter_rows(min_row=1, min_col=1, max_row=30, max_col=12):
                    for cell in row:
                        cell.value = None

                d_1_raw_df_summary = [y for x in raw_df_summary.get(stock, pd.DataFrame({})) for y in x]

                for df_amount, df in enumerate(d_1_raw_df_summary):
                    df_column_list = df.columns.tolist()

                    # 篩選需要的年份
                    years_list = []
                    pattern = r'\d{4}\sY'
                    for years in df_column_list:
                        if re.match(pattern, years):
                            years_list.append(int(years.split()[0]))

                    if years_list:
                        end_year = max(years_list)
                        start_year = end_year - 10
                        drop_column = [
                            x for x in df_column_list
                            if not (re.match(pattern, x) and start_year <= int(x.split()[0]) <= end_year)
                        ]
                        drop_column.pop(0)  # 移除列名的第一欄
                        df = df.drop(columns=drop_column)

                        # 資料轉型為數值型
                        years_data = df.columns[1:]
                        df[years_data] = df[years_data].apply(pd.to_numeric, errors='coerce')

                        # *** 關鍵修改：反轉年份欄位的順序 ***
                        # 保留第一欄（通常是項目名稱），但反轉年份欄位
                        first_col = df.columns[0]  # 第一欄
                        year_cols = df.columns[1:]  # 年份欄位
                        reversed_year_cols = year_cols[::-1]  # 反轉年份欄位順序

                        # 重新組合欄位順序：第一欄 + 反轉的年份欄位
                        new_column_order = [first_col] + list(reversed_year_cols)
                        df_reordered = df[new_column_order]

                        # 將資料寫入 Excel，並設置欄位格式
                        start_row = 1

                        for r_idx, row in enumerate(dataframe_to_rows(df_reordered, index=False, header=True),
                                                    start=start_row):
                            row_data = list(row)

                            # 第一欄（項目名稱）寫入 A 欄
                            cell = ws.cell(row=r_idx, column=1, value=row_data[0])
                            cell.font = Font(size=12, bold=(r_idx == start_row))

                            # 年份數據從右邊開始寫入（從 L 欄開始往左）
                            year_data = row_data[1:]  # 除了第一欄以外的年份數據

                            # L欄是第12欄，從L欄開始往左寫
                            for year_idx, value in enumerate(year_data):
                                column_position = 12 - year_idx  # L=12, K=11, J=10...
                                cell = ws.cell(row=r_idx, column=column_position, value=value)
                                cell.font = Font(size=12, bold=(r_idx == start_row))

                        # 自動調整欄寬
                        for col in ws.columns:
                            max_length = max(len(str(cell.value or '')) for cell in col)
                            ws.column_dimensions[col[0].column_letter].width = max_length + 2

                # 儲存到base64
                output_buffer = io.BytesIO()
                wb.save(output_buffer)
                output_buffer.seek(0)
                modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

                return modified_base64, f"{stock}的Summary成功寫入"

            except Exception as e:
                return excel_base64, f"處理Summary資料時發生錯誤: {e}"
        else:
            return excel_base64, '無原始資料'

    async def process_df_financial(self, raw_df_financial, stock, excel_base64):
        """處理financial數據並寫入Excel base64"""
        if raw_df_financial:
            try:
                excel_binary = base64.b64decode(excel_base64)
                excel_buffer = io.BytesIO(excel_binary)
                wb = load_workbook(excel_buffer)
                ws = wb.worksheets[0]

                # 定義各類財務數據的起始位置
                starting_cell = [("IncomeStatement", 1, 14),  # N1
                                 ("BalanceSheet", 1, 27),  # AA1
                                 ("CashFlowStatement", 1, 40)]  # AN1

                # 清除舊資料
                for row in ws.iter_rows(min_row=1, min_col=14, max_row=100, max_col=25):
                    for cell in row:
                        cell.value = None
                for row in ws.iter_rows(min_row=1, min_col=27, max_row=100, max_col=38):
                    for cell in row:
                        cell.value = None
                for row in ws.iter_rows(min_row=1, min_col=40, max_row=100, max_col=51):
                    for cell in row:
                        cell.value = None

                if raw_df_financial.get(stock) == [f'{stock}是非美國企業，此頁面須付費！']:
                    return excel_base64, f'{stock}是非美國企業，此頁面須付費！'

                else:
                    d_1_raw_df_financial = [y for x in raw_df_financial.get(stock, pd.DataFrame({})) for y in x]

                    for df_amount, df in enumerate(d_1_raw_df_financial):
                        df_column_list = df.columns.tolist()

                        # 篩選需要的年份
                        years_list = []
                        pattern = r'\d{4}\sY'
                        for years in df_column_list:
                            if re.match(pattern, years):
                                years_list.append(int(years.split()[0]))

                        if years_list:
                            end_year = max(years_list)
                            start_year = end_year - 10
                            drop_column = [
                                x for x in df_column_list
                                if not (re.match(pattern, x) and start_year <= int(x.split()[0]) <= end_year)
                            ]
                            drop_column.pop(0)
                            df = df.drop(columns=drop_column)

                            # 資料轉型為數值型
                            years_data = df.columns[1:]
                            df[years_data] = df[years_data].apply(pd.to_numeric, errors='coerce')

                            # *** 關鍵修改：反轉年份欄位的順序 ***
                            first_col = df.columns[0]  # 第一欄
                            year_cols = df.columns[1:]  # 年份欄位
                            reversed_year_cols = year_cols[::-1]  # 反轉年份欄位順序

                            # 重新組合欄位順序：第一欄 + 反轉的年份欄位
                            new_column_order = [first_col] + list(reversed_year_cols)
                            df_reordered = df[new_column_order]

                            # 將表格資料寫入指定位置並調整格式
                            start_row = starting_cell[df_amount][1]
                            start_col = starting_cell[df_amount][2]

                            for r_idx, row in enumerate(dataframe_to_rows(df_reordered, index=False, header=True),
                                                        start=start_row):
                                row_data = list(row)

                                # 第一欄（項目名稱）寫入起始欄位
                                cell = ws.cell(row=r_idx, column=start_col, value=row_data[0])
                                cell.font = Font(size=12, bold=(r_idx == start_row))

                                # 年份數據從右邊開始寫入（從起始欄位+11開始往左）
                                year_data = row_data[1:]  # 除了第一欄以外的年份數據

                                # 從起始欄位+11開始往左寫（假設有12欄的空間）
                                for year_idx, value in enumerate(year_data):
                                    column_position = (start_col + 11) - year_idx  # 從右邊往左寫
                                    cell = ws.cell(row=r_idx, column=column_position, value=value)
                                    cell.font = Font(size=12, bold=(r_idx == start_row))

                            # 自動調整欄寬
                            for col in ws.columns:
                                max_length = max(len(str(cell.value or '')) for cell in col)
                                ws.column_dimensions[col[0].column_letter].width = max_length + 2

                # 儲存到base64
                output_buffer = io.BytesIO()
                wb.save(output_buffer)
                output_buffer.seek(0)
                modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

                return modified_base64, f"{stock}的Financial成功寫入"

            except Exception as e:
                return excel_base64, f"處理Financial資料時發生錯誤: {e}"
        else:
            return excel_base64, '無原始資料'

    async def process_df_ratios(self, raw_df_ratios, stock, excel_base64):
        """處理ratios數據並寫入Excel base64"""
        if raw_df_ratios:
            try:
                excel_binary = base64.b64decode(excel_base64)
                excel_buffer = io.BytesIO(excel_binary)
                wb = load_workbook(excel_buffer)
                ws = wb.worksheets[0]  # 使用第一個工作表

                # 定義各類財務數據的起始位置（對應A.py的7個類別）
                starting_cell = [
                    ('Profitability', 1, 53), ('Credit', 1, 66),
                    ('Liquidity', 1, 79), ('Working Capital', 1, 92),
                    ('Enterprise Value', 1, 105), ('Multiples', 1, 118),
                    ('Per Share Data Items', 1, 131)
                ]

                # 清除舊資料（對應A.py的7個區域）
                for row in ws.iter_rows(min_row=1, min_col=53, max_row=100, max_col=64):
                    for cell in row:
                        cell.value = None
                for row in ws.iter_rows(min_row=1, min_col=66, max_row=100, max_col=77):
                    for cell in row:
                        cell.value = None
                for row in ws.iter_rows(min_row=1, min_col=79, max_row=100, max_col=90):
                    for cell in row:
                        cell.value = None
                for row in ws.iter_rows(min_row=1, min_col=92, max_row=100, max_col=103):
                    for cell in row:
                        cell.value = None
                for row in ws.iter_rows(min_row=1, min_col=105, max_row=100, max_col=116):
                    for cell in row:
                        cell.value = None
                for row in ws.iter_rows(min_row=1, min_col=118, max_row=100, max_col=129):
                    for cell in row:
                        cell.value = None
                for row in ws.iter_rows(min_row=1, min_col=131, max_row=100, max_col=142):
                    for cell in row:
                        cell.value = None

                if raw_df_ratios.get(stock) == [f'{stock}是非美國企業，此頁面須付費！']:
                    return excel_base64, f'{stock}是非美國企業，此頁面須付費！'
                else:
                    d_1_raw_df_ratios = [y for x in raw_df_ratios.get(stock, []) for y in x]

                    for df_amount, df in enumerate(d_1_raw_df_ratios):
                        df_column_list = df.columns.tolist()

                        # 篩選需要的年份
                        years_list = []
                        pattern = r'\d{4}\sY'
                        for years in df_column_list:
                            if re.match(pattern, years):
                                years_list.append(int(years.split()[0]))

                        if years_list:
                            end_year = max(years_list)
                            start_year = end_year - 10
                            drop_column = [
                                x for x in df_column_list
                                if not (re.match(pattern, x) and start_year <= int(x.split()[0]) <= end_year)
                            ]
                            drop_column.pop(0)
                            df = df.drop(columns=drop_column)

                            # 資料轉型為數值型
                            years_data = df.columns[1:]
                            df[years_data] = df[years_data].apply(pd.to_numeric, errors='coerce')

                            # *** 關鍵修改：反轉年份欄位的順序 ***
                            first_col = df.columns[0]  # 第一欄
                            year_cols = df.columns[1:]  # 年份欄位
                            reversed_year_cols = year_cols[::-1]  # 反轉年份欄位順序

                            # 重新組合欄位順序：第一欄 + 反轉的年份欄位
                            new_column_order = [first_col] + list(reversed_year_cols)
                            df_reordered = df[new_column_order]

                            # 將表格資料寫入指定位置並調整格式
                            start_row = starting_cell[df_amount][1]
                            start_col = starting_cell[df_amount][2]

                            for r_idx, row in enumerate(dataframe_to_rows(df_reordered, index=False, header=True),
                                                        start=start_row):
                                row_data = list(row)

                                # 第一欄（項目名稱）寫入起始欄位
                                cell = ws.cell(row=r_idx, column=start_col, value=row_data[0])
                                cell.font = Font(size=12, bold=(r_idx == start_row))

                                # 年份數據從右邊開始寫入（從起始欄位+11開始往左）
                                year_data = row_data[1:]  # 除了第一欄以外的年份數據

                                # 從起始欄位+11開始往左寫（假設有12欄的空間）
                                for year_idx, value in enumerate(year_data):
                                    column_position = (start_col + 11) - year_idx  # 從右邊往左寫
                                    cell = ws.cell(row=r_idx, column=column_position, value=value)
                                    cell.font = Font(size=12, bold=(r_idx == start_row))

                            # 自動調整欄寬
                            for col in ws.columns:
                                max_length = max(len(str(cell.value or '')) for cell in col)
                                ws.column_dimensions[col[0].column_letter].width = max_length + 2

                # 儲存到base64
                output_buffer = io.BytesIO()
                wb.save(output_buffer)
                output_buffer.seek(0)
                modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

                return modified_base64, f"{stock}的Ratios成功寫入"

            except Exception as e:
                return excel_base64, f"處理Ratios資料時發生錯誤: {e}"
        else:
            return excel_base64, '無原始資料'

    async def EPS_PE_MarketCap_data_write_to_excel(self, EPS_PE_MarketCap_content, stock, excel_base64):
        """將 EPS_PE_MarketCap 數據寫入 Excel base64"""
        try:
            excel_binary = base64.b64decode(excel_base64)
            excel_buffer = io.BytesIO(excel_binary)
            wb = load_workbook(excel_buffer)
            ws = wb.worksheets[0]

            # 處理資料
            for data in EPS_PE_MarketCap_content.get(stock, {}):
                start_cell = "EN1"
                start_row = int(start_cell[2:])  # 提取行號，例如 "1"

                for i, (key, value) in enumerate(data.items()):
                    row = start_row + i  # 從起始行開始逐行寫入
                    ws[f"EN{row}"] = key  # 寫入鍵到 EN 列
                    ws[f"EO{row}"] = value  # 寫入值到 EO 列

            # 儲存到 base64
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            return modified_base64, f'{stock}的EPS_PE_MarketCap成功寫入及儲存成功'

        except Exception as e:
            return excel_base64, f"處理 EPS_PE_MarketCap 時發生錯誤: {e}"

    async def _rate_limit(self, api_key="yfinance"):
        """實施速率限制"""
        current_time = time.time()

        if api_key not in self.last_request_time:
            self.last_request_time[api_key] = 0

        time_since_last_request = current_time - self.last_request_time[api_key]

        if time_since_last_request < self.request_delay:
            sleep_time = self.request_delay - time_since_last_request
            # 添加隨機延遲，避免所有請求同時發送
            sleep_time += random.uniform(0.5, 1.5)
            print(f"⏳ 等待 {sleep_time:.1f} 秒以避免API限制...")
            await asyncio.sleep(sleep_time)

        self.last_request_time[api_key] = time.time()

    async def _fetch_stock_data_with_retry(self, stock, max_retries=3):
        """帶重試機制的數據獲取"""
        for attempt in range(max_retries):
            try:
                return await asyncio.to_thread(self._fetch_stock_data, stock)
            except Exception as e:
                if attempt == max_retries - 1:  # 最後一次嘗試
                    raise e

                # 指數退避：每次重試等待時間加倍
                wait_time = (2 ** attempt) * 3 + random.uniform(2, 5)
                print(f"⚠️ 獲取 {stock} 資料失敗，{wait_time:.1f}秒後重試... (嘗試 {attempt + 1}/{max_retries})")
                await asyncio.sleep(wait_time)

    def _fetch_stock_data(self, stock):
        """同步獲取股票數據"""
        # 查詢 10 年期美國國債收益率
        tnx = yf.Ticker("^TNX")
        rf_rate = tnx.info['previousClose'] / 100

        # 獲取股票資料
        Stock = yf.Ticker(stock)
        beta = Stock.info['beta']
        currentPrice = Stock.info['currentPrice']
        symbol = Stock.info['symbol']

        return {
            'Stock': symbol,
            'CurrentPrice': currentPrice,
            'beta': beta,
            'rf_rate': rf_rate
        }

    async def others_data(self, stock, excel_base64):
        """抓取其他數據並寫入Excel base64"""
        async with self.semaphore:  # 限制併發數量
            try:
                # 添加請求延遲，避免頻率過高
                await self._rate_limit("yfinance")

                # 使用重試機制獲取數據
                dic_data = await self._fetch_stock_data_with_retry(stock)

                print(f'{stock}: {dic_data}')

                # 寫入 Excel（移到線程中執行避免阻塞）
                modified_base64 = await self._write_to_excel(excel_base64, dic_data)

                return modified_base64, f'{stock}的其他資料成功寫入'

            except Exception as e:
                return excel_base64, f"獲取 {stock} 資料時發生錯誤：{str(e)}"

    async def _write_to_excel(self, excel_base64, dic_data):
        """寫入Excel文件"""

        def write_excel():
            excel_binary = base64.b64decode(excel_base64)
            excel_buffer = io.BytesIO(excel_binary)
            wb = load_workbook(excel_buffer)

            ws = wb.worksheets[0]  # 選擇第一個工作表
            ws['EQ2'] = dic_data['Stock']
            ws['ER2'] = dic_data['CurrentPrice']

            ws = wb.worksheets[3]
            ws['C31'] = dic_data['beta']
            ws['C36'] = dic_data['rf_rate']

            # 儲存到base64
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            return base64.b64encode(output_buffer.read()).decode('utf-8')

        return await asyncio.to_thread(write_excel)

    def save_excel_to_file(self, base64_data: str, output_path: str) -> bool:
        """將 base64 編碼的 Excel 資料保存為實體檔案"""
        try:
            excel_binary = base64.b64decode(base64_data)
            with open(output_path, 'wb') as f:
                f.write(excel_binary)
            print(f"Excel 檔案已保存至：{output_path}")
            return True
        except Exception as e:
            print(f"保存檔案時發生錯誤: {e}")
            return False


class StockManager:
    def __init__(self, scraper, processor, max_concurrent=3, delay=1):
        self.scraper = scraper
        self.processor = processor
        self.pattern1 = r'^[a-zA-Z\-\.]{1,5}'
        self.pattern2 = r'是非美國企業，此頁面須付費！$'
        self.semaphore = asyncio.Semaphore(max_concurrent)
        self.delay = delay
        self.excel_files = {}  # 儲存每支股票的Excel base64
        self.max_concurrent = max_concurrent

    async def initialize_excel_files(self, stocks):
        """為所有股票初始化Excel檔案"""
        for stock in stocks:
            excel_base64, message = self.processor.create_excel_from_base64(stock)
            if excel_base64:
                self.excel_files[stock] = excel_base64
                print(f"✅ {message}")
            else:
                print(f"❌ {message}")
                return False
        return True

    async def process_summary(self, stocks):
        """處理Summary數據"""
        raw_df_summary = await self.scraper.run_summary()
        for index, stock in enumerate(stocks):
            if stock in self.excel_files:
                modified_base64, message = await self.processor.process_df_summary(
                    raw_df_summary[index], stock, self.excel_files[stock]
                )
                self.excel_files[stock] = modified_base64
                print(f"✅ {message}")

    async def process_financial(self, stocks):
        """處理Financial數據"""
        raw_df_financial = await self.scraper.run_financial()
        for index, stock in enumerate(stocks):
            if stock in self.excel_files:
                modified_base64, message = await self.processor.process_df_financial(
                    raw_df_financial[index], stock, self.excel_files[stock]
                )
                self.excel_files[stock] = modified_base64
                print(f"✅ {message}")

    async def process_ratios(self, stocks):
        """處理Ratios數據"""
        raw_df_ratios = await self.scraper.run_ratios()
        for index, stock in enumerate(stocks):
            if stock in self.excel_files:
                modified_base64, message = await self.processor.process_df_ratios(
                    raw_df_ratios[index], stock, self.excel_files[stock]
                )
                self.excel_files[stock] = modified_base64
                print(f"✅ {message}")

    async def process_EPS_PE_MarketCap(self, stocks):
        """處理EPS/PE/MarketCap數據"""
        raw_df_EPS_PE_MarketCap = await self.scraper.run_EPS_PE_MarketCap()
        for index, stock in enumerate(stocks):
            if stock in self.excel_files:
                modified_base64, message = await self.processor.EPS_PE_MarketCap_data_write_to_excel(
                    raw_df_EPS_PE_MarketCap[index], stock, self.excel_files[stock]
                )
                self.excel_files[stock] = modified_base64
                print(f"✅ {message}")

    async def process_others_data(self, stocks):
        """處理其他數據"""
        for stock in stocks:
            if stock in self.excel_files:
                modified_base64, message = await self.processor.others_data(
                    stock, self.excel_files[stock]
                )
                self.excel_files[stock] = modified_base64
                print(f"✅ {message}")

    async def process_EPS_Growth_Rate(self, stocks):
        """處理EPS成長率"""
        for stock in stocks:
            if stock in self.excel_files:
                message, modified_base64 = await self.scraper.EPS_Growth_Rate_and_write_to_excel(
                    stock, self.excel_files[stock]
                )
                self.excel_files[stock] = modified_base64
                print(f"✅ {message}")

    def save_all_excel_files(self, stocks, output_folder=None):
        """保存所有Excel檔案"""
        if output_folder is None:
            output_folder = os.getcwd()

        saved_files = []
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        for stock in stocks:
            if stock in self.excel_files:
                output_filename = f"STOCK_{stock}.xlsx"
                output_path = os.path.join(output_folder, output_filename)

                if self.processor.save_excel_to_file(self.excel_files[stock], output_path):
                    saved_files.append(output_path)
                    print(f"✅ {stock} 檔案已保存至：{output_path}")
                else:
                    print(f"❌ {stock} 檔案保存失敗")

        return saved_files


# ===== 新增：股票代碼驗證器類別 =====
class StockValidator:
    """股票代碼驗證器"""

    def __init__(self):
        self.valid_stocks = []
        self.invalid_stocks = []

    def validate_single_stock(self, stock):
        """驗證單一股票代碼"""
        try:
            # 使用 yfinance 獲取歷史數據來驗證股票
            ticker = yf.Ticker(stock)

            # 先嘗試獲取基本歷史數據，避免使用 .info 屬性
            hist = ticker.history(period="5d")

            if not hist.empty:
                return True, f"✅ {stock}: 有效股票代碼"
            else:
                return False, f"❌ {stock}: 無法獲得股價資訊"

        except Exception as e:
            return False, f"❌ {stock}: 驗證失敗 - {str(e)}"

    async def validate_stocks_async(self, stocks, log_callback=None):
        """異步驗證多個股票代碼"""
        self.valid_stocks = []
        self.invalid_stocks = []

        if log_callback:
            log_callback("🔍 開始驗證股票代碼...")

        # 使用線程池執行同步的股票驗證
        with ThreadPoolExecutor(max_workers=5) as executor:
            # 創建任務列表
            tasks = []
            for stock in stocks:
                task = asyncio.get_event_loop().run_in_executor(
                    executor, self.validate_single_stock, stock
                )
                tasks.append((stock, task))

            # 等待所有驗證完成
            for stock, task in tasks:
                try:
                    is_valid, message = await task

                    if log_callback:
                        log_callback(message)

                    if is_valid:
                        self.valid_stocks.append(stock)
                    else:
                        self.invalid_stocks.append(stock)

                    # 添加小延遲避免API限制
                    await asyncio.sleep(0.5)

                except Exception as e:
                    error_msg = f"❌ {stock}: 驗證過程發生錯誤 - {str(e)}"
                    if log_callback:
                        log_callback(error_msg)
                    self.invalid_stocks.append(stock)

        if log_callback:
            log_callback(f"🎯 驗證完成！有效股票: {len(self.valid_stocks)}，無效股票: {len(self.invalid_stocks)}")

        return self.valid_stocks, self.invalid_stocks