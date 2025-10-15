# åœ¨æ‚¨çš„ stock_analyzer_gui.py æª”æ¡ˆæœ€ä¸Šæ–¹æ·»åŠ é€™æ®µç¨‹å¼ç¢¼

import os
import sys


def setup_playwright_path():
    """è¨­å®š Playwright ç€è¦½å™¨è·¯å¾‘"""

    # å¦‚æœæ˜¯æ‰“åŒ…å¾Œçš„åŸ·è¡Œæª”
    if hasattr(sys, '_MEIPASS'):
        # PyInstaller æ‰“åŒ…å¾Œçš„è‡¨æ™‚è³‡æ–™å¤¾
        base_path = sys._MEIPASS

        # è¨­å®šç€è¦½å™¨è·¯å¾‘åˆ°æ‰“åŒ…å¾Œçš„ä½ç½®
        browser_path = os.path.join(base_path, 'ms-playwright')

        if os.path.exists(browser_path):
            os.environ['PLAYWRIGHT_BROWSERS_PATH'] = browser_path
            print(f"è¨­å®šç€è¦½å™¨è·¯å¾‘: {browser_path}")
        else:
            # å˜—è©¦åŸå§‹è·¯å¾‘
            original_path = r'C:\Users\2993\AppData\Local\ms-playwright'
            if os.path.exists(original_path):
                os.environ['PLAYWRIGHT_BROWSERS_PATH'] = original_path
                print(f"ä½¿ç”¨åŸå§‹ç€è¦½å™¨è·¯å¾‘: {original_path}")
    else:
        # é–‹ç™¼ç’°å¢ƒï¼Œä½¿ç”¨åŸå§‹è·¯å¾‘
        original_path = r'C:\Users\2993\AppData\Local\ms-playwright'
        if os.path.exists(original_path):
            os.environ['PLAYWRIGHT_BROWSERS_PATH'] = original_path


# åœ¨å°å…¥ playwright ä¹‹å‰å‘¼å«é€™å€‹å‡½æ•¸
setup_playwright_path()

# ç„¶å¾Œæ‰å°å…¥ playwright
from playwright.sync_api import sync_playwright

import asyncio
import base64
import io
from playwright.async_api import async_playwright
import pandas as pd
import random
from io import StringIO
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import aiohttp
import json
import re
import yfinance as yf

class StockScraper:
    def __init__(self, stocks, config=None, headless=True, max_concurrent=5):
        """
        åˆå§‹åŒ–çˆ¬èŸ²é¡åˆ¥ã€‚
        stocks: è‚¡ç¥¨ä»£ç¢¼çš„åˆ—è¡¨
        headless: æ˜¯å¦ä½¿ç”¨ç„¡é ­æ¨¡å¼
        max_concurrent: åŒæ™‚åŸ·è¡Œçš„è‚¡ç¥¨æ•¸é‡ï¼ˆæ§åˆ¶ä½µç™¼æ•¸ï¼‰
        """
        self.stocks = stocks.get('final_stocks')
        self.us_stocks = stocks.get('us_stocks')
        self.non_us_stocks = stocks.get('non_us_stocks')
        self.config = config
        self.headless = headless
        self.max_concurrent = max_concurrent
        self.browser = None
        self.playwright = None
        # é©—è­‰ Schwab API é…ç½®
        self._validate_schwab_config()

    def _validate_schwab_config(self):
        """é©—è­‰ Schwab API é…ç½®æ˜¯å¦å®Œæ•´"""
        if self.config is None:
            print("âš ï¸ è­¦å‘Šï¼šæœªæä¾› Schwab API é…ç½®")
            print("é¸æ“‡æ¬ŠéˆåŠŸèƒ½å°‡ç„¡æ³•ä½¿ç”¨")
            self.schwab_available = False
            return

        required_keys = ['app_key', 'app_secret']
        missing_keys = [key for key in required_keys if not self.config.get(key)]

        if missing_keys:
            print(f"âš ï¸ è­¦å‘Šï¼šSchwab API é…ç½®ä¸å®Œæ•´ï¼Œç¼ºå°‘ï¼š{', '.join(missing_keys)}")
            print("é¸æ“‡æ¬ŠéˆåŠŸèƒ½å°‡ç„¡æ³•ä½¿ç”¨")
            self.schwab_available = False
        else:
            print("âœ“ Schwab API é…ç½®å·²è¼‰å…¥")
            self.schwab_available = True

    async def setup_browser(self):
        """è¨­å®šç€è¦½å™¨ç’°å¢ƒã€‚"""
        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.launch(
            headless=self.headless,
            args=[
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",
                "--disable-accelerated-2d-canvas",
                "--disable-gpu",
                "--disable-features=IsolateOrigins,site-per-process",
                "--disable-background-timer-throttling",
                "--disable-blink-features=AutomationControlled",  # æ–°å¢ï¼šéš±è—è‡ªå‹•åŒ–æ¨™è¨˜
                "--exclude-switches=enable-automation",  # æ–°å¢ï¼šç§»é™¤automationé–‹é—œ
            ],
        )

    async def cleanup(self):
        """æ¸…ç†è³‡æºã€‚"""
        if self.browser:
            await self.browser.close()
        if self.playwright:
            await self.playwright.stop()

    async def fetch_financials_data(self, stock, semaphore):
        """æŠ“å–å–®ä¸€è‚¡ç¥¨çš„æ•¸æ“šï¼ˆfinancialsï¼‰ã€‚"""
        async with semaphore:
            try:
                context = await self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
                    viewport={"width": 800, "height": 600},
                    java_script_enabled=True
                )
                try:
                    page_financials = await context.new_page()
                    financials = await asyncio.gather(self.get_financials(stock, page_financials))
                    return {stock: financials}
                finally:
                    await context.close()
            except Exception as e:
                return {"stock": stock, "error": str(e)}

    async def get_financials(self, stock, page, retries=3):
        """æŠ“å–ç‰¹å®šè‚¡ç¥¨çš„è²¡å‹™è³‡æ–™ä¸¦å›å‚³ DataFrameã€‚"""
        URL = f'https://www.roic.ai/quote/{stock}/financials'
        attempt = 0

        while attempt < retries:
            try:
                await asyncio.sleep(random.uniform(1, 3))
                await page.goto(URL, wait_until='networkidle', timeout=100000) # networkidle

                # 2025/09/23 æ›´æ–°æ–°é‚è¼¯

                await page.wait_for_selector('table.w-full.caption-bottom.text-sm.table-fixed', timeout=100000)
                content = await page.content()
                dfs = pd.read_html(StringIO(content))
                return dfs

                # ä¹‹å‰çš„é‚è¼¯
                # if await page.query_selector(
                #         'div.rounded-lg.bg-card.text-card-foreground.shadow-sm.mx-auto.flex.w-\\[500px\\].flex-col.items-center.border.drop-shadow-lg'):
                #     return f'{stock}æ˜¯éç¾åœ‹ä¼æ¥­ï¼Œæ­¤é é¢é ˆä»˜è²»ï¼'
                # else:
                #     await page.wait_for_selector('table.w-full.caption-bottom.text-sm.table-fixed', timeout=100000)
                #     content = await page.content()
                #     dfs = pd.read_html(StringIO(content))
                #     return dfs

            except Exception as e:
                attempt += 1
                if attempt == retries:
                    return f"Error for {stock}: {e}"

        return f"Failed to retrieve data for {stock}"

    async def run_financial(self):
        await self.setup_browser()
        semaphore = asyncio.Semaphore(self.max_concurrent)
        try:
            tasks = [self.fetch_financials_data(stock, semaphore) for stock in self.us_stocks]
            result = await asyncio.gather(*tasks)
        finally:
            await self.cleanup()
        return result

    async def fetch_ratios_data(self, stock, semaphore):
        """æŠ“å–å–®ä¸€è‚¡ç¥¨çš„æ•¸æ“šï¼ˆRatiosï¼‰ã€‚"""
        async with semaphore:
            try:
                context = await self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
                    viewport={"width": 800, "height": 600},
                    java_script_enabled=True
                )
                try:
                    page_ratios = await context.new_page()
                    ratios = await asyncio.gather(self.get_ratios(stock, page_ratios))
                    # print({stock: ratios})
                    return {stock: ratios}
                finally:
                    await context.close()
            except Exception as e:
                return {"stock": stock, "error": str(e)}

    async def get_ratios(self, stock, page, retries=3):
        """æŠ“å–ç‰¹å®šè‚¡ç¥¨çš„æ¯”ç‡è³‡æ–™ä¸¦å›å‚³ DataFrameã€‚"""
        URL = f'https://www.roic.ai/quote/{stock}/ratios'
        attempt = 0

        while attempt < retries:
            try:
                await asyncio.sleep(random.uniform(1, 3))
                await page.goto(URL, wait_until='load', timeout=50000)

                # 2025/09/23 æ›´æ–°æ–°é‚è¼¯
                await page.wait_for_selector('table.w-full.caption-bottom.text-sm.table-fixed', timeout=100000)
                content = await page.content()
                dfs = pd.read_html(StringIO(content))
                return dfs

                # ä¹‹å‰çš„é‚è¼¯
                # if await page.query_selector(
                #         'div.rounded-lg.bg-card.text-card-foreground.shadow-sm.mx-auto.flex.w-\\[500px\\].flex-col.items-center.border.drop-shadow-lg'):
                #     return f'{stock}æ˜¯éç¾åœ‹ä¼æ¥­ï¼Œæ­¤é é¢é ˆä»˜è²»ï¼'
                # else:
                #     await page.wait_for_selector('table.w-full.caption-bottom.text-sm.table-fixed', timeout=100000)
                #     content = await page.content()
                #     dfs = pd.read_html(StringIO(content))
                #     return dfs

            except Exception as e:
                attempt += 1
                if attempt == retries:
                    return f"Error for {stock}: {e}"

        return f"Failed to retrieve data for {stock}"

    async def run_ratios(self):
        await self.setup_browser()
        semaphore = asyncio.Semaphore(self.max_concurrent)
        try:
            tasks = [self.fetch_ratios_data(stock, semaphore) for stock in self.us_stocks]
            result = await asyncio.gather(*tasks)
        finally:
            await self.cleanup()
        return result

    # async def fetch_EPS_PE_MarketCap_data(self, stock, semaphore):
    #     """æŠ“å–å–®ä¸€è‚¡ç¥¨çš„æ•¸æ“šï¼ˆEPS_PE_MarketCapï¼‰ã€‚"""
    #     async with semaphore:
    #         try:
    #             context = await self.browser.new_context(
    #                 user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
    #                 viewport={"width": 800, "height": 600},
    #             )
    #             try:
    #                 page_EPS_PE_MarketCap = await context.new_page()
    #                 EPS_PE_MarketCap = await asyncio.gather(self.get_EPS_PE_MarketCap(stock, page_EPS_PE_MarketCap))
    #                 return {stock: EPS_PE_MarketCap}
    #             finally:
    #                 await context.close()
    #         except Exception as e:
    #             return {"stock": stock, "error": str(e)}

    async def get_EPS_PE_MarketCap(self, stock, page, retries=3):
        """æŠ“å–ç‰¹å®šè‚¡ç¥¨çš„EPS/PE/MarketCapæ•¸æ“š - æ›´æ–°ç‰ˆæœ¬é©æ‡‰æ–°çš„HTMLçµæ§‹"""
        url = f'https://www.roic.ai/quote/{stock}'
        attempt = 0

        while attempt < retries:
            try:
                await asyncio.sleep(random.uniform(1, 3))
                await page.goto(url, wait_until='load', timeout=30000)

                # ç­‰å¾…é—œéµæŒ‡æ¨™å®¹å™¨è¼‰å…¥
                await page.wait_for_selector('div[data-cy="company_header_ratios"]', timeout=30000)

                content = await page.content()
                soup = BeautifulSoup(content, 'html.parser')

                # æ–¹æ³•1ï¼šä½¿ç”¨æ–°çš„data-cyå±¬æ€§å®šä½
                ratios_container = soup.find('div', {'data-cy': 'company_header_ratios'})

                if ratios_container:
                    print(f"æ‰¾åˆ° {stock} çš„æŒ‡æ¨™å®¹å™¨")

                    # æå–æ‰€æœ‰æŒ‡æ¨™é …ç›®
                    metric_items = ratios_container.find_all('div', class_='shrink-0 flex-col')

                    if len(metric_items) >= 3:  # è‡³å°‘éœ€è¦EPS, P/E, Market Cap
                        dic_data = {}

                        for item in metric_items:
                            # æå–æ•¸å€¼ï¼ˆå¤§å­—ï¼‰
                            value_span = item.find('span', class_='flex text-lg text-foreground')
                            # æå–æ¨™ç±¤ï¼ˆå°å­—ï¼‰
                            label_span = item.find('span', class_='flex text-sm uppercase text-muted-foreground')

                            if value_span and label_span:
                                label = label_span.get_text(strip=True)
                                value_text = value_span.get_text(strip=True)

                                # æ ¹æ“šæ¨™ç±¤é¡å‹é€²è¡Œä¸åŒè™•ç†
                                if label in ['EPS', 'P/E']:
                                    try:
                                        dic_data[label] = float(value_text)
                                    except ValueError:
                                        dic_data[label] = value_text  # å¦‚æœç„¡æ³•è½‰æ›ç‚ºæ•¸å­—ï¼Œä¿æŒåŸå­—ä¸²
                                else:
                                    dic_data[label] = value_text  # Market Cap, Next Earnç­‰ä¿æŒå­—ä¸²

                        print(f"æˆåŠŸæå– {stock} çš„æŒ‡æ¨™æ•¸æ“š: {dic_data}")
                        return dic_data

                    else:
                        print(f"æŒ‡æ¨™é …ç›®æ•¸é‡ä¸è¶³: æ‰¾åˆ° {len(metric_items)} å€‹é …ç›®")

                # æ–¹æ³•2ï¼šå‚™ç”¨æ–¹æ¡ˆ - ä½¿ç”¨é¡åˆ¥é¸æ“‡å™¨
                if not ratios_container:
                    print(f"å˜—è©¦å‚™ç”¨æ–¹æ¡ˆæŠ“å– {stock} çš„æŒ‡æ¨™...")

                    # ç›´æ¥å°‹æ‰¾æ‰€æœ‰ç¬¦åˆæ–°çµæ§‹çš„spanå…ƒç´ 
                    value_spans = soup.find_all('span', class_='flex text-lg text-foreground')
                    label_spans = soup.find_all('span', class_='flex text-sm uppercase text-muted-foreground')

                    if len(value_spans) >= 3 and len(label_spans) >= 3:
                        dic_data = {}

                        # å‡è¨­å‰å¹¾å€‹å°±æ˜¯æˆ‘å€‘è¦çš„æŒ‡æ¨™
                        for i in range(min(len(value_spans), len(label_spans))):
                            label = label_spans[i].get_text(strip=True)
                            value_text = value_spans[i].get_text(strip=True)

                            # åªè™•ç†æˆ‘å€‘é—œå¿ƒçš„æŒ‡æ¨™
                            if label in ['EPS', 'P/E', 'MARKET CAP', 'Market Cap', 'NEXT EARN', 'Next Earn']:
                                if label in ['EPS', 'P/E']:
                                    try:
                                        dic_data[label] = float(value_text)
                                    except ValueError:
                                        dic_data[label] = value_text
                                else:
                                    dic_data[label] = value_text

                        if dic_data:
                            print(f"å‚™ç”¨æ–¹æ¡ˆæˆåŠŸæå– {stock} çš„æŒ‡æ¨™æ•¸æ“š: {dic_data}")
                            return dic_data

                # å¦‚æœæ‰€æœ‰æ–¹æ³•éƒ½å¤±æ•—
                return {'error': f'ç„¡æ³•æ‰¾åˆ° {stock} çš„æŒ‡æ¨™æ•¸æ“š'}

            except Exception as e:
                attempt += 1
                print(f"ç¬¬ {attempt} æ¬¡å˜—è©¦å¤±æ•—: {e}")
                if attempt < retries:
                    await asyncio.sleep(random.uniform(2, 5))
                else:
                    return {'error': f'æŠ“å– {stock} æ•¸æ“šæ™‚ç™¼ç”ŸéŒ¯èª¤: {e}'}

        return {'error': f'Failed to retrieve data for {stock}'}

    async def fetch_combined_summary_and_metrics_data(self, stock, semaphore):
        """åŒæ™‚æŠ“å–Summaryè¡¨æ ¼æ•¸æ“šå’ŒEPS/PE/MarketCapæŒ‡æ¨™æ•¸æ“š"""
        async with semaphore:
            try:
                context = await self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
                    viewport={"width": 800, "height": 600},
                )
                try:
                    page = await context.new_page()

                    # ä¸€æ¬¡æ€§ç²å–å…©ç¨®æ•¸æ“š
                    summary_data, metrics_data = await self.get_combined_data(stock, page)

                    return {
                        stock: {
                            'summary': summary_data,
                            'metrics': metrics_data
                        }
                    }
                finally:
                    await context.close()
            except Exception as e:
                return {"stock": stock, "error": str(e)}

    async def get_combined_data(self, stock, page, retries=3):
        """å¾åŒä¸€é é¢åŒæ™‚ç²å–Summaryè¡¨æ ¼å’ŒæŒ‡æ¨™æ•¸æ“š - æ›´æ–°ç‰ˆæœ¬"""
        URL = f'https://www.roic.ai/quote/{stock}'
        attempt = 0

        while attempt < retries:
            try:
                await asyncio.sleep(random.uniform(1, 3))
                await page.goto(URL, wait_until='load', timeout=50000)

                # ç­‰å¾…å…©ç¨®é—œéµå…ƒç´ è¼‰å…¥å®Œæˆ
                await page.wait_for_selector('table.w-full.caption-bottom.text-sm.table-fixed', timeout=100000)
                await page.wait_for_selector('div[data-cy="company_header_ratios"]', timeout=30000)

                # ç²å–é é¢å…§å®¹
                content = await page.content()

                # 1. ä½¿ç”¨ pandas è§£æè¡¨æ ¼æ•¸æ“šï¼ˆSummaryéƒ¨åˆ†ï¼‰
                summary_data = None
                try:
                    dfs = pd.read_html(StringIO(content))
                    summary_data = dfs
                    print(f"æˆåŠŸè§£æ {stock} çš„è¡¨æ ¼æ•¸æ“šï¼Œå…± {len(dfs)} å€‹è¡¨æ ¼")
                except Exception as e:
                    print(f"è§£æ {stock} è¡¨æ ¼æ•¸æ“šå¤±æ•—: {e}")
                    summary_data = []

                # 2. ä½¿ç”¨ BeautifulSoup è§£ææŒ‡æ¨™æ•¸æ“šï¼ˆæ›´æ–°ç‰ˆæœ¬ï¼‰
                metrics_data = None
                try:
                    soup = BeautifulSoup(content, 'html.parser')

                    # ä½¿ç”¨æ–°çš„é¸æ“‡å™¨
                    ratios_container = soup.find('div', {'data-cy': 'company_header_ratios'})

                    if ratios_container:
                        metric_items = ratios_container.find_all('div', class_='shrink-0 flex-col')

                        if len(metric_items) >= 3:
                            metrics_data = {}

                            for item in metric_items:
                                value_span = item.find('span', class_='flex text-lg text-foreground')
                                label_span = item.find('span', class_='flex text-sm uppercase text-muted-foreground')

                                if value_span and label_span:
                                    label = label_span.get_text(strip=True)
                                    value_text = value_span.get_text(strip=True)

                                    if label in ['EPS', 'P/E']:
                                        try:
                                            metrics_data[label] = float(value_text)
                                        except ValueError:
                                            metrics_data[label] = value_text
                                    else:
                                        metrics_data[label] = value_text

                            print(f"æˆåŠŸè§£æ {stock} çš„æŒ‡æ¨™æ•¸æ“š: {metrics_data}")
                        else:
                            metrics_data = {}
                    else:
                        print(f"æœªæ‰¾åˆ° {stock} çš„æŒ‡æ¨™å®¹å™¨")
                        metrics_data = {}

                except Exception as e:
                    print(f"è§£æ {stock} æŒ‡æ¨™æ•¸æ“šå¤±æ•—: {e}")
                    metrics_data = {}

                return summary_data, metrics_data

            except Exception as e:
                attempt += 1
                print(f"ç¬¬ {attempt} æ¬¡å˜—è©¦å¤±æ•—: {e}")
                if attempt == retries:
                    return [], {}
                await asyncio.sleep(random.uniform(2, 5))

        return [], {}

    async def run_combined_summary_and_metrics(self):
        """åŸ·è¡Œåˆä½µçš„Summaryå’ŒæŒ‡æ¨™æ•¸æ“šæŠ“å–"""
        await self.setup_browser()
        semaphore = asyncio.Semaphore(self.max_concurrent)
        try:
            tasks = [self.fetch_combined_summary_and_metrics_data(stock, semaphore) for stock in self.stocks]
            result = await asyncio.gather(*tasks)

            # åˆ†é›¢çµæœä»¥ä¿æŒèˆ‡ç¾æœ‰ä»£ç¢¼çš„å…¼å®¹æ€§
            summary_results = []
            metrics_results = []

            for item in result:
                for stock, data in item.items():
                    if stock != "stock" and "error" not in item:  # æ’é™¤éŒ¯èª¤é …ç›®
                        summary_results.append({stock: data['summary']})
                        metrics_results.append({stock: data['metrics']})
                    else:
                        # è™•ç†éŒ¯èª¤æƒ…æ³
                        summary_results.append(item)
                        metrics_results.append(item)

            return summary_results, metrics_results

        finally:
            await self.cleanup()


    # async def EPS_Growth_Rate_and_write_to_excel(self, stock, excel_base64):
    #     """æŠ“å–EPSæˆé•·ç‡ä¸¦å¯«å…¥Excel"""
    #     if '-' in stock:
    #         stock = ''.join(['.' if char == '-' else char for char in stock])
    #
    #     async with aiohttp.ClientSession() as session:
    #         async with session.get(f'https://api.stockboss.io/api/symbol?symbol={stock}') as response:
    #             content = await response.text()
    #             dic = json.loads(content)
    #             # print(dic['symbol']['guru_summary']['summary']['summary']['company_data']['wacc'])
    #             # wacc = float(dic['symbol']['guru_summary']['summary']['summary']['company_data']['wacc'])/100
    #             l_eps_growth5y = []
    #             try:
    #                 EPS_Growth_Rate_3_Year = \
    #                     dic['symbol']['keyratio']['keyratio']['annuals']['3-Year EPS Growth Rate %'][-1]
    #                 EPS_Growth_Rate_5_Year = \
    #                     dic['symbol']['keyratio']['keyratio']['annuals']['5-Year EPS Growth Rate %'][-1]
    #                 EPS_Growth_Rate_10_Year = \
    #                     dic['symbol']['keyratio']['keyratio']['annuals']['10-Year EPS Growth Rate %'][-1]
    #
    #                 EPS_Growth_Rate_3_Year = 0 if EPS_Growth_Rate_3_Year == '-' else EPS_Growth_Rate_3_Year
    #                 EPS_Growth_Rate_5_Year = 0 if EPS_Growth_Rate_5_Year == '-' else EPS_Growth_Rate_5_Year
    #                 EPS_Growth_Rate_10_Year = 0 if EPS_Growth_Rate_10_Year == '-' else EPS_Growth_Rate_10_Year
    #
    #                 l_eps_growth5y = l_eps_growth5y + [EPS_Growth_Rate_3_Year, EPS_Growth_Rate_5_Year,
    #                                                    EPS_Growth_Rate_10_Year]
    #
    #             except KeyError as e:
    #                 return f"EPS_Growth_Rateçš„dictionaryéŒ¯èª¤ï¼š{stock}", excel_base64
    #
    #             # é¸æ“‡æˆé•·ç‡ï¼šå¦‚æœæœ€å°å€¼å¤§æ–¼ 0ï¼Œå‰‡å–æœ€å°å€¼ï¼Œå¦å‰‡å–æœ€å¤§å€¼
    #             selected_growth_rate = min(l_eps_growth5y) / 100 if min(l_eps_growth5y) > 0 else max(
    #                 l_eps_growth5y) / 100
    #             # print(selected_growth_rate)
    #             # print(wacc)
    #             # å¯«å…¥ Excel
    #             try:
    #                 excel_binary = base64.b64decode(excel_base64)
    #                 excel_buffer = io.BytesIO(excel_binary)
    #                 wb = load_workbook(excel_buffer)
    #                 ws = wb.worksheets[3]  # å‡è¨­éœ€è¦å¯«å…¥çš„å·¥ä½œè¡¨æ˜¯ç¬¬å››å€‹
    #
    #                 ws['C3'] = None
    #                 ws['C3'] = selected_growth_rate
    #                 # ws['C6'] = wacc
    #
    #                 output_buffer = io.BytesIO()
    #                 wb.save(output_buffer)
    #                 output_buffer.seek(0)
    #                 modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')
    #
    #                 return f"{stock}çš„EPSæˆé•·ç‡åŠWACCæˆåŠŸå¯«å…¥", modified_base64
    #
    #             except Exception as e:
    #                 return f"å¯«å…¥Excelæ™‚ç™¼ç”ŸéŒ¯èª¤ï¼š{e}", excel_base64

    # async def fetch_seekingalpha_data(self, stock, semaphore):
    #     async with semaphore:
    #         try:
    #             context = await self.browser.new_context(
    #                 user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
    #                 viewport={"width": 800, "height": 600},  # å¢åŠ è¦–çª—å¤§å°
    #                 java_script_enabled=True,  # ç¢ºä¿JavaScriptå•Ÿç”¨
    #             )
    #             try:
    #                 page_summary = await context.new_page()
    #                 summary = await self.get_seekingalpha_html(stock, page_summary)
    #                 return {stock: summary}
    #             finally:
    #                 await context.close()
    #         except Exception as e:
    #             return {"stock": stock, "error": str(e)}

    async def get_seekingalpha_html(self, stock, page, retries=3):
        """æŠ“å–ç‰¹å®šè‚¡ç¥¨çš„æ‘˜è¦è³‡æ–™ä¸¦å›å‚³ DataFrameã€‚"""
        if '-' in stock:
            stock = ''.join(['.' if char == '-' else char for char in stock])

        URL = f'https://seekingalpha.com/symbol/{stock}/growth'
        attempt = 0
        # print(URL)
        while attempt < retries:
            try:
                print(f"æ­£åœ¨å˜—è©¦æŠ“å– {stock} çš„è³‡æ–™ (ç¬¬ {attempt + 1} æ¬¡)...")

                # éš¨æ©Ÿç­‰å¾…æ™‚é–“
                await asyncio.sleep(random.uniform(2, 5))

                # å‰å¾€é é¢ - æ”¹ç”¨ domcontentloaded
                await page.goto(URL, wait_until='domcontentloaded', timeout=60000)

                # # === æ¥µç°¡ç‰ˆæ¨¡æ“¬äººé¡è¡Œç‚º ===
                # # æœ€çŸ­åœé “
                # await asyncio.sleep(random.uniform(0.1, 0.4))
                #
                # # åªç§»å‹•ä¸€æ¬¡æ»‘é¼ 
                # await page.mouse.move(
                #     random.randint(300, 500),
                #     random.randint(200, 300)
                # )
                #
                # # æ¥µçŸ­ç­‰å¾…
                # await asyncio.sleep(random.uniform(0.1, 0.3))
                # # === æ¥µç°¡ç‰ˆçµæŸ ===

                # ä½¿ç”¨æ›´ç²¾ç¢ºçš„é¸æ“‡å™¨çµ„åˆ
                try:
                    # å…ˆç­‰å¾…ç‰¹å®šçš„ Growth Rates section
                    await page.wait_for_selector('section[data-test-id="card-container-growth-rates"]', timeout=15000)

                    # å†ç­‰å¾…è©² section å…§çš„è¡¨æ ¼
                    await page.wait_for_selector(
                        'section[data-test-id="card-container-growth-rates"] table[data-test-id="table"]',
                        timeout=10000)

                    # ç­‰å¾… Revenue è¡Œå‡ºç¾ï¼ˆç¢ºä¿å…§å®¹å·²è¼‰å…¥ï¼‰
                    await page.wait_for_selector(
                        'section[data-test-id="card-container-growth-rates"] th:has-text("Revenue")', timeout=10000)

                    # çŸ­æš«ç­‰å¾…ç¢ºä¿æ•¸æ“šæ¸²æŸ“å®Œæˆ
                    await asyncio.sleep(2)

                except Exception as e:
                    print(f"ç­‰å¾…é—œéµå…ƒç´ æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
                    raise e

                # ç²å–é é¢å…§å®¹
                content = await page.content()

                # ä½¿ç”¨BeautifulSoupè§£æç›®æ¨™è¡¨æ ¼
                soup = BeautifulSoup(content, 'html.parser')

                # å…ˆæ‰¾åˆ° Growth Rates sectionï¼Œå†åœ¨å…¶å…§æ‰¾è¡¨æ ¼
                growth_section = soup.find('section', {'data-test-id': 'card-container-growth-rates'})

                if not growth_section:
                    print("æœªæ‰¾åˆ° Growth Rates section")
                    raise Exception("æœªæ‰¾åˆ° Growth Rates section")

                target_table = growth_section.find('table', {'data-test-id': 'table'})

                if target_table:
                    print("æ‰¾åˆ°æ­£ç¢ºçš„ Growth Rates è¡¨æ ¼ï¼Œé–‹å§‹è§£æ...")

                    # è§£æè¡¨é ­
                    header_row = target_table.find('thead').find('tr') if target_table.find('thead') else None
                    headers = []
                    if header_row:
                        header_cells = header_row.find_all('th')
                        for cell in header_cells:
                            div_text = cell.find('div')
                            if div_text:
                                header_text = div_text.get_text(strip=True)
                            else:
                                header_text = cell.get_text(strip=True)
                            headers.append(header_text)

                    print(f"è¡¨é ­: {headers}")

                    # é©—è­‰è¡¨é ­çµæ§‹æ˜¯å¦æ­£ç¢º
                    expected_headers = ['YoY', '3Y', '5Y', '10Y']
                    if not all(h in headers for h in expected_headers):
                        print(f"è¡¨é ­çµæ§‹ä¸ç¬¦åˆé æœŸï¼ŒæœŸæœ›åŒ…å«: {expected_headers}")
                        raise Exception("è¡¨é ­çµæ§‹ä¸æ­£ç¢º")

                    # æ‰¾åˆ° 5Y å’Œ 10Y åœ¨è¡¨é ­ä¸­çš„ä½ç½®
                    try:
                        header_5y_index = headers.index('5Y')
                        header_10y_index = headers.index('10Y')
                        print(f"5Yä½ç½®: {header_5y_index}, 10Yä½ç½®: {header_10y_index}")
                    except ValueError as e:
                        print(f"æ‰¾ä¸åˆ°5Yæˆ–10Yè¡¨é ­: {e}")
                        raise Exception("æ‰¾ä¸åˆ°5Yæˆ–10Yè¡¨é ­")

                    # è§£æè¡¨æ ¼å…§å®¹
                    tbody = target_table.find('tbody')
                    if tbody:
                        rows = tbody.find_all('tr')

                        # åªè™•ç†Revenueè¡Œä¸¦è¿”å›5Yå’Œ10Yæ•¸æ“š
                        for i, row in enumerate(rows):
                            row_data = []

                            # è™•ç†ç¬¬ä¸€å€‹thï¼ˆè¡Œæ¨™é¡Œï¼‰
                            th = row.find('th')
                            if th:
                                div_text = th.find('div')
                                if div_text:
                                    row_name = div_text.get_text(strip=True)
                                else:
                                    row_name = th.get_text(strip=True)
                                row_data.append(row_name)

                            # è™•ç†å…¶ä»–td
                            tds = row.find_all('td')
                            for td in tds:
                                div_text = td.find('div')
                                if div_text:
                                    cell_value = div_text.get_text(strip=True)
                                else:
                                    cell_value = td.get_text(strip=True)
                                row_data.append(cell_value)

                            # æª¢æŸ¥æ˜¯å¦ç‚ºRevenueè¡Œ
                            if 'Revenue' in row_data[0] and 'Revenue per Share' not in row_data[0]:
                                print(f"æ‰¾åˆ°Revenueè¡Œ: {row_data}")

                                # æ ¹æ“šè¡¨é ­ä½ç½®ç²¾ç¢ºæå–5Yå’Œ10Yæ•¸æ“š
                                if len(row_data) > max(header_5y_index, header_10y_index):
                                    result = {
                                        "5Y": row_data[header_5y_index],  # ç›´æ¥ç”¨è¡¨é ­ä½ç½®
                                        "10Y": row_data[header_10y_index]  # ç›´æ¥ç”¨è¡¨é ­ä½ç½®
                                    }
                                    print(f"æå–çµæœ: {result}")
                                    return result
                                else:
                                    return {"error": f"Revenueè¡Œæ•¸æ“šä¸è¶³: {row_data}"}

                        return {"error": "æœªæ‰¾åˆ°Revenueè¡Œ"}
                    else:
                        print("æœªæ‰¾åˆ°tbody")
                        return {"error": "æœªæ‰¾åˆ°tbody"}

                else:
                    print("æœªæ‰¾åˆ°Growth Ratesè¡¨æ ¼")
                    return {"error": "æœªæ‰¾åˆ°Growth Ratesè¡¨æ ¼"}

            except Exception as e:
                print(f"ç¬¬ {attempt + 1} æ¬¡å˜—è©¦å¤±æ•—: {e}")
                attempt += 1
                if attempt < retries:
                    await asyncio.sleep(random.uniform(10, 20))

        return {"error": f"Failed to retrieve data for {stock} after {retries} attempts"}

    async def run_seekingalpha(self):
        await self.setup_browser()

        # å‰µå»ºä¸€å€‹æŒä¹…çš„context
        context = await self.browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
            viewport={"width": 800, "height": 600},
            java_script_enabled=True,
        )

        try:
            page = await context.new_page()
            result = []

            # ä¾åºè™•ç†æ¯å€‹è‚¡ç¥¨
            for stock in self.stocks:
                print(f"æ­£åœ¨è™•ç† {stock}...")
                stock_data = await self.get_seekingalpha_html(stock, page)
                result.append({stock: stock_data})

                # æ¯å€‹è‚¡ç¥¨ä¹‹é–“çš„å»¶é²
                await asyncio.sleep(random.uniform(10, 20))

        finally:
            await context.close()
            await self.cleanup()

        return result

    async def fetch_wacc_data(self, stock, semaphore):
        async with semaphore:
            try:
                context = await self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
                    viewport={"width": 1920, "height": 1080},
                    java_script_enabled=True,
                )
                try:
                    page_summary = await context.new_page()
                    wacc_value = await self.get_wacc_html(stock, page_summary)
                    return {stock: wacc_value}
                finally:
                    await context.close()
            except Exception as e:
                return {stock: None}  # å¦‚æœå‡ºéŒ¯è¿”å›None

    async def get_wacc_html(self, stock, page, retries=3):
        """æŠ“å–ç‰¹å®šè‚¡ç¥¨çš„WACCè³‡æ–™ä¸¦å›å‚³intæ•¸å€¼ã€‚"""
        if '-' in stock:
            stock = ''.join(['.' if char == '-' else char for char in stock])

        URL = f'https://www.gurufocus.com/term/wacc/{stock}'
        attempt = 0
        # print(URL)
        while attempt < retries:
            try:
                print(f"æ­£åœ¨å˜—è©¦æŠ“å– {stock} çš„WACCè³‡æ–™ (ç¬¬ {attempt + 1} æ¬¡)...")

                # éš¨æ©Ÿç­‰å¾…æ™‚é–“
                await asyncio.sleep(random.uniform(2, 5))

                # å‰å¾€é é¢
                await page.goto(URL, wait_until='networkidle', timeout=60000)

                # ç­‰å¾…é é¢è¼‰å…¥å®Œæˆ
                await page.wait_for_load_state('networkidle')

                # ç­‰å¾…é—œéµå…§å®¹è¼‰å…¥
                try:
                    await page.wait_for_selector('h1', timeout=30000)
                    await asyncio.sleep(3)
                except Exception as e:
                    print(f"ç­‰å¾…é é¢è¼‰å…¥æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")

                # ç²å–é é¢å…§å®¹
                content = await page.content()

                # ä½¿ç”¨BeautifulSoupè§£æWACCæ•¸å€¼
                soup = BeautifulSoup(content, 'html.parser')

                # å°‹æ‰¾åŒ…å«WACCæ•¸å€¼çš„ç‰¹å®šå…ƒç´ 
                wacc_value = None

                # æ–¹æ³•1: å°‹æ‰¾åŒ…å«":X.X% (As of"æ¨¡å¼çš„fontæ¨™ç±¤
                font_elements = soup.find_all('font', style=True)
                for font in font_elements:
                    text = font.get_text(strip=True)
                    if '% (As of' in text and text.startswith(':'):
                        # æå–ç™¾åˆ†æ¯”æ•¸å€¼
                        match = re.search(r':(\d+\.?\d*)%', text)
                        if match:
                            wacc_value = float(match.group(1))/100
                            # print(f"æ‰¾åˆ°WACCå€¼: {wacc_value}%")
                            break

                if wacc_value is not None:
                    return wacc_value
                else:
                    print(f"æœªèƒ½æ‰¾åˆ° {stock} çš„WACCæ•¸å€¼")
                    return None

            except Exception as e:
                print(f"ç¬¬ {attempt + 1} æ¬¡å˜—è©¦å¤±æ•—: {e}")
                attempt += 1
                if attempt < retries:
                    await asyncio.sleep(random.uniform(5, 10))

        print(f"Failed to retrieve WACC data for {stock} after {retries} attempts")
        return None

    async def run_wacc(self):
        await self.setup_browser()
        semaphore = asyncio.Semaphore(self.max_concurrent)
        try:
            tasks = [self.fetch_wacc_data(stock, semaphore) for stock in self.stocks]
            result = await asyncio.gather(*tasks)
        finally:
            await self.cleanup()
        return result


    async def fetch_TradingView_data(self, stock, semaphore):
        async with semaphore:
            try:
                context = await self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
                    viewport={"width": 1920, "height": 1080},
                    java_script_enabled=True,
                )
                try:
                    page_summary = await context.new_page()
                    wacc_value = await self.get_TradingView_html(stock, page_summary)
                    return {stock: wacc_value}
                finally:
                    await context.close()
            except Exception as e:
                return {stock: None}  # å¦‚æœå‡ºéŒ¯è¿”å›None

    async def get_TradingView_html(self, stock, page, retries=3):
        """æŠ“å–ç‰¹å®šè‚¡ç¥¨çš„trading-viewè³‡æ–™ä¸¦è™•ç†ç¶²å€è­‰åˆ¸äº¤æ˜“æ‰€å•é¡Œã€‚"""
        url_stock_exchange = yf.Ticker(stock).info.get('fullExchangeName', None)
        if url_stock_exchange in ['NasdaqGS', 'NasdaqGM', 'NasdaqCM']:
            url_stock_exchange = 'NASDAQ'

        if '-' in stock:
            stock = ''.join(['.' if char == '-' else char for char in stock])

        URL = f'https://www.tradingview.com/symbols/{url_stock_exchange}-{stock}/financials-earnings/?earnings-period=FY&revenues-period=FY'
        attempt = 0
        # print(URL)
        while attempt < retries:
            try:
                print(f"æ­£åœ¨å˜—è©¦æŠ“å– {stock} çš„trading-viewè³‡æ–™ (ç¬¬ {attempt + 1} æ¬¡)...")

                # éš¨æ©Ÿç­‰å¾…æ™‚é–“
                await asyncio.sleep(random.uniform(2, 5))

                # å‰å¾€é é¢
                await page.goto(URL, wait_until='networkidle', timeout=60000)

                # ç­‰å¾…é é¢è¼‰å…¥å®Œæˆ
                await page.wait_for_load_state('networkidle')

                # ç­‰å¾…é—œéµå…§å®¹è¼‰å…¥
                try:
                    await page.wait_for_selector('h1', timeout=30000)
                    await asyncio.sleep(3)
                except Exception as e:
                    print(f"ç­‰å¾…é é¢è¼‰å…¥æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")

                # ç²å–é é¢å…§å®¹
                content = await page.content()

                # ä½¿ç”¨BeautifulSoupè§£ætrading-viewæ•¸å€¼
                soup = BeautifulSoup(content, 'html.parser')

                # è§£æå¹´ä»½ - æ‰¾åˆ°æ¨™é¡Œè¡Œä¸­çš„å¹´ä»½
                years = []
                year_elements = soup.find_all('div', class_='value-OxVAcLqi')
                for element in year_elements:
                    text = element.get_text(strip=True)
                    if text.isdigit() and len(text) == 4:  # å¹´ä»½æ˜¯4ä½æ•¸å­—
                        years.append(int(text))

                # å¦‚æœæ²’æ‰¾åˆ°å¹´ä»½ï¼Œå˜—è©¦å¦ä¸€ç¨®æ–¹å¼
                if not years:
                    # æŸ¥æ‰¾åŒ…å«å¹´ä»½çš„å®¹å™¨
                    values_container = soup.find('div', class_='values-AtxjAQkN')
                    if values_container:
                        year_divs = values_container.find_all('div', class_='value-OxVAcLqi')
                        for div in year_divs:
                            text = div.get_text(strip=True)
                            if text.isdigit() and len(text) == 4:
                                years.append(int(text))

                if not years:
                    print(f"ç„¡æ³•æ‰¾åˆ°å¹´ä»½è³‡æ–™å°æ–¼ {stock}")
                    return None

                # åˆå§‹åŒ–è³‡æ–™å­—å…¸
                data = {
                    'Year': years,
                    'Reported': [None] * len(years),
                    'Estimate': [None] * len(years),
                    'Surprise': [None] * len(years)
                }

                # è§£æä¸‰ç¨®é¡å‹çš„è³‡æ–™
                data_types = ['Reported', 'Estimate', 'Surprise']

                for data_type in data_types:
                    # æ‰¾åˆ°å°æ‡‰çš„è³‡æ–™å®¹å™¨
                    container = soup.find('div', {'data-name': data_type})
                    if not container:
                        print(f"æ‰¾ä¸åˆ° {data_type} è³‡æ–™å®¹å™¨")
                        continue

                    # æ‰¾åˆ°è©²å®¹å™¨ä¸­çš„æ•¸å€¼å€åŸŸ
                    values_section = container.find('div', class_='values-C9MdAMrq')
                    if not values_section:
                        print(f"æ‰¾ä¸åˆ° {data_type} çš„æ•¸å€¼å€åŸŸ")
                        continue

                    # ç²å–æ‰€æœ‰æ•¸å€¼å®¹å™¨
                    value_containers = values_section.find_all('div', class_='container-OxVAcLqi')

                    for i, value_container in enumerate(value_containers):
                        if i >= len(years):  # é˜²æ­¢ç´¢å¼•è¶…å‡ºç¯„åœ
                            break

                        # æª¢æŸ¥æ˜¯å¦ç‚ºé–å®šè³‡æ–™ï¼ˆè·³éä»˜è²»å…§å®¹ï¼‰
                        lock_button = value_container.find('button', class_='lockButton-N_j3rnsK')
                        if lock_button:
                            continue  # è·³éé–å®šçš„è³‡æ–™

                        # æå–æ•¸å€¼
                        value_div = value_container.find('div', class_='value-OxVAcLqi')
                        if value_div:
                            value = value_div.get_text(strip=True)
                            # è™•ç†ç‰¹æ®Šç¬¦è™Ÿ
                            if value == 'â€”' or value == '-':
                                value = None
                            elif value.startswith('â€ª') and value.endswith('â€¬'):
                                # ç§»é™¤Unicodeæ§åˆ¶å­—ç¬¦
                                value = value.strip('â€ªâ€¬')

                            # å„²å­˜æ•¸å€¼
                            data[data_type][i] = value

                # å»ºç«‹DataFrameï¼ˆåŸå§‹æ ¼å¼ï¼‰
                df_original = pd.DataFrame(data)

                # åªä¿ç•™æœ‰è³‡æ–™çš„è¡Œï¼ˆè‡³å°‘æœ‰ä¸€å€‹éNoneå€¼ï¼‰
                mask = df_original[['Reported', 'Estimate', 'Surprise']].notna().any(axis=1)
                df_filtered = df_original[mask].reset_index(drop=True)

                # è½‰æ›æˆæ©«å‘æ ¼å¼
                if len(df_filtered) > 0:
                    # å‰µå»ºæ–°çš„DataFrameï¼Œå¹´ä»½ä½œç‚ºåˆ—æ¨™é¡Œ
                    years_list = df_filtered['Year'].tolist()

                    # å‰µå»ºæ©«å‘æ ¼å¼çš„æ•¸æ“š
                    transposed_data = {
                        'Year': years_list,
                        'Reported': df_filtered['Reported'].tolist(),
                        'Estimate': df_filtered['Estimate'].tolist(),
                        'Surprise': df_filtered['Surprise'].tolist()
                    }

                    # è½‰ç½®æ•¸æ“šï¼šå¹´ä»½ä½œç‚ºåˆ—æ¨™é¡Œï¼ŒæŒ‡æ¨™ä½œç‚ºè¡Œ
                    result_dict = {}

                    # ç¬¬ä¸€è¡Œï¼šå¹´ä»½
                    for i, year in enumerate(years_list):
                        result_dict[str(year)] = [
                            transposed_data['Reported'][i],
                            transposed_data['Estimate'][i],
                            transposed_data['Surprise'][i]
                        ]

                    # å‰µå»ºæœ€çµ‚çš„DataFrameï¼Œä»¥æŒ‡æ¨™åç¨±ä½œç‚ºç´¢å¼•
                    df_final = pd.DataFrame(result_dict, index=['Reported', 'Estimate', 'Surprise'])

                    print(f"æˆåŠŸè§£æ {stock} çš„è³‡æ–™ï¼Œæ ¼å¼ç‚º {df_final.shape[1]} å¹´ä»½ x {df_final.shape[0]} æŒ‡æ¨™")
                    return df_final
                else:
                    print(f"æœªæ‰¾åˆ° {stock} çš„æœ‰æ•ˆè³‡æ–™")
                    return None

            except Exception as e:
                print(f"ç¬¬ {attempt + 1} æ¬¡å˜—è©¦å¤±æ•—: {e}")
                attempt += 1
                if attempt < retries:
                    await asyncio.sleep(random.uniform(5, 10))

        print(f"Failed to retrieve TradingView data for {stock} after {retries} attempts")
        return None

    async def run_TradingView(self):
        await self.setup_browser()
        semaphore = asyncio.Semaphore(self.max_concurrent)
        try:
            tasks = [self.fetch_TradingView_data(stock, semaphore) for stock in self.stocks]
            result = await asyncio.gather(*tasks)
        finally:
            await self.cleanup()
        return result

    async def fetch_barchart_data(self, stock, semaphore):
        """æŠ“å–å–®ä¸€è‚¡ç¥¨çš„æ•¸æ“šï¼ˆBarchart Volatilityï¼‰"""
        async with semaphore:
            try:
                context = await self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
                    viewport={"width": 1920, "height": 1080},
                    java_script_enabled=True,
                )
                try:
                    page = await context.new_page()
                    html_content = await self.get_barchart_html(stock, page)
                    return {stock: html_content}
                finally:
                    await context.close()
            except Exception as e:
                return {stock: {"error": str(e)}}

    async def get_barchart_html(self, stock, page, retries=3):
        """æŠ“å–ç‰¹å®šè‚¡ç¥¨çš„Barcharté é¢ä¸¦å›å‚³å®Œæ•´HTML"""
        URL = f'https://www.barchart.com/stocks/quotes/{stock}/volatility-charts'
        attempt = 0

        while attempt < retries:
            try:
                print(f"æ­£åœ¨å˜—è©¦æŠ“å– {stock} çš„Barcharté é¢ (ç¬¬ {attempt + 1} æ¬¡)...")

                await asyncio.sleep(random.uniform(2, 5))
                await page.goto(URL, wait_until='domcontentloaded', timeout=60000)

                # ç­‰å¾…é é¢è¼‰å…¥
                await asyncio.sleep(3)

                # ç²å–å®Œæ•´HTMLå…§å®¹
                content = await page.content()

                # print(f"âœ“ æˆåŠŸç²å– {stock} çš„HTMLï¼Œé•·åº¦: {len(content)}")
                bs = BeautifulSoup(content, 'html.parser')

                div = bs.find('div', {'class':'bc-datatable-toolbar bc-options-toolbar volatility'})
                # print(div)
                return div.text.replace('\xa0', ' ')
                # return content

            except Exception as e:
                print(f"ç¬¬ {attempt + 1} æ¬¡å˜—è©¦å¤±æ•—: {e}")
                attempt += 1
                if attempt < retries:
                    await asyncio.sleep(random.uniform(5, 10))

        return None

    async def run_barchart(self):
        """åŸ·è¡ŒBarchartæ•¸æ“šæŠ“å–"""
        await self.setup_browser()
        semaphore = asyncio.Semaphore(self.max_concurrent)
        try:
            tasks = [self.fetch_barchart_data(stock, semaphore) for stock in self.stocks]
            result = await asyncio.gather(*tasks)
        finally:
            await self.cleanup()
        return result

    async def fetch_option_chain_data(self, stock, semaphore):
        """æŠ“å–å–®ä¸€è‚¡ç¥¨çš„é¸æ“‡æ¬Šéˆæ•¸æ“š"""
        async with semaphore:
            try:
                # æª¢æŸ¥ Schwab API æ˜¯å¦å¯ç”¨
                if not self.schwab_available:
                    return {stock: {"error": "Schwab API é…ç½®æœªå®Œæ•´è¨­å®š"}}

                # ä½¿ç”¨ schwabdev å®¢æˆ¶ç«¯
                option_data = await asyncio.to_thread(
                    self._get_option_chain_sync, stock
                )
                return {stock: option_data}
            except Exception as e:
                return {stock: {"error": str(e)}}

    def _get_option_chain_sync(self, stock):
        """åŒæ­¥ç²å–é¸æ“‡æ¬Šéˆæ•¸æ“š - ä½¿ç”¨å‚³å…¥çš„é…ç½®"""
        import schwabdev
        import os
        import sys

        # æª¢æŸ¥é…ç½®æ˜¯å¦å¯ç”¨
        if not self.schwab_available or not self.config:
            raise ValueError(
                "Schwab API é…ç½®æœªè¨­å®šæˆ–ä¸å®Œæ•´ã€‚\n"
                "è«‹ç¢ºèªå·²å®Œæˆ OAuth èªè­‰æµç¨‹ã€‚"
            )

        # å¾é…ç½®ä¸­è®€å–æ†‘è­‰
        app_key = self.config.get('app_key')
        app_secret = self.config.get('app_secret')
        callback_url = self.config.get('callback_url', 'https://127.0.0.1')

        # é©—è­‰å¿…è¦åƒæ•¸
        if not app_key or not app_secret:
            raise ValueError(
                "Schwab API æ†‘è­‰ç¼ºå¤±ã€‚\n"
                f"app_key: {'å·²è¨­å®š' if app_key else 'âŒ æœªè¨­å®š'}\n"
                f"app_secret: {'å·²è¨­å®š' if app_secret else 'âŒ æœªè¨­å®š'}"
            )

        # ğŸ”¥ é—œéµä¿®æ”¹ï¼šè¨ˆç®— tokens.json çš„å®Œæ•´è·¯å¾‘åˆ° schwab/ è³‡æ–™å¤¾
        if getattr(sys, 'frozen', False):
            # æ‰“åŒ…å¾Œçš„åŸ·è¡Œæª”
            base_path = os.path.dirname(sys.executable)
            tokens_folder = os.path.join(base_path, 'schwab')
        else:
            # é–‹ç™¼ç’°å¢ƒ
            # StockScraper.py ä½æ–¼: pythonProject1/stock_class/StockScraper.py
            current_file = os.path.abspath(__file__)
            project_root = os.path.dirname(os.path.dirname(current_file))  # pythonProject1/
            tokens_folder = os.path.join(project_root, 'schwab')  # pythonProject1/schwab/

        tokens_file_path = os.path.join(tokens_folder, 'tokens.json')

        print(f"ğŸ” ä½¿ç”¨ Schwab API ç²å– {stock} çš„é¸æ“‡æ¬Šæ•¸æ“š...")
        print(f"ğŸ“ Token ä½ç½®: {tokens_file_path}")

        # æª¢æŸ¥ tokens.json æ˜¯å¦å­˜åœ¨
        if not os.path.exists(tokens_file_path):
            raise FileNotFoundError(
                f"æ‰¾ä¸åˆ° Token æª”æ¡ˆ: {tokens_file_path}\n"
                "è«‹å…ˆå®Œæˆ OAuth èªè­‰æµç¨‹ã€‚"
            )

        # å‰µå»ºå®¢æˆ¶ç«¯
        client = schwabdev.Client(
            app_key,
            app_secret,
            callback_url,
            tokens_file=tokens_file_path  # ğŸ‘ˆ ä½¿ç”¨å®Œæ•´è·¯å¾‘
        )

        # ç²å–é¸æ“‡æ¬Šæ•¸æ“š
        response = client.option_chains(stock)
        return response.json()

    async def run_option_chains(self):
        """æ‰¹æ¬¡åŸ·è¡Œé¸æ“‡æ¬ŠéˆæŠ“å–"""
        await self.setup_browser()  # å¦‚æœéœ€è¦çš„è©±
        semaphore = asyncio.Semaphore(self.max_concurrent)
        try:
            tasks = [
                self.fetch_option_chain_data(stock, semaphore)
                for stock in self.stocks
            ]
            result = await asyncio.gather(*tasks)
        finally:
            pass  # é¸æ“‡æ¬ŠAPIä¸éœ€è¦æ¸…ç†ç€è¦½å™¨
        return result