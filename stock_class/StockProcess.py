import asyncio
import time
import base64
import io
import pandas as pd
import random
import re
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
import yfinance as yf
from excel_template.fundamental_excel_template import Fundamental_Excel_Template_Base64
# from excel_template.option_chain_excel_template import Option_Chain_Excel_Template_Base64
from stock_class.RareLimitManager import RateLimitManager
import os
import tempfile
import xlwings as xw
class StockProcess:
    def __init__(self, max_concurrent=2, request_delay=2.0):
        # 將 semaphore 移到類別層級，確保全域限制
        self.semaphore = asyncio.Semaphore(max_concurrent)
        self.request_delay = request_delay  # 請求之間的延遲（秒）
        self.last_request_time = {}  # 記錄每個API的上次請求時間

    def create_excel_from_base64(self, stock):
        """從base64模板創建Excel文件的base64"""
        try:
            if Fundamental_Excel_Template_Base64.strip() == "" or "請將您從轉換工具得到的" in Fundamental_Excel_Template_Base64:
                return "", "❌ 錯誤：請先設定 Fundamental_Excel_Template_Base64 變數"

            excel_binary = base64.b64decode(Fundamental_Excel_Template_Base64.strip())
            excel_buffer = io.BytesIO(excel_binary)
            workbook = load_workbook(excel_buffer)

            # 儲存修改後的檔案到記憶體
            output_buffer = io.BytesIO()
            workbook.save(output_buffer)
            output_buffer.seek(0)
            excel_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            return excel_base64, f"成功為 {stock} 創建Excel檔案"

        except Exception as e:
            return "", f"創建Excel檔案時發生錯誤: {e}"

    async def process_df_summary(self, raw_df_summary, stock, excel_base64):
        """處理summary數據並寫入Excel base64 - 修復版本"""
        try:
            excel_binary = base64.b64decode(excel_base64)
            excel_buffer = io.BytesIO(excel_binary)
            wb = load_workbook(excel_buffer)
            ws = wb.worksheets[0]

            # 清除舊資料=
            for row in ws.iter_rows(min_row=1, min_col=1, max_row=30, max_col=12):
                for cell in row:
                    cell.value = None

            # 【關鍵修復】立即保存清除後的版本
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            cleaned_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            # 檢查是否有原始數據
            if not raw_df_summary:
                return cleaned_base64, 'Summary: 無原始資料，已清空舊數據'

            # 處理數據結構
            if isinstance(raw_df_summary, list):
                d_1_raw_df_summary = raw_df_summary
            elif isinstance(raw_df_summary, dict) and stock in raw_df_summary:
                stock_data = raw_df_summary[stock]
                if isinstance(stock_data, list):
                    d_1_raw_df_summary = stock_data
                else:
                    d_1_raw_df_summary = [stock_data] if stock_data is not None else []
            else:
                d_1_raw_df_summary = [raw_df_summary] if raw_df_summary is not None else []

            for df_amount, df in enumerate(d_1_raw_df_summary):
                df_column_list = df.columns.tolist()

                # 篩選需要的年份
                years_list = []
                pattern = r'\d{4}\sY'
                for years in df_column_list:
                    if re.match(pattern, years):
                        years_list.append(int(years.split()[0]))

                if years_list:
                    end_year = max(years_list)
                    start_year = end_year - 10
                    drop_column = [
                        x for x in df_column_list
                        if not (re.match(pattern, x) and start_year <= int(x.split()[0]) <= end_year)
                    ]
                    drop_column.pop(0)
                    df = df.drop(columns=drop_column)

                    # 資料轉型為數值型
                    years_data = df.columns[1:]
                    df[years_data] = df[years_data].apply(pd.to_numeric, errors='coerce')

                    # 反轉年份欄位的順序
                    first_col = df.columns[0]
                    year_cols = df.columns[1:]
                    reversed_year_cols = year_cols[::-1]

                    new_column_order = [first_col] + list(reversed_year_cols)
                    df_reordered = df[new_column_order]

                    # 將資料寫入 Excel
                    start_row = 1

                    for r_idx, row in enumerate(dataframe_to_rows(df_reordered, index=False, header=True),
                                                start=start_row):
                        row_data = list(row)

                        # 第一欄寫入 A 欄
                        cell = ws.cell(row=r_idx, column=1, value=row_data[0])
                        cell.font = Font(name='新細明體', size=12, bold=(r_idx == start_row))

                        # 年份數據從右邊開始寫入
                        year_data = row_data[1:]

                        for year_idx, value in enumerate(year_data):
                            column_position = 12 - year_idx
                            cell = ws.cell(row=r_idx, column=column_position, value=value)
                            cell.font = Font(name='新細明體', size=12, bold=(r_idx == start_row))

                    # 自動調整欄寬
                    for col in ws.columns:
                        max_length = max(len(str(cell.value or '')) for cell in col)
                        ws.column_dimensions[col[0].column_letter].width = max_length + 2

            # 儲存到base64
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            return modified_base64, f"{stock}的Summary成功寫入"

        except Exception as e:
            # 【關鍵修復】返回清除後的版本
            try:
                excel_binary = base64.b64decode(excel_base64)
                excel_buffer = io.BytesIO(excel_binary)
                wb = load_workbook(excel_buffer)
                ws = wb.worksheets[0]

                # 清除舊資料
                for row in ws.iter_rows(min_row=1, min_col=1, max_row=30, max_col=12):
                    for cell in row:
                        cell.value = None

                # 保存清除後的版本
                output_buffer = io.BytesIO()
                wb.save(output_buffer)
                output_buffer.seek(0)
                cleaned_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

                return cleaned_base64, f"處理Summary資料時發生錯誤，已清空相關區域: {e}"
            except:
                return excel_base64, f"處理Summary資料時發生嚴重錯誤: {e}"

    async def process_df_financial(self, raw_df_financial, stock, excel_base64):
        """處理financial數據並寫入Excel base64 - 修復版本"""
        try:
            excel_binary = base64.b64decode(excel_base64)
            excel_buffer = io.BytesIO(excel_binary)
            wb = load_workbook(excel_buffer)
            ws = wb.worksheets[0]

            # 定義各類財務數據的起始位置
            starting_cell = [("IncomeStatement", 1, 14),  # N1
                             ("BalanceSheet", 1, 27),  # AA1
                             ("CashFlowStatement", 1, 40)]  # AN1

            # 清除舊資料
            for row in ws.iter_rows(min_row=1, min_col=14, max_row=100, max_col=25):
                for cell in row:
                    cell.value = None
            for row in ws.iter_rows(min_row=1, min_col=27, max_row=100, max_col=38):
                for cell in row:
                    cell.value = None
            for row in ws.iter_rows(min_row=1, min_col=40, max_row=100, max_col=51):
                for cell in row:
                    cell.value = None

            # 【關鍵修復】立即保存清除後的版本
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            cleaned_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            # 檢查是否為非美國企業錯誤訊息
            # if raw_df_financial.get(stock) == [f'{stock}是非美國企業，此頁面須付費！']:
            #     return cleaned_base64, f'Financial: {stock}是非美國企業，此頁面須付費！'
            # print(raw_df_financial)
            if raw_df_financial is None:
                return cleaned_base64, f'Financial: {stock}是非美國企業，此頁面須付費！'

            # 開始處理數據
            d_1_raw_df_financial = [y for x in raw_df_financial.get(stock, pd.DataFrame({})) for y in x]

            for df_amount, df in enumerate(d_1_raw_df_financial):
                df_column_list = df.columns.tolist()

                # 篩選需要的年份
                years_list = []
                pattern = r'\d{4}\sY'
                for years in df_column_list:
                    if re.match(pattern, years):
                        years_list.append(int(years.split()[0]))

                if years_list:
                    end_year = max(years_list)
                    start_year = end_year - 10
                    drop_column = [
                        x for x in df_column_list
                        if not (re.match(pattern, x) and start_year <= int(x.split()[0]) <= end_year)
                    ]
                    drop_column.pop(0)
                    df = df.drop(columns=drop_column)

                    # 資料轉型為數值型
                    years_data = df.columns[1:]
                    df[years_data] = df[years_data].apply(pd.to_numeric, errors='coerce')

                    # 反轉年份欄位的順序
                    first_col = df.columns[0]
                    year_cols = df.columns[1:]
                    reversed_year_cols = year_cols[::-1]

                    new_column_order = [first_col] + list(reversed_year_cols)
                    df_reordered = df[new_column_order]

                    # 將表格資料寫入指定位置並調整格式
                    start_row = starting_cell[df_amount][1]
                    start_col = starting_cell[df_amount][2]

                    for r_idx, row in enumerate(dataframe_to_rows(df_reordered, index=False, header=True),
                                                start=start_row):
                        row_data = list(row)

                        # 第一欄（項目名稱）寫入起始欄位
                        cell = ws.cell(row=r_idx, column=start_col, value=row_data[0])
                        cell.font = Font(size=12, bold=(r_idx == start_row))

                        # 年份數據從右邊開始寫入
                        year_data = row_data[1:]

                        for year_idx, value in enumerate(year_data):
                            column_position = (start_col + 11) - year_idx
                            cell = ws.cell(row=r_idx, column=column_position, value=value)
                            cell.font = Font(size=12, bold=(r_idx == start_row))

                    # 自動調整欄寬
                    for col in ws.columns:
                        max_length = max(len(str(cell.value or '')) for cell in col)
                        ws.column_dimensions[col[0].column_letter].width = max_length + 2

            # 保存到base64
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            return modified_base64, f"{stock}的Financial成功寫入"

        except Exception as e:
            # 【關鍵修復】返回清除後的版本而非原始版本
            try:
                excel_binary = base64.b64decode(excel_base64)
                excel_buffer = io.BytesIO(excel_binary)
                wb = load_workbook(excel_buffer)
                ws = wb.worksheets[0]

                # 清除舊資料
                for row in ws.iter_rows(min_row=1, min_col=14, max_row=100, max_col=25):
                    for cell in row:
                        cell.value = None
                for row in ws.iter_rows(min_row=1, min_col=27, max_row=100, max_col=38):
                    for cell in row:
                        cell.value = None
                for row in ws.iter_rows(min_row=1, min_col=40, max_row=100, max_col=51):
                    for cell in row:
                        cell.value = None

                # 保存清除後的版本
                output_buffer = io.BytesIO()
                wb.save(output_buffer)
                output_buffer.seek(0)
                cleaned_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')
                print(raw_df_financial)
                return cleaned_base64, f"處理{stock}的Financial資料時發生錯誤，已清空相關區域: {e}"
            except:
                # 如果連清除都失敗，只能返回原始版本
                return excel_base64, f"處理Financial資料時發生嚴重錯誤: {e}"

    async def process_df_ratios(self, raw_df_ratios, stock, excel_base64):
        """處理ratios數據並寫入Excel base64 - 修復版本"""
        try:
            excel_binary = base64.b64decode(excel_base64)
            excel_buffer = io.BytesIO(excel_binary)
            wb = load_workbook(excel_buffer)
            ws = wb.worksheets[0]

            # 清除舊資料（對應7個區域）
            clear_ranges = [
                (1, 53, 100, 64), (1, 66, 100, 77), (1, 79, 100, 90),
                (1, 92, 100, 103), (1, 105, 100, 116), (1, 118, 100, 129),
                (1, 131, 100, 142)
            ]

            for min_row, min_col, max_row, max_col in clear_ranges:
                for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
                    for cell in row:
                        cell.value = None

            # 【關鍵修復】立即保存清除後的版本
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            cleaned_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')
            # print(raw_df_ratios)

            # 檢查是否為非美國企業錯誤訊息
            if raw_df_ratios is None:
                return cleaned_base64, f'Ratios: {stock}是非美國企業，此頁面須付費！'
            # print('程式有到這邊')
            # 定義各類財務數據的起始位置（對應A.py的7個類別）
            starting_cell = [
                ('Profitability', 1, 53), ('Credit', 1, 66),
                ('Liquidity', 1, 79), ('Working Capital', 1, 92),
                ('Enterprise Value', 1, 105), ('Multiples', 1, 118),
                ('Per Share Data Items', 1, 131)
            ]

            # 處理數據...
            d_1_raw_df_ratios = [y for x in raw_df_ratios.get(stock, []) for y in x]

            for df_amount, df in enumerate(d_1_raw_df_ratios):
                df_column_list = df.columns.tolist()

                # 篩選需要的年份
                years_list = []
                pattern = r'\d{4}\sY'
                for years in df_column_list:
                    if re.match(pattern, years):
                        years_list.append(int(years.split()[0]))

                if years_list:
                    end_year = max(years_list)
                    start_year = end_year - 10
                    drop_column = [
                        x for x in df_column_list
                        if not (re.match(pattern, x) and start_year <= int(x.split()[0]) <= end_year)
                    ]
                    drop_column.pop(0)
                    df = df.drop(columns=drop_column)

                    # 資料轉型為數值型
                    years_data = df.columns[1:]
                    df[years_data] = df[years_data].apply(pd.to_numeric, errors='coerce')

                    # *** 關鍵修改：反轉年份欄位的順序 ***
                    first_col = df.columns[0]  # 第一欄
                    year_cols = df.columns[1:]  # 年份欄位
                    reversed_year_cols = year_cols[::-1]  # 反轉年份欄位順序

                    # 重新組合欄位順序：第一欄 + 反轉的年份欄位
                    new_column_order = [first_col] + list(reversed_year_cols)
                    df_reordered = df[new_column_order]

                    # 將表格資料寫入指定位置並調整格式
                    start_row = starting_cell[df_amount][1]
                    start_col = starting_cell[df_amount][2]

                    for r_idx, row in enumerate(dataframe_to_rows(df_reordered, index=False, header=True),
                                                start=start_row):
                        row_data = list(row)

                        # 第一欄（項目名稱）寫入起始欄位
                        cell = ws.cell(row=r_idx, column=start_col, value=row_data[0])
                        cell.font = Font(size=12, bold=(r_idx == start_row))

                        # 年份數據從右邊開始寫入（從起始欄位+11開始往左）
                        year_data = row_data[1:]  # 除了第一欄以外的年份數據

                        # 從起始欄位+11開始往左寫（假設有12欄的空間）
                        for year_idx, value in enumerate(year_data):
                            column_position = (start_col + 11) - year_idx  # 從右邊往左寫
                            cell = ws.cell(row=r_idx, column=column_position, value=value)
                            cell.font = Font(size=12, bold=(r_idx == start_row))

                    # 自動調整欄寬
                    for col in ws.columns:
                        max_length = max(len(str(cell.value or '')) for cell in col)
                        ws.column_dimensions[col[0].column_letter].width = max_length + 2

            # print('程式有執行到這邊哦')
            # 儲存到base64
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            return modified_base64, f"{stock}的Ratios成功寫入"

        except Exception as e:
            # 【關鍵修復】返回清除後的版本
            try:
                excel_binary = base64.b64decode(excel_base64)
                excel_buffer = io.BytesIO(excel_binary)
                wb = load_workbook(excel_buffer)
                ws = wb.worksheets[0]

                # 清除舊資料
                clear_ranges = [
                    (1, 53, 100, 64), (1, 66, 100, 77), (1, 79, 100, 90),
                    (1, 92, 100, 103), (1, 105, 100, 116), (1, 118, 100, 129),
                    (1, 131, 100, 142)
                ]

                for min_row, min_col, max_row, max_col in clear_ranges:
                    for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
                        for cell in row:
                            cell.value = None

                # 保存清除後的版本
                output_buffer = io.BytesIO()
                wb.save(output_buffer)
                output_buffer.seek(0)
                cleaned_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

                return cleaned_base64, f"處理Ratios資料時發生錯誤，已清空相關區域: {e}"
            except:
                return excel_base64, f"處理Ratios資料時發生嚴重錯誤: {e}"

    async def EPS_PE_MarketCap_data_write_to_excel(self, EPS_PE_MarketCap_content, stock, excel_base64):
        """將 EPS_PE_MarketCap 數據寫入 Excel base64 - 完整格式化處理"""
        try:
            excel_binary = base64.b64decode(excel_base64)
            excel_buffer = io.BytesIO(excel_binary)
            wb = load_workbook(excel_buffer)
            ws = wb.worksheets[0]

            # 清除舊資料 - 清空 EN1 到 EO1
            for i in range(1, 5):
                ws[f'EN{i}'].value = None
                ws[f'EO{i}'].value = None

            # 【關鍵修復】立即保存清除後的版本
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            cleaned_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            # 檢查是否有原始數據
            if not EPS_PE_MarketCap_content:
                return cleaned_base64, 'EPS_PE_MarketCap: 無原始資料，已清空舊數據'

            # 處理資料
            for data in EPS_PE_MarketCap_content.get(stock, {}):
                start_cell = "EN1"
                start_row = int(start_cell[2:])  # 提取行號，例如 "1"

                for i, (key, value) in enumerate(data.items()):
                    row = start_row + i  # 從起始行開始逐行寫入

                    # 寫入鍵到 EY 列
                    key_cell = ws[f"EN{row}"]
                    key_cell.value = key
                    key_cell.font = Font(size=12, bold=False)  # 統一字體設定

                    # 寫入值到 EZ 列
                    value_cell = ws[f"EO{row}"]
                    value_cell.value = value
                    value_cell.font = Font(size=12, bold=False)  # 統一字體設定

                # 自動調整欄寬
                for col in ws.columns:
                    max_length = max(len(str(cell.value or '')) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = max_length + 2

            # 儲存到 base64
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            return modified_base64, f'{stock}的EPS_PE_MarketCap成功寫入並儲存成功'

        except Exception as e:
            try:
                excel_binary = base64.b64decode(excel_base64)
                excel_buffer = io.BytesIO(excel_binary)
                wb = load_workbook(excel_buffer)
                ws = wb.worksheets[0]

                # 清除舊資料 - 清空 EN1 到 EO1
                for i in range(1, 5):
                    ws[f'EN{i}'].value = None
                    ws[f'EO{i}'].value = None

                output_buffer = io.BytesIO()
                wb.save(output_buffer)
                output_buffer.seek(0)
                cleaned_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')
                return cleaned_base64, f"處理 EPS_PE_MarketCap 時發生錯誤: {e}"
            except Exception as e:
                return excel_base64, f"處理 EPS_PE_MarketCap 資料時發生嚴重錯誤: {e}"

    async def _rate_limit(self, api_key="yfinance"):
        """實施速率限制"""
        current_time = time.time()

        if api_key not in self.last_request_time:
            self.last_request_time[api_key] = 0

        time_since_last_request = current_time - self.last_request_time[api_key]

        if time_since_last_request < self.request_delay:
            sleep_time = self.request_delay - time_since_last_request
            # 添加隨機延遲，避免所有請求同時發送
            sleep_time += random.uniform(0.5, 1.5)
            print(f"⏳ 等待 {sleep_time:.1f} 秒以避免API限制...")
            await asyncio.sleep(sleep_time)

        self.last_request_time[api_key] = time.time()

    async def _fetch_stock_data_with_retry(self, stock, max_retries=3):
        """帶重試機制的數據獲取"""
        for attempt in range(max_retries):
            try:
                return await asyncio.to_thread(self._fetch_stock_data, stock)
            except Exception as e:
                if attempt == max_retries - 1:  # 最後一次嘗試
                    raise e

                # 指數退避：每次重試等待時間加倍
                wait_time = (2 ** attempt) * 3 + random.uniform(2, 5)
                print(f"⚠️ 獲取 {stock} 資料失敗，{wait_time:.1f}秒後重試... (嘗試 {attempt + 1}/{max_retries})")
                await asyncio.sleep(wait_time)

    def _fetch_stock_data(self, stock):
        """同步獲取股票數據"""
        # 查詢 10 年期美國國債收益率
        # tnx = yf.Ticker("^TNX")
        # rf_rate = tnx.info['previousClose'] / 100

        # 獲取股票資料
        Stock = yf.Ticker(stock)
        # beta = Stock.info['beta']
        currentPrice = Stock.info['currentPrice']
        symbol = Stock.info['symbol']

        return {
            'Stock': symbol,
            'CurrentPrice': currentPrice,
            # 'beta': beta,
            # 'rf_rate': rf_rate
        }

    async def others_data(self, stock, excel_base64):
        """抓取其他數據並寫入Excel base64"""
        async with self.semaphore:  # 限制併發數量
            try:
                # 添加請求延遲，避免頻率過高
                await self._rate_limit("yfinance")

                # 使用重試機制獲取數據
                dic_data = await self._fetch_stock_data_with_retry(stock)

                print(f'{stock}: {dic_data}')

                # 寫入 Excel（移到線程中執行避免阻塞）
                modified_base64 = await self._write_to_excel(excel_base64, dic_data)

                return modified_base64, f'{stock}的其他資料成功寫入'

            except Exception as e:
                return excel_base64, f"獲取 {stock} 資料時發生錯誤：{str(e)}"

    def write_wacc_data_to_excel(self, stock, wacc_value, excel_base64):
        """將WACC數據寫入Excel"""
        try:
            print(f"正在處理 {stock} 的WACC值: {wacc_value}")

            # 解碼Excel
            excel_binary = base64.b64decode(excel_base64)
            excel_buffer = io.BytesIO(excel_binary)
            wb = load_workbook(excel_buffer)
            ws = wb.worksheets[3]  # 使用第四個工作表

            # 清除舊資料
            ws['C6'] = None  # 你需要根據實際Excel模板調整位置
            ws['C7'] = None
            # 【關鍵修復】立即保存清除後的版本
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            cleaned_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            # 檢查是否有原始數據
            if not wacc_value:
                return cleaned_base64, 'EPS_PE_MarketCap: 無原始資料，已清空舊數據'

            # 例如：假設WACC值寫入C6C7儲存格
            ws['C6'] = wacc_value  # 你需要根據實際Excel模板調整位置
            ws['C7'] = wacc_value
            # 保存修改後的Excel
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            return modified_base64, f"成功將 {stock} 的WACC值 {wacc_value} 寫入Excel"

        except Exception as e:
            try:
                # 解碼Excel
                excel_binary = base64.b64decode(excel_base64)
                excel_buffer = io.BytesIO(excel_binary)
                wb = load_workbook(excel_buffer)
                ws = wb.worksheets[3]  # 使用第四個工作表

                # 清除舊資料
                ws['C5'] = None  # 你需要根據實際Excel模板調整位置

                # 【關鍵修復】立即保存清除後的版本
                output_buffer = io.BytesIO()
                wb.save(output_buffer)
                output_buffer.seek(0)
                cleaned_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')
                return cleaned_base64, f"寫入 {stock} 的WACC數據時發生錯誤: {e}"
            except Exception as e:
                return excel_base64, f"處理 WACC 資料時發生嚴重錯誤: {e}"

    def write_TradeingView_data_to_excel(self, stock, tradingview_data, excel_base64):
        """將TradingView數據寫入Excel"""
        try:
            print(f"正在處理 {stock} 的TradingView數據")

            # 解碼Excel
            excel_binary = base64.b64decode(excel_base64)
            excel_buffer = io.BytesIO(excel_binary)
            wb = load_workbook(excel_buffer)
            ws = wb.worksheets[0]  # 使用第一個工作表

            # 檢查數據類型
            if tradingview_data is None:
                print(f"{stock} 的數據為空")
                return None, f"{stock} 的數據為空"

            # 從 DataFrame 獲取數據
            if hasattr(tradingview_data, 'columns') and hasattr(tradingview_data, 'index'):
                # 確認這是一個 DataFrame

                # 清除舊數據範圍 (可選，根據需要調整範圍)
                for row in range(6, 10):  # 清除前10行
                    for col in range(144, 154):  # 清除EN到EW列 (大概10列)
                        ws.cell(row=row, column=col).value = None

                # 指標名稱從EN開始，年份從EO開始 (EN = 第144列, EO = 第145列)
                start_row = 6
                label_col = 144  # EN列 - 指標名稱列
                start_col = 145  # EO列 - 數據開始列

                # 寫入年份標題 (第一行，從EO1開始)
                for col_idx, year in enumerate(tradingview_data.columns):
                    ws.cell(row=start_row, column=start_col + col_idx).value = int(year)

                # 寫入數據行 (從第二行開始)
                for row_idx, index_name in enumerate(tradingview_data.index):
                    # 寫入行標題 (指標名稱) 到EN列
                    ws.cell(row=start_row + row_idx + 1, column=label_col).value = index_name

                    # 寫入該行的所有數據
                    for col_idx, year in enumerate(tradingview_data.columns):
                        value = tradingview_data.loc[index_name, year]

                        # 處理不同類型的數據
                        if value is None or (isinstance(value, str) and value.lower() == 'none'):
                            cell_value = None
                        elif isinstance(value, str) and ('%' in value):
                            # 保持百分比格式
                            cell_value = value
                        else:
                            try:
                                # 嘗試轉換為數字
                                cell_value = float(value)
                            except (ValueError, TypeError):
                                # 如果無法轉換，保持原始字符串
                                cell_value = value

                        ws.cell(row=start_row + row_idx + 1, column=start_col + col_idx).value = cell_value

                print(f"成功將 {stock} 的數據寫入Excel，範圍: EN{start_row}:EO{start_row + len(tradingview_data.index)}")

                # *** 新增：自動調整欄寬 ***
                # 計算受影響的欄位範圍
                affected_columns = list(range(label_col, start_col + len(tradingview_data.columns)))

                for col_num in affected_columns:
                    try:
                        # 取得該欄位的所有儲存格
                        column_cells = [ws.cell(row=r, column=col_num) for r in
                                        range(1, start_row + len(tradingview_data.index) + 2)]

                        # 計算最大寬度
                        max_length = 0
                        for cell in column_cells:
                            if cell.value is not None:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length

                        # 設定欄寬 (最小寬度為8，最大寬度為50)
                        column_width = min(max(max_length + 2, 8), 50)

                        # 取得欄位字母
                        from openpyxl.utils import get_column_letter
                        column_letter = get_column_letter(col_num)
                        ws.column_dimensions[column_letter].width = column_width

                    except Exception as col_error:
                        print(f"調整欄位 {col_num} 寬度時發生錯誤: {col_error}")
                        continue

            else:
                return None, f"{stock} 的數據格式不正確，預期為DataFrame"

            # 保存修改後的Excel
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            return modified_base64, f"成功將 {stock} 的TradingView數據寫入Excel"

        except Exception as e:
            return None, f"寫入 {stock} 的TradingView數據時發生錯誤: {e}"

    # def write_beta_to_option_excel(self, stock, beta_value, excel_base64):
    #     """將Beta值寫入選擇權Excel的第二個工作表C8儲存格 - 使用xlwings"""
    #     try:
    #         print(f"正在處理 {stock} 的Beta值（選擇權模板）: {beta_value}")
    #
    #         # 解碼並創建臨時檔案
    #         excel_binary = base64.b64decode(excel_base64)
    #
    #         with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsm') as tmp_file:
    #             tmp_file.write(excel_binary)
    #             tmp_path = tmp_file.name
    #
    #         try:
    #             # 用 xlwings 打開
    #             app = xw.App(visible=False)
    #             wb = app.books.open(tmp_path)
    #             ws = wb.sheets[1]  # 第二個工作表
    #
    #             # 清除舊資料
    #             ws.range('B10').value = None
    #             wb.save()
    #
    #             # 檢查是否有原始數據
    #             if beta_value is None:
    #                 wb.close()
    #                 app.quit()
    #
    #                 # 讀取回 base64
    #                 with open(tmp_path, 'rb') as f:
    #                     cleaned_binary = f.read()
    #                 cleaned_base64 = base64.b64encode(cleaned_binary).decode('utf-8')
    #
    #                 return cleaned_base64, f'Beta: {stock} 無原始資料，已清空舊數據'
    #
    #             # 寫入Beta值
    #             ws.range('B10').value = beta_value
    #             wb.save()
    #             wb.close()
    #             app.quit()
    #
    #             # 讀取回 base64
    #             with open(tmp_path, 'rb') as f:
    #                 modified_binary = f.read()
    #             modified_base64 = base64.b64encode(modified_binary).decode('utf-8')
    #
    #             return modified_base64, f"成功將 {stock} 的Beta值 {beta_value} 寫入選擇權模板 (C8)"
    #
    #         finally:
    #             if os.path.exists(tmp_path):
    #                 os.unlink(tmp_path)
    #
    #     except Exception as e:
    #         return excel_base64, f"❌ 處理 {stock} 的Beta資料時發生錯誤: {e}"

    def write_seekingalpha_data_to_excel(self, stock, raw_revenue_growth, excel_base64):
        """將revenue_growth數據寫入Excel"""
        try:
            print(f"正在處理 {stock} 的revenue_growth值: {raw_revenue_growth}")

            # 解碼Excel
            excel_binary = base64.b64decode(excel_base64)
            excel_buffer = io.BytesIO(excel_binary)
            wb = load_workbook(excel_buffer)
            ws = wb.worksheets[3]  # 假設需要寫入的工作表是第四個

            # 提取5Y和10Y的數值（去掉%符號，轉換為float）
            if "5Y" in raw_revenue_growth and "10Y" in raw_revenue_growth:
                # 從 '27.89%' 提取 27.89
                revenue_5y_str = raw_revenue_growth["5Y"].replace('%', '')
                revenue_10y_str = raw_revenue_growth["10Y"].replace('%', '')

                try:
                    if revenue_5y_str == '-':
                        revenue_5y = None
                    else:
                        revenue_5y = float(revenue_5y_str) / 100

                    if revenue_10y_str == '-':
                        revenue_10y = None
                    else:
                        revenue_10y = float(revenue_10y_str) / 100

                except ValueError:
                    return None, f"無法轉換 {stock} 的revenue數值為浮點數: 5Y={revenue_5y_str}, 10Y={revenue_10y_str}"

                # 寫入對應的儲存格
                ws['F5'] = None
                ws['F6'] = None
                ws['F3'] = None

                # 寫入對應的儲存格
                ws['F5'] = revenue_5y  # 5Y數值寫入F5
                ws['F6'] = revenue_10y  # 10Y數值寫入F6
                ws['F3'] = revenue_10y  # 10Y數值寫入F3

                # 可以添加一些格式設定
                # ws['F4'].font = Font(bold=True)
                # ws['F5'].font = Font(bold=True)

                # 保存修改後的Excel
                output_buffer = io.BytesIO()
                wb.save(output_buffer)
                output_buffer.seek(0)
                modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

                return modified_base64, f"成功將 {stock} 的revenue growth數據寫入Excel (5Y: {revenue_5y}%, 10Y: {revenue_10y}%)"

            else:
                return None, f"{stock} 的revenue數據缺少5Y或10Y欄位: {raw_revenue_growth}"

        except Exception as e:
            return None, f"寫入 {stock} 的revenue growth數據時發生錯誤: {e}"

    def write_earnings_date_to_fundamental_excel(self, stock, earnings_data, excel_base64):
        """
        將財報日期寫入股票分析模板（Fundamental Excel）

        ⚠️ 注意：請根據你的模板修改以下儲存格位置

        參數:
            stock: 股票代碼
            earnings_data: {'earnings_date': '2026年3月20日 週五 下午9:00', 'status': 'ESTIMATE'}
            excel_base64: Excel 的 base64 字串
        """
        try:
            print(f"正在處理 {stock} 的財報日期（Fundamental 模板）: {earnings_data}")

            # 解碼 Excel
            excel_binary = base64.b64decode(excel_base64)
            excel_buffer = io.BytesIO(excel_binary)
            wb = load_workbook(excel_buffer)

            # 寫入第一個工作表
            ws = wb.worksheets[0]

            ws['EO12'] = None  # 清除舊資料
            ws['EO13'] = None

            if earnings_data:
                ws['EO12'] = earnings_data.get('earnings_date', '')  # 完整日期時間
                ws['EO13'] = earnings_data.get('status', '')  # ESTIMATE/CONFIRMED

            # 保存修改
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            return modified_base64, f"✅ 成功將 {stock} 的財報日期寫入 Fundamental 模板"

        except Exception as e:
            return excel_base64, f"❌ 寫入 {stock} Fundamental 模板時發生錯誤: {e}"

    def write_earnings_date_to_option_excel(self, stock, earnings_data, file_path):
        """
        將財報日期寫入選擇權模板（Option Chain Template）

        ⚠️ 注意：請根據你的模板修改以下儲存格位置

        參數:
            stock: 股票代碼
            earnings_data: {'earnings_date': '2026年3月20日 週五 下午9:00', 'status': 'ESTIMATE'}
            file_path: Excel 檔案的實體路徑（非 base64）
        """
        import xlwings as xw

        try:
            print(f"正在處理 {stock} 的財報日期（Option 模板）: {earnings_data}")

            # 檢查檔案是否存在
            if not os.path.exists(file_path):
                return file_path, f"❌ 檔案不存在: {file_path}"

            # 用 xlwings 打開（不顯示視窗）
            app = xw.App(visible=False)

            try:
                wb = app.books.open(file_path)
                ws = wb.sheets[2]  # 第3個工作表（或根據你的需求修改）

                ws.range('I9').value = None  # 清除舊資料
                ws.range('I10').value = None

                if earnings_data:
                    ws.range('I9').value = earnings_data.get('earnings_date', '')  # 完整日期時間
                    ws.range('I10').value = earnings_data.get('status', '')  # ESTIMATE/CONFIRMED

                wb.save()
                wb.close()

                return file_path, f"✅ 成功將 {stock} 的財報日期寫入 Option 模板"

            finally:
                app.quit()

        except Exception as e:
            return file_path, f"❌ 寫入 {stock} Option 模板時發生錯誤: {e}"

    async def _write_to_excel(self, excel_base64, dic_data):
        """寫入Excel文件"""

        def write_excel():
            excel_binary = base64.b64decode(excel_base64)
            excel_buffer = io.BytesIO(excel_binary)
            wb = load_workbook(excel_buffer)

            ws = wb.worksheets[0]  # 選擇第一個工作表

            ws['EQ2'] = dic_data['Stock']
            ws['ER2'] = dic_data['CurrentPrice']

            # ws = wb.worksheets[3]
            # ws['C31'] = dic_data['beta']
            # ws['C36'] = dic_data['rf_rate']

            # 儲存到base64
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            return base64.b64encode(output_buffer.read()).decode('utf-8')

        return await asyncio.to_thread(write_excel)

    def save_excel_to_file(self, base64_data: str, output_path: str) -> bool:
        """將 base64 編碼的 Excel 資料保存為實體檔案"""
        try:
            excel_binary = base64.b64decode(base64_data)
            with open(output_path, 'wb') as f:
                f.write(excel_binary)
            print(f"Excel 檔案已保存至：{output_path}")
            return True
        except Exception as e:
            print(f"保存檔案時發生錯誤: {e}")
            return False

    # def create_option_excel_from_base64(self, stock):
    #     """從base64模板創建選擇權Excel文件的base64 - 使用xlwings"""
    #     try:
    #         if Option_Chain_Excel_Template_Base64.strip() == "" or "請將您從轉換工具得到的" in Option_Chain_Excel_Template_Base64:
    #             return "", "❌ 錯誤：請先設定 Option_Chain_Excel_Template_Base64 變數"
    #
    #         # 解碼 base64 並創建臨時檔案
    #         excel_binary = base64.b64decode(Option_Chain_Excel_Template_Base64.strip())
    #
    #         with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsm') as tmp_file:
    #             tmp_file.write(excel_binary)
    #             tmp_path = tmp_file.name
    #
    #         try:
    #             # 用 xlwings 打開並立即關閉（確保格式正確）
    #             app = xw.App(visible=False)
    #             wb = app.books.open(tmp_path)
    #             wb.save()
    #             wb.close()
    #             app.quit()
    #
    #             # 讀取回 base64
    #             with open(tmp_path, 'rb') as f:
    #                 modified_binary = f.read()
    #
    #             excel_base64 = base64.b64encode(modified_binary).decode('utf-8')
    #
    #             return excel_base64, f"成功為 {stock} 創建選擇權Excel檔案"
    #
    #         finally:
    #             # 清理臨時檔案
    #             if os.path.exists(tmp_path):
    #                 os.unlink(tmp_path)
    #
    #     except Exception as e:
    #         return "", f"創建選擇權Excel檔案時發生錯誤: {e}"

    # def write_barchart_data_to_excel(self, stock, barchart_text, excel_base64):
    #     """將Barchart波動率數據寫入選擇權Excel base64 - 使用xlwings"""
    #     try:
    #         print(f"正在處理 {stock} 的Barchart數據")
    #
    #         # 解碼並創建臨時檔案
    #         excel_binary = base64.b64decode(excel_base64)
    #
    #         with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsm') as tmp_file:
    #             tmp_file.write(excel_binary)
    #             tmp_path = tmp_file.name
    #
    #         try:
    #             # 用 xlwings 打開
    #             app = xw.App(visible=False)
    #             wb = app.books.open(tmp_path)
    #             ws = wb.sheets[1]  # 第二個工作表
    #
    #             # 清除舊資料
    #             ws.range('B5').value = None
    #             ws.range('B6').value = None
    #             ws.range('B7').value = None
    #             ws.range('B8').value = None
    #             wb.save()
    #
    #             # 檢查是否有原始數據
    #             if not barchart_text or isinstance(barchart_text, dict):
    #                 wb.close()
    #                 app.quit()
    #
    #                 with open(tmp_path, 'rb') as f:
    #                     cleaned_binary = f.read()
    #                 cleaned_base64 = base64.b64encode(cleaned_binary).decode('utf-8')
    #
    #                 return cleaned_base64, f'Barchart: {stock} 無原始資料，已清空舊數據'
    #
    #             # 解析字串提取數值
    #             iv_match = re.search(r'IV:\s*(\d+\.?\d*)%', barchart_text)
    #             hv_match = re.search(r'HV:\s*(\d+\.?\d*)%', barchart_text)
    #             iv_pctl_match = re.search(r'IV Pctl:\s*(\d+\.?\d*)%', barchart_text)
    #             iv_rank_match = re.search(r'IV Rank:\s*(\d+\.?\d*)%', barchart_text)
    #
    #             iv_value = float(iv_match.group(1)) / 100 if iv_match else None
    #             hv_value = float(hv_match.group(1)) / 100 if hv_match else None
    #             iv_pctl_value = float(iv_pctl_match.group(1)) / 100 if iv_pctl_match else None
    #             iv_rank_value = float(iv_rank_match.group(1)) / 100 if iv_rank_match else None
    #
    #             if all(v is None for v in [iv_value, hv_value, iv_pctl_value, iv_rank_value]):
    #                 wb.close()
    #                 app.quit()
    #
    #                 with open(tmp_path, 'rb') as f:
    #                     cleaned_binary = f.read()
    #                 cleaned_base64 = base64.b64encode(cleaned_binary).decode('utf-8')
    #
    #                 return cleaned_base64, f'❌ {stock} 無法提取Barchart數據，網頁HTML結構可能已改變'
    #
    #             # 寫入數值
    #             ws.range('B5').value = iv_value
    #             ws.range('B6').value = hv_value
    #             ws.range('B7').value = iv_pctl_value
    #             ws.range('B8').value = iv_rank_value
    #
    #             wb.save()
    #             wb.close()
    #             app.quit()
    #
    #             # 讀取回 base64
    #             with open(tmp_path, 'rb') as f:
    #                 modified_binary = f.read()
    #             modified_base64 = base64.b64encode(modified_binary).decode('utf-8')
    #
    #             # 構建成功訊息
    #             extracted_values = []
    #             if iv_value is not None:
    #                 extracted_values.append(f"IV={iv_value:.4f}")
    #             if hv_value is not None:
    #                 extracted_values.append(f"HV={hv_value:.4f}")
    #             if iv_pctl_value is not None:
    #                 extracted_values.append(f"IV Pctl={iv_pctl_value:.4f}")
    #             if iv_rank_value is not None:
    #                 extracted_values.append(f"IV Rank={iv_rank_value:.4f}")
    #
    #             success_msg = f"成功將 {stock} 的Barchart數據寫入Excel ({', '.join(extracted_values)})"
    #
    #             if None in [iv_value, hv_value, iv_pctl_value, iv_rank_value]:
    #                 missing = []
    #                 if iv_value is None:
    #                     missing.append("IV")
    #                 if hv_value is None:
    #                     missing.append("HV")
    #                 if iv_pctl_value is None:
    #                     missing.append("IV Pctl")
    #                 if iv_rank_value is None:
    #                     missing.append("IV Rank")
    #                 success_msg += f" [警告: 無法提取 {', '.join(missing)}]"
    #
    #             return modified_base64, success_msg
    #
    #         finally:
    #             if os.path.exists(tmp_path):
    #                 os.unlink(tmp_path)
    #
    #     except Exception as e:
    #         return excel_base64, f"❌ 處理 {stock} 的Barchart資料時發生錯誤: {e}"

    def flatten_option_chain(self, option_data, stock):
        """
        將選擇權鏈數據展平為DataFrame格式，並確保所有數據都是Excel兼容的
        返回: DataFrame
        """
        try:
            all_options = []

            # 提取基本股票資訊
            base_info = {
                'symbol': option_data.get('symbol'),
                'status': option_data.get('status'),
                'underlying': option_data.get('underlying'),
                'strategy': option_data.get('strategy'),
                'interval': option_data.get('interval'),
                'isDelayed': option_data.get('isDelayed'),
                'isIndex': option_data.get('isIndex'),
                'interestRate': option_data.get('interestRate'),
                'underlyingPrice': option_data.get('underlyingPrice'),
                'volatility': option_data.get('volatility'),
                'daysToExpiration': option_data.get('daysToExpiration'),
                'dividendYield': option_data.get('dividendYield'),
                'numberOfContracts': option_data.get('numberOfContracts'),
                'assetMainType': option_data.get('assetMainType'),
                'assetSubType': option_data.get('assetSubType'),
                'isChainTruncated': option_data.get('isChainTruncated')
            }

            # 處理 Call 選擇權
            if 'callExpDateMap' in option_data:
                for exp_date_key, strikes in option_data['callExpDateMap'].items():
                    for strike_price, contracts in strikes.items():
                        for contract in contracts:
                            option_record = base_info.copy()
                            option_record.update(contract)
                            option_record['expDateKey'] = exp_date_key
                            option_record['strikeKey'] = strike_price
                            # option_record['optionType'] = 'CALL'  # 標記類型
                            all_options.append(option_record)

            # 處理 Put 選擇權
            if 'putExpDateMap' in option_data:
                for exp_date_key, strikes in option_data['putExpDateMap'].items():
                    for strike_price, contracts in strikes.items():
                        for contract in contracts:
                            option_record = base_info.copy()
                            option_record.update(contract)
                            option_record['expDateKey'] = exp_date_key
                            option_record['strikeKey'] = strike_price
                            # option_record['optionType'] = 'PUT'  # 標記類型
                            all_options.append(option_record)

            # 轉換為DataFrame
            df = pd.DataFrame(all_options)

            # 關鍵修復：將複雜數據類型轉換為字串
            df = self._convert_complex_types_to_string(df)

            # 計算 Bid-Ask Spread
            df = self._calculate_bid_ask_spread(df)

            # 計算 Bid-Ask Score
            df = self._calculate_bid_ask_score(df)

            # 計算 Volume Score
            df = self._calculate_volume_score(df)  # 新增

            # 計算 OI Score
            df = self._calculate_oi_score(df)  # 新增

            # 計算 Liquidity Score
            df = self._calculate_liquidity_score(df)  # 新增

            # 計算 Gamma Exposure
            df = self._calculate_Gamma_Exposure(df)  # 新增

            # ✨ 新增：重新排序欄位
            desired_columns = [
                'symbol', 'status', 'underlying', 'strategy', 'interval', 'isDelayed',
                'isIndex', 'interestRate', 'underlyingPrice', 'volatility', 'daysToExpiration',
                'dividendYield', 'numberOfContracts', 'assetMainType', 'assetSubType',
                'isChainTruncated', 'putCall', 'description', 'exchangeName', 'bid', 'ask',
                'last', 'mark', 'bidSize', 'askSize', 'bidAskSize', 'lastSize', 'highPrice',
                'lowPrice', 'openPrice', 'closePrice', 'totalVolume', 'tradeTimeInLong',
                'quoteTimeInLong', 'netChange', 'delta', 'gamma', 'theta', 'vega', 'rho',
                'openInterest', 'timeValue', 'theoreticalOptionValue', 'theoreticalVolatility',
                'optionDeliverablesList', 'strikePrice', 'expirationDate', 'expirationType',
                'lastTradingDay', 'multiplier', 'settlementType', 'deliverableNote',
                'percentChange', 'markChange', 'markPercentChange', 'intrinsicValue',
                'extrinsicValue', 'optionRoot', 'exerciseType', 'high52Week', 'low52Week',
                'nonStandard', 'inTheMoney', 'mini', 'pennyPilot', 'expDateKey', 'strikeKey',
                'Bid-Ask Spread', 'Bid-Ask Score', 'Volume Score', 'OI Score', 'Liquidity Score',
                'Gamma Exposure'
            ]

            # 只保留存在於 DataFrame 中的欄位,並按照指定順序排列
            existing_columns = [col for col in desired_columns if col in df.columns]
            # 加入任何不在desired_columns中但存在於df的欄位
            remaining_columns = [col for col in df.columns if col not in existing_columns]
            final_columns = existing_columns + remaining_columns

            df = df[final_columns]

            return df

        except Exception as e:
            print(f"展平 {stock} 選擇權數據時發生錯誤: {e}")
            import traceback
            traceback.print_exc()
            return None

    def _calculate_bid_ask_spread(self, df):
        """
        計算 Bid-Ask Spread（買賣價差比例）
        公式：=ABS(bid - ask) / ((bid + ask) / 2)

        參數:
            df: DataFrame，必須包含 'bid' 和 'ask' 欄位

        返回:
            DataFrame，新增 'Bid-Ask Spread' 欄位
        """
        try:
            # 確保 bid 和 ask 欄位存在
            if 'bid' not in df.columns or 'ask' not in df.columns:
                print("⚠️ 警告：缺少 bid 或 ask 欄位，無法計算 Bid-Ask Spread")
                df['Bid-Ask Spread'] = None
                return df

            # 將 bid 和 ask 轉換為數值型態（處理可能的字串或 None）
            df['bid'] = pd.to_numeric(df['bid'], errors='coerce')
            df['ask'] = pd.to_numeric(df['ask'], errors='coerce')

            # 計算 Bid-Ask Spread
            # 公式：|bid - ask| / ((bid + ask) / 2)
            def calculate_spread(row):
                bid = row['bid']
                ask = row['ask']

                # 檢查數據有效性
                if pd.isna(bid) or pd.isna(ask):
                    return None

                # 避免除以零
                if bid + ask == 0:
                    return None

                # 計算價差比例
                spread = abs(bid - ask) / ((bid + ask) / 2)

                # 四捨五入到 4 位小數
                return round(spread, 4)

            df['Bid-Ask Spread'] = df.apply(calculate_spread, axis=1)

            # 統計資訊
            valid_spreads = df['Bid-Ask Spread'].dropna()
            if len(valid_spreads) > 0:
                print(f"✓ Bid-Ask Spread 計算完成：{len(valid_spreads)} 筆有效數據")
                print(f"  平均價差: {valid_spreads.mean():.4f}")
                print(f"  最小價差: {valid_spreads.min():.4f}")
                print(f"  最大價差: {valid_spreads.max():.4f}")
            else:
                print("⚠️ 警告：沒有有效的 Bid-Ask Spread 數據")

            return df

        except Exception as e:
            print(f"❌ 計算 Bid-Ask Spread 時發生錯誤: {e}")
            df['Bid-Ask Spread'] = None
            return df

    def _calculate_bid_ask_score(self, df):
        """
        計算 Bid-Ask Score（標準化分數，0-1之間）
        公式：=1 - MIN(1, Spread / PERCENTILE_95)

        原始 Excel 公式：
        =1 - MIN(1, BP2 / PERCENTILE.INC(FILTER($BP:$BP,ISNUMBER($BP:$BP)),0.95))

        參數:
            df: DataFrame，必須包含 'Bid-Ask Spread' 欄位

        返回:
            DataFrame，新增 'Bid-Ask Score' 欄位
        """
        try:
            # 確保 Bid-Ask Spread 欄位存在
            if 'Bid-Ask Spread' not in df.columns:
                print("⚠️ 警告：缺少 Bid-Ask Spread 欄位，無法計算 Bid-Ask Score")
                df['Bid-Ask Score'] = None
                return df

            # 過濾出有效的 Spread 數據（非空值且為數字）
            valid_spreads = df['Bid-Ask Spread'].dropna()

            if len(valid_spreads) == 0:
                print("⚠️ 警告：沒有有效的 Bid-Ask Spread，無法計算 Score")
                df['Bid-Ask Score'] = None
                return df

            # 計算 95th percentile（第95百分位數）
            percentile_95 = valid_spreads.quantile(0.95)

            print(f"✓ Bid-Ask Score 基準（95th percentile）: {percentile_95:.4f}")

            # 計算 Bid-Ask Score
            def calculate_score(spread):
                # 如果 spread 為空值，返回 None
                if pd.isna(spread):
                    return None

                # 避免除以零
                if percentile_95 == 0:
                    return 1.0  # 如果基準為0，所有分數都是完美的

                # 公式：1 - MIN(1, spread / percentile_95)
                normalized_spread = spread / percentile_95
                score = 1 - min(1, normalized_spread)

                # 確保分數在 0-1 之間
                score = max(0, min(1, score))

                # 四捨五入到 4 位小數
                return round(score, 4)

            df['Bid-Ask Score'] = df['Bid-Ask Spread'].apply(calculate_score)

            # 統計資訊
            valid_scores = df['Bid-Ask Score'].dropna()
            if len(valid_scores) > 0:
                print(f"✓ Bid-Ask Score 計算完成：{len(valid_scores)} 筆有效數據")
                print(f"  平均分數: {valid_scores.mean():.4f}")
                print(f"  最小分數: {valid_scores.min():.4f}")
                print(f"  最大分數: {valid_scores.max():.4f}")

                # 分數分布統計
                excellent = (valid_scores >= 0.9).sum()
                good = ((valid_scores >= 0.7) & (valid_scores < 0.9)).sum()
                fair = ((valid_scores >= 0.5) & (valid_scores < 0.7)).sum()
                poor = (valid_scores < 0.5).sum()

                print(f"  分數分布：")
                print(f"    優秀 (≥0.9): {excellent} 筆")
                print(f"    良好 (0.7-0.9): {good} 筆")
                print(f"    普通 (0.5-0.7): {fair} 筆")
                print(f"    較差 (<0.5): {poor} 筆")
            else:
                print("⚠️ 警告：沒有有效的 Bid-Ask Score 數據")

            return df

        except Exception as e:
            print(f"❌ 計算 Bid-Ask Score 時發生錯誤: {e}")
            import traceback
            traceback.print_exc()
            df['Bid-Ask Score'] = None
            return df

    def _calculate_volume_score(self, df):
        """
        計算 Volume Score（成交量分數，對數標準化）
        ✅ 修正：Percentile 計算應包含 0 值
        """
        try:
            import numpy as np

            # 確認欄位存在
            if 'totalVolume' not in df.columns:
                print("⚠️ 警告：缺少 totalVolume 欄位")
                df['Volume Score'] = None
                return df

            # 轉換為數值型態
            df['totalVolume'] = pd.to_numeric(df['totalVolume'], errors='coerce')

            # ✅ 修正：只移除 NaN，保留 0
            valid_volumes = df['totalVolume'].dropna()

            # ✅ 修正：篩選數字（包含0），對齊 Excel 的 ISNUMBER
            valid_volumes = valid_volumes[valid_volumes >= 0]  # 改用 >= 0

            if len(valid_volumes) == 0:
                print("⚠️ 警告：沒有有效的成交量數據")
                df['Volume Score'] = 0
                return df

            # ✅ 計算第95百分位數（包含0的數據）
            percentile_95 = valid_volumes.quantile(0.95)

            print(f"✓ Volume Score 基準（95th percentile）: {percentile_95:,.0f}")
            print(f"  有效數據筆數: {len(valid_volumes)}")
            print(f"  其中為0的筆數: {(valid_volumes == 0).sum()}")

            # 計算分數
            def calculate_volume_score(volume):
                # ✅ 對齊 Excel: IF(AF2=0, 0, ...)
                if pd.isna(volume) or volume == 0:
                    return 0

                # 避免 log(1) = 0 導致除以零
                denominator = np.log(1 + percentile_95)
                if denominator == 0:
                    return 0

                # 公式：LN(1 + volume) / LN(1 + percentile_95)
                score = np.log(1 + volume) / denominator

                # 四捨五入到 4 位小數
                return round(score, 4)

            df['Volume Score'] = df['totalVolume'].apply(calculate_volume_score)

            # 統計資訊
            valid_scores = df['Volume Score'][df['Volume Score'] > 0]
            if len(valid_scores) > 0:
                print(f"✓ Volume Score 計算完成：{len(valid_scores)} 筆有效數據")
                print(f"  平均分數: {valid_scores.mean():.4f}")
                print(f"  最大分數: {valid_scores.max():.4f}")

            return df

        except Exception as e:
            print(f"❌ 計算 Volume Score 時發生錯誤: {e}")
            import traceback
            traceback.print_exc()
            df['Volume Score'] = 0
            return df

    def _calculate_oi_score(self, df):
        """
        計算 OI Score（未平倉量分數，對數標準化）
        ✅ 修正：Percentile 計算應包含 0 值
        """
        try:
            import numpy as np

            if 'openInterest' not in df.columns:
                print("⚠️ 警告：缺少 openInterest 欄位")
                df['OI Score'] = None
                return df

            df['openInterest'] = pd.to_numeric(df['openInterest'], errors='coerce')

            # ✅ 修正：保留 0 值
            valid_oi = df['openInterest'].dropna()
            valid_oi = valid_oi[valid_oi >= 0]  # 改用 >= 0

            if len(valid_oi) == 0:
                print("⚠️ 警告：沒有有效的未平倉量數據")
                df['OI Score'] = 0
                return df

            percentile_95 = valid_oi.quantile(0.95)

            print(f"✓ OI Score 基準（95th percentile）: {percentile_95:,.0f}")
            print(f"  有效數據筆數: {len(valid_oi)}")
            print(f"  其中為0的筆數: {(valid_oi == 0).sum()}")

            def calculate_oi_score(oi):
                # ✅ 對齊 Excel
                if pd.isna(oi) or oi == 0:
                    return 0

                denominator = np.log(1 + percentile_95)
                if denominator == 0:
                    return 0

                score = np.log(1 + oi) / denominator
                return round(score, 4)

            df['OI Score'] = df['openInterest'].apply(calculate_oi_score)

            valid_scores = df['OI Score'][df['OI Score'] > 0]
            if len(valid_scores) > 0:
                print(f"✓ OI Score 計算完成：{len(valid_scores)} 筆有效數據")
                print(f"  平均分數: {valid_scores.mean():.4f}")

            return df

        except Exception as e:
            print(f"❌ 計算 OI Score 時發生錯誤: {e}")
            import traceback
            traceback.print_exc()
            df['OI Score'] = 0
            return df

    def _calculate_liquidity_score(self, df):
        """
        計算 Liquidity Score（綜合流動性分數）
        公式：= 0.4 * Bid-Ask Score + 0.3 * Volume Score + 0.3 * OI Score

        權重說明：
        - Bid-Ask Score (40%)：價差最重要，直接影響交易成本
        - Volume Score (30%)：成交量代表市場活躍度
        - OI Score (30%)：未平倉量代表市場深度

        參數:
            df: DataFrame，必須包含 'Bid-Ask Score', 'Volume Score', 'OI Score'

        返回:
            DataFrame，新增 'Liquidity Score' 欄位
        """
        try:
            # 確認必要欄位存在
            required_columns = ['Bid-Ask Score', 'Volume Score', 'OI Score']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                print(f"⚠️ 警告：缺少必要欄位 {missing_columns}，無法計算 Liquidity Score")
                df['Liquidity Score'] = None
                return df

            # 定義權重
            WEIGHT_BID_ASK = 0.4
            WEIGHT_VOLUME = 0.3
            WEIGHT_OI = 0.3

            # 計算綜合分數
            def calculate_liquidity(row):
                bid_ask_score = row['Bid-Ask Score']
                volume_score = row['Volume Score']
                oi_score = row['OI Score']

                # 檢查是否所有分數都有效
                # 如果任何一個是 None 或 NaN，返回 None
                if pd.isna(bid_ask_score):
                    bid_ask_score = 0
                if pd.isna(volume_score):
                    volume_score = 0
                if pd.isna(oi_score):
                    oi_score = 0

                # 加權平均
                liquidity = (
                        WEIGHT_BID_ASK * bid_ask_score +
                        WEIGHT_VOLUME * volume_score +
                        WEIGHT_OI * oi_score
                )

                # 只限制 ≥ 0，允許 > 1
                # liquidity = max(0, min(1, liquidity))
                liquidity = max(0, liquidity)

                # 四捨五入到 4 位小數
                return round(liquidity, 4)

            df['Liquidity Score'] = df.apply(calculate_liquidity, axis=1)

            # 統計資訊
            valid_scores = df['Liquidity Score'].dropna()
            if len(valid_scores) > 0:
                print(f"✓ Liquidity Score 計算完成：{len(valid_scores)} 筆數據")
                print(f"  平均分數: {valid_scores.mean():.4f}")
                print(f"  最小分數: {valid_scores.min():.4f}")
                print(f"  最大分數: {valid_scores.max():.4f}")

                # 流動性等級分布
                excellent = (valid_scores >= 0.8).sum()
                good = ((valid_scores >= 0.6) & (valid_scores < 0.8)).sum()
                fair = ((valid_scores >= 0.4) & (valid_scores < 0.6)).sum()
                poor = (valid_scores < 0.4).sum()

                print(f"  流動性等級分布：")
                print(f"    🟢 優秀 (≥0.8): {excellent} 筆 ({excellent / len(valid_scores) * 100:.1f}%)")
                print(f"    🔵 良好 (0.6-0.8): {good} 筆 ({good / len(valid_scores) * 100:.1f}%)")
                print(f"    🟡 普通 (0.4-0.6): {fair} 筆 ({fair / len(valid_scores) * 100:.1f}%)")
                print(f"    🔴 較差 (<0.4): {poor} 筆 ({poor / len(valid_scores) * 100:.1f}%)")

                # 推薦交易的合約數量
                tradeable = (valid_scores >= 0.6).sum()
                print(f"  💡 建議交易合約：{tradeable} 筆（流動性≥0.6）")

            return df

        except Exception as e:
            print(f"❌ 計算 Liquidity Score 時發生錯誤: {e}")
            import traceback
            traceback.print_exc()
            df['Liquidity Score'] = None
            return df

    def _calculate_Gamma_Exposure(self, df):
        """
        計算 Gamma Exposure（Gamma 風險敞口）
        公式：gamma * openInterest

        參數:
            df: DataFrame，必須包含 'gamma' 和 'openInterest' 欄位

        返回:
            DataFrame，新增 'Gamma Exposure' 欄位
        """
        try:
            # 確認必要欄位存在
            if 'gamma' not in df.columns or 'openInterest' not in df.columns:
                print("⚠️ 警告：缺少 gamma 或 openInterest 欄位，無法計算 Gamma Exposure")
                df['Gamma Exposure'] = None
                return df

            # 轉換為數值型態（防禦性編程）
            df['gamma'] = pd.to_numeric(df['gamma'], errors='coerce')
            df['openInterest'] = pd.to_numeric(df['openInterest'], errors='coerce')

            # 計算 Gamma Exposure
            df['Gamma Exposure'] = df['gamma'] * df['openInterest']

            return df

        except Exception as e:
            print(f"❌ 計算 Gamma Exposure 時發生錯誤: {e}")
            import traceback
            traceback.print_exc()
            df['Gamma Exposure'] = None
            return df

    def _convert_complex_types_to_string(self, df):
        """
        將DataFrame中的複雜數據類型（字典、列表）轉換為字串
        """
        import json

        for col in df.columns:
            # 檢查該列是否包含複雜類型
            if df[col].dtype == 'object':
                def convert_value(val):
                    if val is None:
                        return None
                    elif isinstance(val, (dict, list)):
                        # 將字典或列表轉換為JSON字串
                        return json.dumps(val, ensure_ascii=False)
                    elif isinstance(val, (int, float, str, bool)):
                        return val
                    else:
                        # 其他類型嘗試轉換為字串
                        return str(val)

                df[col] = df[col].apply(convert_value)

        return df

    # def write_option_chain_to_excel(self, stock, option_df, excel_base64):
    #     """將選擇權鏈DataFrame寫入Excel base64 - 使用xlwings"""
    #     try:
    #         if option_df is None or option_df.empty:
    #             return excel_base64, f"{stock} 選擇權數據為空"
    #
    #         print(f"準備寫入 {stock} 的選擇權數據: {len(option_df)} 筆合約, {len(option_df.columns)} 個欄位")
    #
    #         # 解碼並創建臨時檔案
    #         excel_binary = base64.b64decode(excel_base64)
    #
    #         with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsm') as tmp_file:
    #             tmp_file.write(excel_binary)
    #             tmp_path = tmp_file.name
    #
    #         try:
    #             # 用 xlwings 打開
    #             app = xw.App(visible=False)
    #             wb = app.books.open(tmp_path)
    #
    #             # 找到或創建 OptionChain 工作表
    #             sheet_name = 'OptionChain'
    #             if sheet_name in [sheet.name for sheet in wb.sheets]:
    #                 ws = wb.sheets[sheet_name]
    #                 # 清除舊數據
    #                 ws.clear()
    #             else:
    #                 ws = wb.sheets.add(sheet_name)
    #
    #             # 寫入 DataFrame（xlwings 可以直接寫入 DataFrame）
    #             ws.range('A1').options(index=False).value = option_df
    #
    #             # 設置表頭格式
    #             header_range = ws.range(f'A1:{ws.range("A1").end("right").address}')
    #             header_range.font.bold = True
    #             header_range.font.size = 11
    #
    #             # 自動調整欄寬
    #             ws.autofit(axis='columns')
    #
    #             wb.save()
    #             wb.close()
    #             app.quit()
    #
    #             # 讀取回 base64
    #             with open(tmp_path, 'rb') as f:
    #                 modified_binary = f.read()
    #             modified_base64 = base64.b64encode(modified_binary).decode('utf-8')
    #
    #             print(f"✅ 成功寫入 {stock} 的選擇權數據到Excel")
    #             return modified_base64, f"✅ 成功將 {stock} 的選擇權數據寫入Excel ({len(option_df)} 筆合約)"
    #
    #         finally:
    #             if os.path.exists(tmp_path):
    #                 os.unlink(tmp_path)
    #
    #     except Exception as e:
    #         print(f"❌ 寫入 {stock} 選擇權數據時發生錯誤: {e}")
    #         import traceback
    #         traceback.print_exc()
    #         return excel_base64, f"❌ 寫入 {stock} 選擇權數據時發生錯誤: {e}"

    def batch_write_options_to_excel(self, stock_data_dict, excel_files_dict):
        """
        批次處理多支股票的選擇權數據寫入 (直接操作檔案,不用 base64)

        參數:
            stock_data_dict: {stock: {'option_chain': df, 'beta': value, 'barchart': text}}
            excel_files_dict: {stock: file_path}  👈 改成檔案路徑

        返回:
            updated_excel_files: {stock: file_path}  👈 返回檔案路徑
            messages: {stock: message}
        """
        import xlwings as xw
        import re

        updated_files = {}
        messages = {}

        try:
            # 🔥 步驟 1: 一次啟動 xlwings,批次處理所有股票
            print(f"🚀 啟動 Excel 批次處理 {len(excel_files_dict)} 支股票...")
            app = xw.App(visible=False)

            try:
                for stock, file_path in excel_files_dict.items():
                    try:
                        if stock not in stock_data_dict:
                            messages[stock] = "⚠️ 無數據可寫入"
                            updated_files[stock] = file_path
                            continue

                        # 🔥 檢查檔案是否存在
                        if not os.path.exists(file_path):
                            messages[stock] = f"❌ 檔案不存在: {file_path}"
                            continue

                        data = stock_data_dict[stock]
                        wb = app.books.open(file_path)
                        ws = wb.sheets[1]  # 第二個工作表

                        # 寫入 Beta
                        if 'beta' in data and data['beta'] is not None:
                            ws.range('B10').value = data['beta']

                        # 寫入 Barchart
                        if 'barchart' in data and data['barchart'] is not None:
                            barchart_text = data['barchart']

                            if isinstance(barchart_text, str):
                                iv_match = re.search(r'IV:\s*(\d+\.?\d*)%', barchart_text)
                                hv_match = re.search(r'HV:\s*(\d+\.?\d*)%', barchart_text)
                                iv_pctl_match = re.search(r'IV Pctl:\s*(\d+\.?\d*)%', barchart_text)
                                iv_rank_match = re.search(r'IV Rank:\s*(\d+\.?\d*)%', barchart_text)

                                if iv_match:
                                    ws.range('B5').value = float(iv_match.group(1)) / 100
                                if hv_match:
                                    ws.range('B6').value = float(hv_match.group(1)) / 100
                                if iv_pctl_match:
                                    ws.range('B7').value = float(iv_pctl_match.group(1)) / 100
                                if iv_rank_match:
                                    ws.range('B8').value = float(iv_rank_match.group(1)) / 100

                        # 寫入 Option Chain
                        if 'option_chain' in data and data['option_chain'] is not None:
                            option_df = data['option_chain']

                            if not option_df.empty:
                                sheet_name = 'OptionChain'
                                if sheet_name in [sheet.name for sheet in wb.sheets]:
                                    ws_option = wb.sheets[sheet_name]
                                    ws_option.clear()
                                else:
                                    ws_option = wb.sheets.add(sheet_name)

                                ws_option.range('A1').options(index=False).value = option_df

                                # 設置表頭格式
                                header_range = ws_option.range(f'A1:{ws_option.range("A1").end("right").address}')
                                header_range.font.bold = True
                                header_range.font.size = 11

                                ws_option.autofit(axis='columns')

                        wb.save()  # 🔥 直接儲存到原檔案
                        wb.close()

                        # 🔥 返回檔案路徑 (不是 base64)
                        updated_files[stock] = file_path
                        messages[stock] = f"✅ 批次寫入成功"

                    except Exception as e:
                        messages[stock] = f"❌ 寫入失敗: {e}"
                        if 'wb' in locals():
                            try:
                                wb.close()
                            except:
                                pass
                        continue

            finally:
                app.quit()
                print("✅ Excel 應用程式已關閉")

        except Exception as e:
            print(f"❌ 批次寫入過程發生錯誤: {e}")

        return updated_files, messages

    def log(self, message):
        """簡易日誌方法 (如果沒有的話)"""
        print(message)

    # def batch_create_option_excels_from_base64(self, stocks):
    #     """批次創建多支股票的選擇權Excel檔案"""
    #     import tempfile
    #     import os
    #
    #     results = {}
    #     temp_dir = tempfile.mkdtemp()
    #
    #     try:
    #         # 🔥 只啟動一次 xlwings
    #         app = xw.App(visible=False)
    #
    #         try:
    #             for stock in stocks:
    #                 try:
    #                     # 解碼模板
    #                     excel_binary = base64.b64decode(
    #                         Option_Chain_Excel_Template_Base64.strip()
    #                     )
    #
    #                     # 寫入臨時檔案
    #                     temp_path = os.path.join(temp_dir, f"{stock}_temp.xlsm")
    #                     with open(temp_path, 'wb') as f:
    #                         f.write(excel_binary)
    #
    #                     # 開啟並立即儲存 (確保格式正確)
    #                     wb = app.books.open(temp_path)
    #                     wb.save()
    #                     wb.close()
    #
    #                     # 讀回 base64
    #                     with open(temp_path, 'rb') as f:
    #                         modified_binary = f.read()
    #
    #                     excel_base64 = base64.b64encode(modified_binary).decode('utf-8')
    #                     results[stock] = (excel_base64, f"✅ 成功為 {stock} 創建選擇權Excel檔案")
    #
    #                 except Exception as e:
    #                     results[stock] = ("", f"❌ 創建 {stock} 檔案時發生錯誤: {e}")
    #
    #         finally:
    #             app.quit()  # 🔥 只關閉一次
    #
    #     finally:
    #         # 清理臨時檔案
    #         for stock in stocks:
    #             temp_path = os.path.join(temp_dir, f"{stock}_temp.xlsm")
    #             if os.path.exists(temp_path):
    #                 os.unlink(temp_path)
    #         try:
    #             os.rmdir(temp_dir)
    #         except:
    #             pass
    #
    #     return results