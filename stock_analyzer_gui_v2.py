# 在您的 stock_analyzer_gui.py 檔案最上方添加這段程式碼

import os
import sys


def setup_playwright_path():
    """設定 Playwright 瀏覽器路徑"""

    # 如果是打包後的執行檔
    if hasattr(sys, '_MEIPASS'):
        # PyInstaller 打包後的臨時資料夾
        base_path = sys._MEIPASS

        # 設定瀏覽器路徑到打包後的位置
        browser_path = os.path.join(base_path, 'ms-playwright')

        if os.path.exists(browser_path):
            os.environ['PLAYWRIGHT_BROWSERS_PATH'] = browser_path
            print(f"設定瀏覽器路徑: {browser_path}")
        else:
            # 嘗試原始路徑
            original_path = r'C:\Users\2993\AppData\Local\ms-playwright'
            if os.path.exists(original_path):
                os.environ['PLAYWRIGHT_BROWSERS_PATH'] = original_path
                print(f"使用原始瀏覽器路徑: {original_path}")
    else:
        # 開發環境，使用原始路徑
        original_path = r'C:\Users\2993\AppData\Local\ms-playwright'
        if os.path.exists(original_path):
            os.environ['PLAYWRIGHT_BROWSERS_PATH'] = original_path


# 在導入 playwright 之前呼叫這個函數
setup_playwright_path()

# 然後才導入 playwright
from playwright.sync_api import sync_playwright

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
import asyncio
import sys
import os
from datetime import datetime
import time

# 匯入您原本的模組
import base64
import io
from playwright.async_api import async_playwright
import pandas as pd
import random
from io import StringIO
import re
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from bs4 import BeautifulSoup
import yfinance as yf
import aiohttp
import json
from concurrent.futures import ThreadPoolExecutor

# 您需要將您的 Excel 模板轉換後的 base64 字串放在這裡
EXCEL_TEMPLATE_BASE64 = """
我的模板
"""


# ===== 將您原本的 StockScraper, StockProcess, StockManager 類別複製到這裡 =====
# [這裡應該包含您原本的所有類別，由於篇幅限制我省略了，您需要將它們完整複製過來]

class StockScraper:
    def __init__(self, stocks, headless=True, max_concurrent=3):
        """
        初始化爬蟲類別。
        stocks: 股票代碼的列表
        headless: 是否使用無頭模式
        max_concurrent: 同時執行的股票數量（控制併發數）
        """
        self.stocks = stocks
        self.headless = headless
        self.max_concurrent = max_concurrent
        self.browser = None
        self.playwright = None

    async def setup_browser(self):
        """設定瀏覽器環境。"""
        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.launch(
            headless=self.headless,
            args=[
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",
                "--disable-accelerated-2d-canvas",
                "--disable-gpu",
                "--disable-features=IsolateOrigins,site-per-process",
                "--disable-background-timer-throttling",
            ],
        )

    async def cleanup(self):
        """清理資源。"""
        if self.browser:
            await self.browser.close()
        if self.playwright:
            await self.playwright.stop()

    async def fetch_summary_data(self, stock, semaphore):
        """抓取單一股票的數據（summary）。"""
        async with semaphore:
            try:
                context = await self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
                    viewport={"width": 800, "height": 600},
                )
                try:
                    page_summary = await context.new_page()
                    summary = await asyncio.gather(self.get_summary(stock, page_summary))
                    return {stock: summary}
                finally:
                    await context.close()
            except Exception as e:
                return {"stock": stock, "error": str(e)}

    async def get_summary(self, stock, page, retries=1):
        """抓取特定股票的摘要資料並回傳 DataFrame。"""
        URL = f'https://www.roic.ai/quote/{stock}'
        attempt = 0

        while attempt < retries:
            try:
                await asyncio.sleep(random.uniform(1, 3))
                await page.goto(URL, wait_until='load', timeout=50000)
                await page.wait_for_selector('table.w-full.caption-bottom.text-sm.table-fixed', timeout=100000)
                content = await page.content()
                dfs = pd.read_html(StringIO(content))
                return dfs
            except Exception as e:
                attempt += 1
                if attempt == retries:
                    return f"Error for {stock}: {e}"

        return f"Failed to retrieve data for {stock}"

    async def run_summary(self):
        await self.setup_browser()
        semaphore = asyncio.Semaphore(self.max_concurrent)
        try:
            tasks = [self.fetch_summary_data(stock, semaphore) for stock in self.stocks]
            result = await asyncio.gather(*tasks)
        finally:
            await self.cleanup()
        return result

    async def fetch_financials_data(self, stock, semaphore):
        """抓取單一股票的數據（financials）。"""
        async with semaphore:
            try:
                context = await self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
                    viewport={"width": 800, "height": 600},
                )
                try:
                    page_financials = await context.new_page()
                    financials = await asyncio.gather(self.get_financials(stock, page_financials))
                    return {stock: financials}
                finally:
                    await context.close()
            except Exception as e:
                return {"stock": stock, "error": str(e)}

    async def get_financials(self, stock, page, retries=1):
        """抓取特定股票的財務資料並回傳 DataFrame。"""
        URL = f'https://www.roic.ai/quote/{stock}/financials'
        attempt = 0

        while attempt < retries:
            try:
                await asyncio.sleep(random.uniform(1, 3))
                await page.goto(URL, wait_until='load', timeout=50000)

                if await page.query_selector(
                        'div.rounded-lg.bg-card.text-card-foreground.shadow-sm.mx-auto.flex.w-\\[500px\\].flex-col.items-center.border.drop-shadow-lg'):
                    return f'{stock}是非美國企業，此頁面須付費！'
                else:
                    await page.wait_for_selector('table.w-full.caption-bottom.text-sm.table-fixed', timeout=100000)
                    content = await page.content()
                    dfs = pd.read_html(StringIO(content))
                    return dfs

            except Exception as e:
                attempt += 1
                if attempt == retries:
                    return f"Error for {stock}: {e}"

        return f"Failed to retrieve data for {stock}"

    async def run_financial(self):
        await self.setup_browser()
        semaphore = asyncio.Semaphore(self.max_concurrent)
        try:
            tasks = [self.fetch_financials_data(stock, semaphore) for stock in self.stocks]
            result = await asyncio.gather(*tasks)
        finally:
            await self.cleanup()
        return result

    async def fetch_ratios_data(self, stock, semaphore):
        """抓取單一股票的數據（Ratios）。"""
        async with semaphore:
            try:
                context = await self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
                    viewport={"width": 800, "height": 600},
                )
                try:
                    page_ratios = await context.new_page()
                    ratios = await asyncio.gather(self.get_ratios(stock, page_ratios))
                    return {stock: ratios}
                finally:
                    await context.close()
            except Exception as e:
                return {"stock": stock, "error": str(e)}

    async def get_ratios(self, stock, page, retries=1):
        """抓取特定股票的比率資料並回傳 DataFrame。"""
        URL = f'https://www.roic.ai/quote/{stock}/ratios'
        attempt = 0

        while attempt < retries:
            try:
                await asyncio.sleep(random.uniform(1, 3))
                await page.goto(URL, wait_until='load', timeout=50000)

                if await page.query_selector(
                        'div.rounded-lg.bg-card.text-card-foreground.shadow-sm.mx-auto.flex.w-\\[500px\\].flex-col.items-center.border.drop-shadow-lg'):
                    return f'{stock}是非美國企業，此頁面須付費！'
                else:
                    await page.wait_for_selector('table.w-full.caption-bottom.text-sm.table-fixed', timeout=100000)
                    content = await page.content()
                    dfs = pd.read_html(StringIO(content))
                    return dfs

            except Exception as e:
                attempt += 1
                if attempt == retries:
                    return f"Error for {stock}: {e}"

        return f"Failed to retrieve data for {stock}"

    async def run_ratios(self):
        await self.setup_browser()
        semaphore = asyncio.Semaphore(self.max_concurrent)
        try:
            tasks = [self.fetch_ratios_data(stock, semaphore) for stock in self.stocks]
            result = await asyncio.gather(*tasks)
        finally:
            await self.cleanup()
        return result

    async def fetch_EPS_PE_MarketCap_data(self, stock, semaphore):
        """抓取單一股票的數據（EPS_PE_MarketCap）。"""
        async with semaphore:
            try:
                context = await self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
                    viewport={"width": 800, "height": 600},
                )
                try:
                    page_EPS_PE_MarketCap = await context.new_page()
                    EPS_PE_MarketCap = await asyncio.gather(self.get_EPS_PE_MarketCap(stock, page_EPS_PE_MarketCap))
                    return {stock: EPS_PE_MarketCap}
                finally:
                    await context.close()
            except Exception as e:
                return {"stock": stock, "error": str(e)}

    async def get_EPS_PE_MarketCap(self, stock, page, retries=3):
        url = f'https://www.roic.ai/quote/{stock}'
        attempt = 0
        while attempt < retries:
            try:
                await asyncio.sleep(random.uniform(1, 3))
                await page.goto(url, wait_until='load', timeout=30000)
                await page.wait_for_selector('table', timeout=30000)
                content = await page.content()
                soup = BeautifulSoup(content, 'html.parser')
                span_string = soup.find_all('span', attrs={'stock_class': 'flex text-sm uppercase text-muted-foreground'})
                span_int_value = soup.find_all('span', attrs={'stock_class': 'flex text-lg text-foreground'})

                dic_data = {
                    span_string[0].text: float(span_int_value[0].text),
                    span_string[1].text: float(span_int_value[1].text),
                    span_string[2].text: span_int_value[2].text,
                    span_string[3].text: span_int_value[3].text
                }
                return dic_data
            except Exception as e:
                return f'error message:{e}'

    async def run_EPS_PE_MarketCap(self):
        await self.setup_browser()
        semaphore = asyncio.Semaphore(self.max_concurrent)
        try:
            tasks = [self.fetch_EPS_PE_MarketCap_data(stock, semaphore) for stock in self.stocks]
            result = await asyncio.gather(*tasks)
        finally:
            await self.cleanup()
        return result

    async def EPS_Growth_Rate_and_write_to_excel(self, stock, excel_base64):
        """抓取EPS成長率並寫入Excel"""
        if '-' in stock:
            stock = ''.join(['.' if char == '-' else char for char in stock])

        async with aiohttp.ClientSession() as session:
            async with session.get(f'https://api.stockboss.io/api/symbol?symbol={stock}') as response:
                content = await response.text()
                dic = json.loads(content)
                # print(dic['symbol']['guru_summary']['summary']['summary']['company_data']['wacc'])
                # wacc = float(dic['symbol']['guru_summary']['summary']['summary']['company_data']['wacc'])/100
                l_eps_growth5y = []
                try:
                    EPS_Growth_Rate_3_Year = \
                        dic['symbol']['keyratio']['keyratio']['annuals']['3-Year EPS Growth Rate %'][-1]
                    EPS_Growth_Rate_5_Year = \
                        dic['symbol']['keyratio']['keyratio']['annuals']['5-Year EPS Growth Rate %'][-1]
                    EPS_Growth_Rate_10_Year = \
                        dic['symbol']['keyratio']['keyratio']['annuals']['10-Year EPS Growth Rate %'][-1]

                    EPS_Growth_Rate_3_Year = 0 if EPS_Growth_Rate_3_Year == '-' else EPS_Growth_Rate_3_Year
                    EPS_Growth_Rate_5_Year = 0 if EPS_Growth_Rate_5_Year == '-' else EPS_Growth_Rate_5_Year
                    EPS_Growth_Rate_10_Year = 0 if EPS_Growth_Rate_10_Year == '-' else EPS_Growth_Rate_10_Year

                    l_eps_growth5y = l_eps_growth5y + [EPS_Growth_Rate_3_Year, EPS_Growth_Rate_5_Year,
                                                       EPS_Growth_Rate_10_Year]

                except KeyError as e:
                    return f"EPS_Growth_Rate的dictionary錯誤：{stock}", excel_base64

                # 選擇成長率：如果最小值大於 0，則取最小值，否則取最大值
                selected_growth_rate = min(l_eps_growth5y) / 100 if min(l_eps_growth5y) > 0 else max(
                    l_eps_growth5y) / 100
                # print(selected_growth_rate)
                # print(wacc)
                # 寫入 Excel
                try:
                    excel_binary = base64.b64decode(excel_base64)
                    excel_buffer = io.BytesIO(excel_binary)
                    wb = load_workbook(excel_buffer)
                    ws = wb.worksheets[3]  # 假設需要寫入的工作表是第四個
                    ws['C4'] = selected_growth_rate
                    # ws['C6'] = wacc

                    output_buffer = io.BytesIO()
                    wb.save(output_buffer)
                    output_buffer.seek(0)
                    modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

                    return f"{stock}的EPS成長率成功寫入", modified_base64

                except Exception as e:
                    return f"寫入Excel時發生錯誤：{e}", excel_base64


class StockProcess:
    def __init__(self, max_concurrent=2, request_delay=2.0):
        # 將 semaphore 移到類別層級，確保全域限制
        self.semaphore = asyncio.Semaphore(max_concurrent)
        self.request_delay = request_delay  # 請求之間的延遲（秒）
        self.last_request_time = {}  # 記錄每個API的上次請求時間

    def create_excel_from_base64(self, stock):
        """從base64模板創建Excel文件的base64"""
        try:
            if EXCEL_TEMPLATE_BASE64.strip() == "" or "請將您從轉換工具得到的" in EXCEL_TEMPLATE_BASE64:
                return "", "❌ 錯誤：請先設定 EXCEL_TEMPLATE_BASE64 變數"

            excel_binary = base64.b64decode(EXCEL_TEMPLATE_BASE64.strip())
            excel_buffer = io.BytesIO(excel_binary)
            workbook = load_workbook(excel_buffer)

            # 儲存修改後的檔案到記憶體
            output_buffer = io.BytesIO()
            workbook.save(output_buffer)
            output_buffer.seek(0)
            excel_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            return excel_base64, f"成功為 {stock} 創建Excel檔案"

        except Exception as e:
            return "", f"創建Excel檔案時發生錯誤: {e}"

    def write_wacc_to_excel_base64(self, stock_code: str, wacc_value: str, excel_base64: str) -> tuple:
        """將WACC值寫入Excel的base64"""
        try:
            excel_binary = base64.b64decode(excel_base64)
            excel_buffer = io.BytesIO(excel_binary)
            workbook = load_workbook(excel_buffer)

            if "現金流折現法" in workbook.sheetnames:
                worksheet = workbook["現金流折現法"]
            else:
                worksheet = workbook.active

            # 寫入 WACC 值到 C6 儲存格
            if wacc_value != "未知":
                if wacc_value.endswith('%'):
                    numeric_value = float(wacc_value.rstrip('%')) / 100
                    worksheet['C6'] = numeric_value
                    worksheet['C6'].number_format = '0.00%'
                else:
                    worksheet['C6'] = wacc_value
            else:
                worksheet['C6'] = "無法取得"

            # 儲存修改後的檔案到記憶體
            output_buffer = io.BytesIO()
            workbook.save(output_buffer)
            output_buffer.seek(0)
            modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            return modified_base64, f"已將 {stock_code} 的 WACC 值 ({wacc_value}) 寫入 C6 儲存格"

        except Exception as e:
            return excel_base64, f"操作 Excel 檔案時發生錯誤: {e}"

    async def process_df_summary(self, raw_df_summary, stock, excel_base64):
        """處理summary數據並寫入Excel base64"""
        if raw_df_summary:
            try:
                excel_binary = base64.b64decode(excel_base64)
                excel_buffer = io.BytesIO(excel_binary)
                wb = load_workbook(excel_buffer)
                ws = wb.worksheets[0]

                # 清除舊資料
                for row in ws.iter_rows(min_row=1, min_col=1, max_row=30, max_col=12):
                    for cell in row:
                        cell.value = None

                d_1_raw_df_summary = [y for x in raw_df_summary.get(stock, pd.DataFrame({})) for y in x]

                for df_amount, df in enumerate(d_1_raw_df_summary):
                    df_column_list = df.columns.tolist()

                    # 篩選需要的年份
                    years_list = []
                    pattern = r'\d{4}\sY'
                    for years in df_column_list:
                        if re.match(pattern, years):
                            years_list.append(int(years.split()[0]))

                    if years_list:
                        end_year = max(years_list)
                        start_year = end_year - 10
                        drop_column = [
                            x for x in df_column_list
                            if not (re.match(pattern, x) and start_year <= int(x.split()[0]) <= end_year)
                        ]
                        drop_column.pop(0)  # 移除列名的第一欄
                        df = df.drop(columns=drop_column)

                        # 資料轉型為數值型
                        years_data = df.columns[1:]
                        df[years_data] = df[years_data].apply(pd.to_numeric, errors='coerce')

                        # *** 關鍵修改：反轉年份欄位的順序 ***
                        # 保留第一欄（通常是項目名稱），但反轉年份欄位
                        first_col = df.columns[0]  # 第一欄
                        year_cols = df.columns[1:]  # 年份欄位
                        reversed_year_cols = year_cols[::-1]  # 反轉年份欄位順序

                        # 重新組合欄位順序：第一欄 + 反轉的年份欄位
                        new_column_order = [first_col] + list(reversed_year_cols)
                        df_reordered = df[new_column_order]

                        # 將資料寫入 Excel，並設置欄位格式
                        start_row = 1

                        for r_idx, row in enumerate(dataframe_to_rows(df_reordered, index=False, header=True),
                                                    start=start_row):
                            row_data = list(row)

                            # 第一欄（項目名稱）寫入 A 欄
                            cell = ws.cell(row=r_idx, column=1, value=row_data[0])
                            cell.font = Font(size=12, bold=(r_idx == start_row))

                            # 年份數據從右邊開始寫入（從 L 欄開始往左）
                            year_data = row_data[1:]  # 除了第一欄以外的年份數據

                            # L欄是第12欄，從L欄開始往左寫
                            for year_idx, value in enumerate(year_data):
                                column_position = 12 - year_idx  # L=12, K=11, J=10...
                                cell = ws.cell(row=r_idx, column=column_position, value=value)
                                cell.font = Font(size=12, bold=(r_idx == start_row))

                        # 自動調整欄寬
                        for col in ws.columns:
                            max_length = max(len(str(cell.value or '')) for cell in col)
                            ws.column_dimensions[col[0].column_letter].width = max_length + 2

                # 儲存到base64
                output_buffer = io.BytesIO()
                wb.save(output_buffer)
                output_buffer.seek(0)
                modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

                return modified_base64, f"{stock}的Summary成功寫入"

            except Exception as e:
                return excel_base64, f"處理Summary資料時發生錯誤: {e}"
        else:
            return excel_base64, '無原始資料'

    async def process_df_financial(self, raw_df_financial, stock, excel_base64):
        """處理financial數據並寫入Excel base64"""
        if raw_df_financial:
            try:
                excel_binary = base64.b64decode(excel_base64)
                excel_buffer = io.BytesIO(excel_binary)
                wb = load_workbook(excel_buffer)
                ws = wb.worksheets[0]

                # 定義各類財務數據的起始位置
                starting_cell = [("IncomeStatement", 1, 14),  # N1
                                 ("BalanceSheet", 1, 27),  # AA1
                                 ("CashFlowStatement", 1, 40)]  # AN1

                # 清除舊資料
                for row in ws.iter_rows(min_row=1, min_col=14, max_row=100, max_col=25):
                    for cell in row:
                        cell.value = None
                for row in ws.iter_rows(min_row=1, min_col=27, max_row=100, max_col=38):
                    for cell in row:
                        cell.value = None
                for row in ws.iter_rows(min_row=1, min_col=40, max_row=100, max_col=51):
                    for cell in row:
                        cell.value = None

                if raw_df_financial.get(stock) == [f'{stock}是非美國企業，此頁面須付費！']:
                    return excel_base64, f'{stock}是非美國企業，此頁面須付費！'

                else:
                    d_1_raw_df_financial = [y for x in raw_df_financial.get(stock, pd.DataFrame({})) for y in x]

                    for df_amount, df in enumerate(d_1_raw_df_financial):
                        df_column_list = df.columns.tolist()

                        # 篩選需要的年份
                        years_list = []
                        pattern = r'\d{4}\sY'
                        for years in df_column_list:
                            if re.match(pattern, years):
                                years_list.append(int(years.split()[0]))

                        if years_list:
                            end_year = max(years_list)
                            start_year = end_year - 10
                            drop_column = [
                                x for x in df_column_list
                                if not (re.match(pattern, x) and start_year <= int(x.split()[0]) <= end_year)
                            ]
                            drop_column.pop(0)
                            df = df.drop(columns=drop_column)

                            # 資料轉型為數值型
                            years_data = df.columns[1:]
                            df[years_data] = df[years_data].apply(pd.to_numeric, errors='coerce')

                            # *** 關鍵修改：反轉年份欄位的順序 ***
                            first_col = df.columns[0]  # 第一欄
                            year_cols = df.columns[1:]  # 年份欄位
                            reversed_year_cols = year_cols[::-1]  # 反轉年份欄位順序

                            # 重新組合欄位順序：第一欄 + 反轉的年份欄位
                            new_column_order = [first_col] + list(reversed_year_cols)
                            df_reordered = df[new_column_order]

                            # 將表格資料寫入指定位置並調整格式
                            start_row = starting_cell[df_amount][1]
                            start_col = starting_cell[df_amount][2]

                            for r_idx, row in enumerate(dataframe_to_rows(df_reordered, index=False, header=True),
                                                        start=start_row):
                                row_data = list(row)

                                # 第一欄（項目名稱）寫入起始欄位
                                cell = ws.cell(row=r_idx, column=start_col, value=row_data[0])
                                cell.font = Font(size=12, bold=(r_idx == start_row))

                                # 年份數據從右邊開始寫入（從起始欄位+11開始往左）
                                year_data = row_data[1:]  # 除了第一欄以外的年份數據

                                # 從起始欄位+11開始往左寫（假設有12欄的空間）
                                for year_idx, value in enumerate(year_data):
                                    column_position = (start_col + 11) - year_idx  # 從右邊往左寫
                                    cell = ws.cell(row=r_idx, column=column_position, value=value)
                                    cell.font = Font(size=12, bold=(r_idx == start_row))

                            # 自動調整欄寬
                            for col in ws.columns:
                                max_length = max(len(str(cell.value or '')) for cell in col)
                                ws.column_dimensions[col[0].column_letter].width = max_length + 2

                # 儲存到base64
                output_buffer = io.BytesIO()
                wb.save(output_buffer)
                output_buffer.seek(0)
                modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

                return modified_base64, f"{stock}的Financial成功寫入"

            except Exception as e:
                return excel_base64, f"處理Financial資料時發生錯誤: {e}"
        else:
            return excel_base64, '無原始資料'

    async def process_df_ratios(self, raw_df_ratios, stock, excel_base64):
        """處理ratios數據並寫入Excel base64"""
        if raw_df_ratios:
            try:
                excel_binary = base64.b64decode(excel_base64)
                excel_buffer = io.BytesIO(excel_binary)
                wb = load_workbook(excel_buffer)
                ws = wb.worksheets[0]  # 使用第一個工作表

                # 定義各類財務數據的起始位置（對應A.py的7個類別）
                starting_cell = [
                    ('Profitability', 1, 53), ('Credit', 1, 66),
                    ('Liquidity', 1, 79), ('Working Capital', 1, 92),
                    ('Enterprise Value', 1, 105), ('Multiples', 1, 118),
                    ('Per Share Data Items', 1, 131)
                ]

                # 清除舊資料（對應A.py的7個區域）
                for row in ws.iter_rows(min_row=1, min_col=53, max_row=100, max_col=64):
                    for cell in row:
                        cell.value = None
                for row in ws.iter_rows(min_row=1, min_col=66, max_row=100, max_col=77):
                    for cell in row:
                        cell.value = None
                for row in ws.iter_rows(min_row=1, min_col=79, max_row=100, max_col=90):
                    for cell in row:
                        cell.value = None
                for row in ws.iter_rows(min_row=1, min_col=92, max_row=100, max_col=103):
                    for cell in row:
                        cell.value = None
                for row in ws.iter_rows(min_row=1, min_col=105, max_row=100, max_col=116):
                    for cell in row:
                        cell.value = None
                for row in ws.iter_rows(min_row=1, min_col=118, max_row=100, max_col=129):
                    for cell in row:
                        cell.value = None
                for row in ws.iter_rows(min_row=1, min_col=131, max_row=100, max_col=142):
                    for cell in row:
                        cell.value = None

                if raw_df_ratios.get(stock) == [f'{stock}是非美國企業，此頁面須付費！']:
                    return excel_base64, f'{stock}是非美國企業，此頁面須付費！'
                else:
                    d_1_raw_df_ratios = [y for x in raw_df_ratios.get(stock, []) for y in x]

                    for df_amount, df in enumerate(d_1_raw_df_ratios):
                        df_column_list = df.columns.tolist()

                        # 篩選需要的年份
                        years_list = []
                        pattern = r'\d{4}\sY'
                        for years in df_column_list:
                            if re.match(pattern, years):
                                years_list.append(int(years.split()[0]))

                        if years_list:
                            end_year = max(years_list)
                            start_year = end_year - 10
                            drop_column = [
                                x for x in df_column_list
                                if not (re.match(pattern, x) and start_year <= int(x.split()[0]) <= end_year)
                            ]
                            drop_column.pop(0)
                            df = df.drop(columns=drop_column)

                            # 資料轉型為數值型
                            years_data = df.columns[1:]
                            df[years_data] = df[years_data].apply(pd.to_numeric, errors='coerce')

                            # *** 關鍵修改：反轉年份欄位的順序 ***
                            first_col = df.columns[0]  # 第一欄
                            year_cols = df.columns[1:]  # 年份欄位
                            reversed_year_cols = year_cols[::-1]  # 反轉年份欄位順序

                            # 重新組合欄位順序：第一欄 + 反轉的年份欄位
                            new_column_order = [first_col] + list(reversed_year_cols)
                            df_reordered = df[new_column_order]

                            # 將表格資料寫入指定位置並調整格式
                            start_row = starting_cell[df_amount][1]
                            start_col = starting_cell[df_amount][2]

                            for r_idx, row in enumerate(dataframe_to_rows(df_reordered, index=False, header=True),
                                                        start=start_row):
                                row_data = list(row)

                                # 第一欄（項目名稱）寫入起始欄位
                                cell = ws.cell(row=r_idx, column=start_col, value=row_data[0])
                                cell.font = Font(size=12, bold=(r_idx == start_row))

                                # 年份數據從右邊開始寫入（從起始欄位+11開始往左）
                                year_data = row_data[1:]  # 除了第一欄以外的年份數據

                                # 從起始欄位+11開始往左寫（假設有12欄的空間）
                                for year_idx, value in enumerate(year_data):
                                    column_position = (start_col + 11) - year_idx  # 從右邊往左寫
                                    cell = ws.cell(row=r_idx, column=column_position, value=value)
                                    cell.font = Font(size=12, bold=(r_idx == start_row))

                            # 自動調整欄寬
                            for col in ws.columns:
                                max_length = max(len(str(cell.value or '')) for cell in col)
                                ws.column_dimensions[col[0].column_letter].width = max_length + 2

                # 儲存到base64
                output_buffer = io.BytesIO()
                wb.save(output_buffer)
                output_buffer.seek(0)
                modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

                return modified_base64, f"{stock}的Ratios成功寫入"

            except Exception as e:
                return excel_base64, f"處理Ratios資料時發生錯誤: {e}"
        else:
            return excel_base64, '無原始資料'

    async def EPS_PE_MarketCap_data_write_to_excel(self, EPS_PE_MarketCap_content, stock, excel_base64):
        """將 EPS_PE_MarketCap 數據寫入 Excel base64"""
        try:
            excel_binary = base64.b64decode(excel_base64)
            excel_buffer = io.BytesIO(excel_binary)
            wb = load_workbook(excel_buffer)
            ws = wb.worksheets[0]

            # 處理資料
            for data in EPS_PE_MarketCap_content.get(stock, {}):
                start_cell = "EN1"
                start_row = int(start_cell[2:])  # 提取行號，例如 "1"

                for i, (key, value) in enumerate(data.items()):
                    row = start_row + i  # 從起始行開始逐行寫入
                    ws[f"EN{row}"] = key  # 寫入鍵到 EN 列
                    ws[f"EO{row}"] = value  # 寫入值到 EO 列

            # 儲存到 base64
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            return modified_base64, f'{stock}的EPS_PE_MarketCap成功寫入及儲存成功'

        except Exception as e:
            return excel_base64, f"處理 EPS_PE_MarketCap 時發生錯誤: {e}"

    async def _rate_limit(self, api_key="yfinance"):
        """實施速率限制"""
        current_time = time.time()

        if api_key not in self.last_request_time:
            self.last_request_time[api_key] = 0

        time_since_last_request = current_time - self.last_request_time[api_key]

        if time_since_last_request < self.request_delay:
            sleep_time = self.request_delay - time_since_last_request
            # 添加隨機延遲，避免所有請求同時發送
            sleep_time += random.uniform(0.5, 1.5)
            print(f"⏳ 等待 {sleep_time:.1f} 秒以避免API限制...")
            await asyncio.sleep(sleep_time)

        self.last_request_time[api_key] = time.time()

    async def _fetch_stock_data_with_retry(self, stock, max_retries=3):
        """帶重試機制的數據獲取"""
        for attempt in range(max_retries):
            try:
                return await asyncio.to_thread(self._fetch_stock_data, stock)
            except Exception as e:
                if attempt == max_retries - 1:  # 最後一次嘗試
                    raise e

                # 指數退避：每次重試等待時間加倍
                wait_time = (2 ** attempt) * 3 + random.uniform(2, 5)
                print(f"⚠️ 獲取 {stock} 資料失敗，{wait_time:.1f}秒後重試... (嘗試 {attempt + 1}/{max_retries})")
                await asyncio.sleep(wait_time)

    def _fetch_stock_data(self, stock):
        """同步獲取股票數據"""
        # 查詢 10 年期美國國債收益率
        tnx = yf.Ticker("^TNX")
        rf_rate = tnx.info['previousClose'] / 100

        # 獲取股票資料
        Stock = yf.Ticker(stock)
        beta = Stock.info['beta']
        currentPrice = Stock.info['currentPrice']
        symbol = Stock.info['symbol']

        return {
            'Stock': symbol,
            'CurrentPrice': currentPrice,
            'beta': beta,
            'rf_rate': rf_rate
        }

    async def others_data(self, stock, excel_base64):
        """抓取其他數據並寫入Excel base64"""
        async with self.semaphore:  # 限制併發數量
            try:
                # 添加請求延遲，避免頻率過高
                await self._rate_limit("yfinance")

                # 使用重試機制獲取數據
                dic_data = await self._fetch_stock_data_with_retry(stock)

                print(f'{stock}: {dic_data}')

                # 寫入 Excel（移到線程中執行避免阻塞）
                modified_base64 = await self._write_to_excel(excel_base64, dic_data)

                return modified_base64, f'{stock}的其他資料成功寫入'

            except Exception as e:
                return excel_base64, f"獲取 {stock} 資料時發生錯誤：{str(e)}"

    async def _write_to_excel(self, excel_base64, dic_data):
        """寫入Excel文件"""

        def write_excel():
            excel_binary = base64.b64decode(excel_base64)
            excel_buffer = io.BytesIO(excel_binary)
            wb = load_workbook(excel_buffer)

            ws = wb.worksheets[0]  # 選擇第一個工作表
            ws['EQ2'] = dic_data['Stock']
            ws['ER2'] = dic_data['CurrentPrice']

            ws = wb.worksheets[3]
            ws['C31'] = dic_data['beta']
            ws['C36'] = dic_data['rf_rate']

            # 儲存到base64
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            return base64.b64encode(output_buffer.read()).decode('utf-8')

        return await asyncio.to_thread(write_excel)

    def save_excel_to_file(self, base64_data: str, output_path: str) -> bool:
        """將 base64 編碼的 Excel 資料保存為實體檔案"""
        try:
            excel_binary = base64.b64decode(base64_data)
            with open(output_path, 'wb') as f:
                f.write(excel_binary)
            print(f"Excel 檔案已保存至：{output_path}")
            return True
        except Exception as e:
            print(f"保存檔案時發生錯誤: {e}")
            return False


class StockManager:
    def __init__(self, scraper, processor, max_concurrent=3, delay=1):
        self.scraper = scraper
        self.processor = processor
        self.pattern1 = r'^[a-zA-Z\-\.]{1,5}'
        self.pattern2 = r'是非美國企業，此頁面須付費！$'
        self.semaphore = asyncio.Semaphore(max_concurrent)
        self.delay = delay
        self.excel_files = {}  # 儲存每支股票的Excel base64
        self.max_concurrent = max_concurrent

    async def initialize_excel_files(self, stocks):
        """為所有股票初始化Excel檔案"""
        for stock in stocks:
            excel_base64, message = self.processor.create_excel_from_base64(stock)
            if excel_base64:
                self.excel_files[stock] = excel_base64
                print(f"✅ {message}")
            else:
                print(f"❌ {message}")
                return False
        return True

    async def process_summary(self, stocks):
        """處理Summary數據"""
        raw_df_summary = await self.scraper.run_summary()
        for index, stock in enumerate(stocks):
            if stock in self.excel_files:
                modified_base64, message = await self.processor.process_df_summary(
                    raw_df_summary[index], stock, self.excel_files[stock]
                )
                self.excel_files[stock] = modified_base64
                print(f"✅ {message}")

    async def process_financial(self, stocks):
        """處理Financial數據"""
        raw_df_financial = await self.scraper.run_financial()
        for index, stock in enumerate(stocks):
            if stock in self.excel_files:
                modified_base64, message = await self.processor.process_df_financial(
                    raw_df_financial[index], stock, self.excel_files[stock]
                )
                self.excel_files[stock] = modified_base64
                print(f"✅ {message}")

    async def process_ratios(self, stocks):
        """處理Ratios數據"""
        raw_df_ratios = await self.scraper.run_ratios()
        for index, stock in enumerate(stocks):
            if stock in self.excel_files:
                modified_base64, message = await self.processor.process_df_ratios(
                    raw_df_ratios[index], stock, self.excel_files[stock]
                )
                self.excel_files[stock] = modified_base64
                print(f"✅ {message}")

    async def process_EPS_PE_MarketCap(self, stocks):
        """處理EPS/PE/MarketCap數據"""
        raw_df_EPS_PE_MarketCap = await self.scraper.run_EPS_PE_MarketCap()
        for index, stock in enumerate(stocks):
            if stock in self.excel_files:
                modified_base64, message = await self.processor.EPS_PE_MarketCap_data_write_to_excel(
                    raw_df_EPS_PE_MarketCap[index], stock, self.excel_files[stock]
                )
                self.excel_files[stock] = modified_base64
                print(f"✅ {message}")

    async def process_others_data(self, stocks):
        """處理其他數據"""
        for stock in stocks:
            if stock in self.excel_files:
                modified_base64, message = await self.processor.others_data(
                    stock, self.excel_files[stock]
                )
                self.excel_files[stock] = modified_base64
                print(f"✅ {message}")

    async def process_EPS_Growth_Rate(self, stocks):
        """處理EPS成長率"""
        for stock in stocks:
            if stock in self.excel_files:
                message, modified_base64 = await self.scraper.EPS_Growth_Rate_and_write_to_excel(
                    stock, self.excel_files[stock]
                )
                self.excel_files[stock] = modified_base64
                print(f"✅ {message}")

    def save_all_excel_files(self, stocks, output_folder=None):
        """保存所有Excel檔案"""
        if output_folder is None:
            output_folder = os.getcwd()

        saved_files = []
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        for stock in stocks:
            if stock in self.excel_files:
                output_filename = f"STOCK_{stock}.xlsx"
                output_path = os.path.join(output_folder, output_filename)

                if self.processor.save_excel_to_file(self.excel_files[stock], output_path):
                    saved_files.append(output_path)
                    print(f"✅ {stock} 檔案已保存至：{output_path}")
                else:
                    print(f"❌ {stock} 檔案保存失敗")

        return saved_files

    async def process_wacc_data(self, stocks):
        """處理WACC數據"""
        wacc_scraper = StockScraper(stocks, headless=True, max_concurrent=2)
        await wacc_scraper.setup_browser()

        try:
            wacc_results = await wacc_scraper.get_multiple_wacc_data(stocks)

            for stock in stocks:
                if stock in self.excel_files and stock in wacc_results:
                    modified_base64, message = self.processor.write_wacc_to_excel_base64(
                        stock, wacc_results[stock], self.excel_files[stock]
                    )
                    self.excel_files[stock] = modified_base64
                    print(f"✅ {message}")
        finally:
            await wacc_scraper.cleanup()


# ===== 新增：股票代碼驗證器類別 =====
class StockValidator:
    """股票代碼驗證器"""

    def __init__(self):
        self.valid_stocks = []
        self.invalid_stocks = []

    def validate_single_stock(self, stock):
        """驗證單一股票代碼"""
        try:
            # 使用 yfinance 獲取歷史數據來驗證股票
            ticker = yf.Ticker(stock)

            # 先嘗試獲取基本歷史數據，避免使用 .info 屬性
            hist = ticker.history(period="5d")

            if not hist.empty:
                return True, f"✅ {stock}: 有效股票代碼"
            else:
                return False, f"❌ {stock}: 無法獲得股價資訊"

        except Exception as e:
            return False, f"❌ {stock}: 驗證失敗 - {str(e)}"

    async def validate_stocks_async(self, stocks, log_callback=None):
        """異步驗證多個股票代碼"""
        self.valid_stocks = []
        self.invalid_stocks = []

        if log_callback:
            log_callback("🔍 開始驗證股票代碼...")

        # 使用線程池執行同步的股票驗證
        with ThreadPoolExecutor(max_workers=5) as executor:
            # 創建任務列表
            tasks = []
            for stock in stocks:
                task = asyncio.get_event_loop().run_in_executor(
                    executor, self.validate_single_stock, stock
                )
                tasks.append((stock, task))

            # 等待所有驗證完成
            for stock, task in tasks:
                try:
                    is_valid, message = await task

                    if log_callback:
                        log_callback(message)

                    if is_valid:
                        self.valid_stocks.append(stock)
                    else:
                        self.invalid_stocks.append(stock)

                    # 添加小延遲避免API限制
                    await asyncio.sleep(0.5)

                except Exception as e:
                    error_msg = f"❌ {stock}: 驗證過程發生錯誤 - {str(e)}"
                    if log_callback:
                        log_callback(error_msg)
                    self.invalid_stocks.append(stock)

        if log_callback:
            log_callback(f"🎯 驗證完成！有效股票: {len(self.valid_stocks)}，無效股票: {len(self.invalid_stocks)}")

        return self.valid_stocks, self.invalid_stocks


# ===== GUI 部分 =====
class StockAnalyzerGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("股票爬蟲程式 v2.0")
        self.root.geometry("1400x1000")
        self.root.configure(bg='#1a1a1a')  # 深色背景
        self.root.minsize(1200, 900)

        # 設定樣式
        self.style = ttk.Style()
        self.style.theme_use('clam')

        # 自訂顏色主題
        self.setup_custom_styles()

        # 變數
        self.stocks_var = tk.StringVar()
        self.output_folder_var = tk.StringVar(value=os.getcwd())
        self.is_running = False

        self.setup_ui()

    def setup_custom_styles(self):
        """設定現代化樣式"""
        # 深色主題配色
        bg_dark = '#1a1a1a'
        bg_card = '#2d2d2d'
        accent_blue = '#00d4aa'
        accent_orange = '#ff6b35'
        text_primary = '#ffffff'
        text_secondary = '#b0b0b0'

        # 配置主框架樣式
        self.style.configure('Card.TFrame',
                             background=bg_card,
                             relief='flat',
                             borderwidth=1)

        # 配置標籤框架樣式
        self.style.configure('Card.TLabelframe',
                             background=bg_card,
                             foreground=text_primary,
                             borderwidth=2,
                             relief='flat')

        self.style.configure('Card.TLabelframe.Label',
                             background=bg_card,
                             foreground=accent_blue,
                             font=('Microsoft JhengHei', 12, 'bold'))

        # 主要按鈕樣式
        self.style.configure('Primary.TButton',
                             font=('Microsoft JhengHei', 11, 'bold'),
                             foreground='white',
                             focuscolor='none',
                             borderwidth=0,
                             padding=(20, 10))
        self.style.map('Primary.TButton',
                       background=[('active', '#00b894'), ('!active', accent_blue)])

        # 停止按鈕樣式
        self.style.configure('Danger.TButton',
                             font=('Microsoft JhengHei', 11, 'bold'),
                             foreground='white',
                             focuscolor='none',
                             borderwidth=0,
                             padding=(20, 10))
        self.style.map('Danger.TButton',
                       background=[('active', '#e84393'), ('!active', accent_orange)])

        # 瀏覽按鈕樣式
        self.style.configure('Secondary.TButton',
                             font=('Microsoft JhengHei', 9),
                             foreground=text_primary,
                             focuscolor='none',
                             borderwidth=1,
                             padding=(15, 8))
        self.style.map('Secondary.TButton',
                       background=[('active', '#636e72'), ('!active', '#74b9ff')])

        # 標籤樣式
        self.style.configure('Title.TLabel',
                             background=bg_card,
                             foreground=text_primary,
                             font=('Microsoft JhengHei', 14))

        self.style.configure('Subtitle.TLabel',
                             background=bg_card,
                             foreground=text_secondary,
                             font=('Microsoft JhengHei', 10))

        # 輸入框樣式
        self.style.configure('Modern.TEntry',
                             fieldbackground='#3d3d3d',
                             foreground=text_primary,
                             borderwidth=1,
                             insertcolor=text_primary,
                             selectbackground=accent_blue)

        # 進度條樣式
        self.style.configure('Modern.Horizontal.TProgressbar',
                             background=accent_blue,
                             troughcolor='#3d3d3d',
                             borderwidth=0,
                             lightcolor=accent_blue,
                             darkcolor=accent_blue)

    def setup_ui(self):
        # 主框架 - 添加漸層效果
        main_frame = tk.Frame(self.root, bg='#1a1a1a')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # 標題區域 - 縮小高度
        title_frame = tk.Frame(main_frame, bg='#2d2d2d', relief='flat', bd=2)
        title_frame.pack(fill=tk.X, pady=(0, 15))

        # 縮小標題區域的內邊距
        title_content = tk.Frame(title_frame, bg='#2d2d2d')
        title_content.pack(fill=tk.X, padx=25, pady=15)

        # 縮小主標題字體
        title_label = tk.Label(title_content,
                               text="📊 股票爬蟲程式",
                               font=('標楷體', 22, 'bold'),  # 從28減少到22
                               foreground='#00d4aa',
                               bg='#2d2d2d')
        title_label.pack()

        # 縮小副標題字體和內容
        subtitle_label = tk.Label(title_content,
                                  text="專業級股票數據爬蟲工具 | Version 2.0",  # 合併成一行
                                  font=('標楷體', 16),  # 從18減少到12
                                  foreground='#b0b0b0',
                                  bg='#2d2d2d')
        subtitle_label.pack(pady=(5, 0))

        # 輸入區域框架 - 縮小間距
        input_frame = tk.Frame(main_frame, bg='#2d2d2d', relief='flat', bd=2)
        input_frame.pack(fill=tk.X, pady=(0, 15))

        input_content = tk.Frame(input_frame, bg='#2d2d2d')
        input_content.pack(fill=tk.X, padx=20, pady=15)

        # 縮小區域標題
        input_title = tk.Label(input_content,
                               text="🔍 爬蟲設定",
                               font=('標楷體', 16, 'bold'),  # 從18減少到14
                               foreground='#00d4aa',
                               bg='#2d2d2d')
        input_title.pack(anchor=tk.W, pady=(0, 10))

        # 股票代碼輸入區 - 縮小間距和字體
        stock_frame = tk.Frame(input_content, bg='#2d2d2d')
        stock_frame.pack(fill=tk.X, pady=(0, 10))

        tk.Label(stock_frame,
                 text="💼 股票代碼",
                 font=('標楷體', 14, 'bold'),  # 從14減少到12
                 foreground='#ffffff',
                 bg='#2d2d2d').pack(anchor=tk.W, pady=(0, 5))

        stocks_entry = tk.Entry(stock_frame,
                                textvariable=self.stocks_var,
                                font=('Consolas', 12),  # 從12減少到11
                                bg='#3d3d3d',
                                fg='#ffffff',
                                insertbackground='#00d4aa',
                                selectbackground='#00d4aa',
                                selectforeground='#000000',
                                relief='flat',
                                bd=2)
        stocks_entry.pack(fill=tk.X, ipady=6)

        # 縮小說明文字
        help_label = tk.Label(stock_frame,
                              text="💡 輸入股票代碼，多個代碼請用逗號分隔 (例如: NVDA, MSFT, AAPL, GOOGL)\n💡 請勿輸入非美國股票代碼",
                              font=('Times New Roman', 12),  # 從12減少到10
                              foreground='#ffb347',
                              bg='#2d2d2d',
                              justify=tk.LEFT)
        help_label.pack(anchor=tk.W, pady=(5, 0))

        # 輸出資料夾選擇 - 縮小間距
        folder_frame = tk.Frame(input_content, bg='#2d2d2d')
        folder_frame.pack(fill=tk.X, pady=(10, 0))

        tk.Label(folder_frame,
                 text="📁 輸出資料夾",
                 font=('標楷體', 14, 'bold'),  # 從14減少到12
                 foreground='#ffffff',
                 bg='#2d2d2d').pack(anchor=tk.W, pady=(0, 5))

        folder_input_frame = tk.Frame(folder_frame, bg='#2d2d2d')
        folder_input_frame.pack(fill=tk.X)

        folder_entry = tk.Entry(folder_input_frame,
                                textvariable=self.output_folder_var,
                                font=('Consolas', 12),  # 從11減少到10
                                bg='#3d3d3d',
                                fg='#ffffff',
                                insertbackground='#00d4aa',
                                relief='flat',
                                bd=2)
        folder_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=5)

        browse_btn = tk.Button(folder_input_frame,
                               text="🔍 瀏覽",
                               command=self.browse_folder,
                               font=('新細明體', 12, 'bold'),  # 從12減少到10
                               bg='#74b9ff',
                               fg='white',
                               activebackground='#0984e3',
                               activeforeground='white',
                               relief='flat',
                               bd=0,
                               cursor='hand2')
        browse_btn.pack(side=tk.RIGHT, padx=(8, 0), ipady=5, ipadx=12)

        # 控制區域框架 - 縮小間距
        control_frame = tk.Frame(main_frame, bg='#2d2d2d', relief='flat', bd=2)
        control_frame.pack(fill=tk.X, pady=(0, 15))

        control_content = tk.Frame(control_frame, bg='#2d2d2d')
        control_content.pack(fill=tk.X, padx=20, pady=15)

        # 縮小控制區域標題
        control_title = tk.Label(control_content,
                                 text="🎮 分析控制",
                                 font=('標楷體', 16, 'bold'),  # 從18減少到14
                                 foreground='#00d4aa',
                                 bg='#2d2d2d')
        control_title.pack(anchor=tk.W, pady=(0, 10))

        # 按鈕區 - 縮小按鈕大小
        button_frame = tk.Frame(control_content, bg='#2d2d2d')
        button_frame.pack(pady=(0, 15))

        self.start_btn = tk.Button(button_frame,
                                   text="🚀 開始爬蟲",
                                   command=self.start_analysis,
                                   font=('標楷體', 15, 'bold'),  # 從16減少到13
                                   bg='#00d4aa',
                                   fg='white',
                                   activebackground='#00b894',
                                   activeforeground='white',
                                   relief='flat',
                                   bd=0,
                                   cursor='hand2',
                                   width=15,  # 從15減少到12
                                   height=2)  # 從2減少到1
        self.start_btn.pack(side=tk.LEFT, padx=(0, 15))

        self.stop_btn = tk.Button(button_frame,
                                  text="⏹️ 停止爬蟲",
                                  command=self.stop_analysis,
                                  font=('標楷體', 15, 'bold'),  # 從16減少到13
                                  bg='#ff6b35',
                                  fg='white',
                                  activebackground='#e84393',
                                  activeforeground='white',
                                  relief='flat',
                                  bd=0,
                                  cursor='hand2',
                                  width=15,  # 從15減少到12
                                  height=2,  # 從2減少到1
                                  state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT)

        # 進度區域 - 縮小間距
        progress_frame = tk.Frame(control_content, bg='#2d2d2d')
        progress_frame.pack(fill=tk.X, pady=(0, 10))

        tk.Label(progress_frame,
                 text="📊 爬蟲進度",
                 font=('標楷體', 12, 'bold'),  # 從12減少到11
                 foreground='#ffffff',
                 bg='#2d2d2d').pack(anchor=tk.W, pady=(0, 5))

        # 縮小進度條高度
        progress_container = tk.Frame(progress_frame, bg='#3d3d3d', height=8)  # 從8減少到6
        progress_container.pack(fill=tk.X, pady=(0, 8))
        progress_container.pack_propagate(False)

        self.progress = ttk.Progressbar(progress_container,
                                        mode='indeterminate',
                                        style='Modern.Horizontal.TProgressbar')
        self.progress.pack(fill=tk.BOTH, expand=True)

        # 縮小狀態標籤
        self.status_label = tk.Label(control_content,
                                     text="✅ 系統準備就緒",
                                     font=('標楷體', 13, 'bold'),  # 從13減少到11
                                     foreground='#00d4aa',
                                     bg='#2d2d2d')
        self.status_label.pack()

        # 日誌區域框架 - 這裡是最重要的部分，讓它佔用更多空間
        log_frame = tk.Frame(main_frame, bg='#2d2d2d', relief='flat', bd=2)
        log_frame.pack(fill=tk.BOTH, expand=True)  # 確保日誌區域可以擴展

        log_content = tk.Frame(log_frame, bg='#2d2d2d')
        log_content.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        # 縮小日誌標題
        log_title = tk.Label(log_content,
                             text="📋 執行日誌",
                             font=('標楷體', 16, 'bold'),  # 從18減少到14
                             foreground='#00d4aa',
                             bg='#2d2d2d')
        log_title.pack(anchor=tk.W, pady=(0, 8))

        # 放大滾動文字框 - 這是關鍵改進
        self.log_text = scrolledtext.ScrolledText(log_content,
                                                  font=('Consolas', 12),  # 稍微增加字體大小，從11到12
                                                  bg='#1a1a1a',
                                                  fg='#00ff00',
                                                  insertbackground='#00d4aa',
                                                  selectbackground='#00d4aa',
                                                  selectforeground='#000000',
                                                  relief='flat',
                                                  bd=2,
                                                  wrap=tk.WORD)  # 添加自動換行
        self.log_text.pack(fill=tk.BOTH, expand=True)  # 確保日誌文字框能夠擴展

        # 初始化日誌
        self.log_text.insert(tk.END, "=== 股票爬蟲程式已啟動 ===\n")
        self.log_text.insert(tk.END, "系統準備就緒，請輸入股票代碼開始爬蟲...\n\n")

    def browse_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.output_folder_var.set(folder)

    def log(self, message):
        """現代化日誌顯示"""
        timestamp = datetime.now().strftime("%H:%M:%S")

        # 根據訊息類型選擇顏色
        if "✅" in message or "成功" in message:
            color = "#00ff00"  # 綠色
        elif "❌" in message or "錯誤" in message or "失敗" in message:
            color = "#ff4757"  # 紅色
        elif "⚠️" in message or "警告" in message:
            color = "#ffa502"  # 橙色
        elif "🔄" in message or "處理" in message:
            color = "#3742fa"  # 藍色
        elif "🚀" in message or "開始" in message:
            color = "#ff6b35"  # 橙紅色
        else:
            color = "#ffffff"  # 白色

        # 配置顏色標籤
        tag_name = f"color_{color.replace('#', '')}"
        self.log_text.tag_configure(tag_name, foreground=color)
        self.log_text.tag_configure("timestamp", foreground="#70a1ff")

        # 插入訊息
        self.log_text.insert(tk.END, f"[{timestamp}] ", "timestamp")
        self.log_text.insert(tk.END, f"{message}\n", tag_name)

        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def update_status(self, status):
        """更新狀態標籤"""
        if "完成" in status or "成功" in status:
            color = "#00d4aa"
            icon = "✅"
        elif "失敗" in status or "錯誤" in status:
            color = "#ff4757"
            icon = "❌"
        elif "停止" in status:
            color = "#ffa502"
            icon = "⏹️"
        elif "步驟" in status or "處理" in status:
            color = "#3742fa"
            icon = "🔄"
        else:
            color = "#ffffff"
            icon = "📊"

        self.status_label.config(text=f"{icon} {status}", foreground=color)
        self.root.update_idletasks()

    def start_analysis(self):
        """開始分析 - 加入輸入驗證"""
        # 檢查Excel模板
        if EXCEL_TEMPLATE_BASE64.strip() == "" or "我的模板" in EXCEL_TEMPLATE_BASE64:
            messagebox.showerror("❌ 錯誤",
                                 "請先設定 EXCEL_TEMPLATE_BASE64 變數！\n請將Excel模板轉換為base64後貼入程式碼中。")
            return

        # 獲取輸入的股票代碼
        stocks_input = self.stocks_var.get().strip()
        if not stocks_input:
            messagebox.showwarning("⚠️ 警告", "請輸入至少一個股票代碼！")
            return

        # 處理股票代碼列表，移除空白和重複
        stocks_raw = [s.strip().upper() for s in stocks_input.split(',')]
        stocks = []

        # 過濾空白和重複的股票代碼
        seen = set()
        for stock in stocks_raw:
            if stock and stock not in seen:
                stocks.append(stock)
                seen.add(stock)

        if not stocks:
            messagebox.showwarning("⚠️ 警告", "請輸入有效的股票代碼！")
            return

        # 確認開始（顯示即將驗證的股票）
        confirmation_message = (
            f"即將驗證並爬蟲以下股票：\n"
            f"📈 {', '.join(stocks)}\n\n"
            f"🔍 系統將先驗證股票代碼有效性\n"
            f"📊 僅爬蟲有效的股票代碼\n"
            f"🔥 預計需要數分鐘時間\n\n"
            f"是否開始？"
        )

        if not messagebox.askyesno("🚀 確認開始", confirmation_message):
            return

        # 禁用按鈕
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.is_running = True

        # 清空日誌
        self.log_text.delete(1.0, tk.END)

        # 開始進度條
        self.progress.start()

        # 在新線程中執行分析
        thread = threading.Thread(target=self.run_analysis, args=(stocks,))
        thread.daemon = True
        thread.start()

    def stop_analysis(self):
        """停止分析"""
        self.is_running = False
        self.update_status("正在停止爬蟲...")
        self.log("🛑 使用者請求停止爬蟲")

    def run_analysis(self, stocks):
        """執行分析的主函數"""
        try:
            # 創建新的事件循環
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)

            # 執行異步分析
            loop.run_until_complete(self.async_analysis(stocks))

        except Exception as e:
            self.log(f"❌ 發生錯誤：{str(e)}")
            messagebox.showerror("❌ 錯誤", f"爬蟲過程中發生錯誤：\n{str(e)}")

        finally:
            # 恢復按鈕狀態
            self.start_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)
            self.progress.stop()
            self.is_running = False

    async def async_analysis(self, stocks):
        """異步執行分析 - 增強日誌顯示並加入股票代碼驗證"""
        try:
            self.log("🎯" + "=" * 80)
            self.log("🚀 股票爬蟲系統啟動")
            self.log(f"📊 輸入股票：{', '.join(stocks)}")
            self.log(f"🔢 輸入數量：{len(stocks)} 支")
            self.log("🎯" + "=" * 80)

            start_time = time.time()

            # 新增：股票代碼驗證步驟
            self.update_status("驗證股票代碼有效性")
            self.log("\n🔍 步驟 0/7：正在驗證股票代碼...")

            validator = StockValidator()
            valid_stocks, invalid_stocks = await validator.validate_stocks_async(
                stocks, log_callback=self.log
            )

            # 如果有無效股票，顯示警告
            if invalid_stocks:
                self.log("\n⚠️ 發現無效股票代碼:")
                for invalid_stock in invalid_stocks:
                    self.log(f"   ❌ {invalid_stock}")

            # 如果沒有有效股票，停止分析
            if not valid_stocks:
                self.log("❌ 沒有找到任何有效的股票代碼，停止爬蟲")
                self.update_status("爬蟲失敗：無有效股票代碼")
                return

            # 更新要分析的股票列表
            stocks = valid_stocks
            self.log(f"\n✅ 將爬蟲以下有效股票：{', '.join(stocks)}")
            self.log("🎯" + "=" * 80)

            # 檢查是否被停止
            if not self.is_running:
                self.log("🛑 爬蟲被使用者停止")
                return

            # 創建分析物件（使用有效股票列表）
            self.update_status("初始化爬蟲系統")
            self.log("🔧 正在初始化爬蟲系統...")
            scraper = StockScraper(stocks=stocks, max_concurrent=3)
            processor = StockProcess(max_concurrent=2, request_delay=2.5)
            manager = StockManager(scraper, processor, max_concurrent=3)
            self.log("✅ 爬蟲系統初始化完成")

            # 步驟 1：初始化 Excel 檔案
            if not self.is_running:
                return

            self.update_status("初始化 Excel 檔案")
            self.log("\n📄 步驟 1/7：正在初始化 Excel 檔案...")

            success = await manager.initialize_excel_files(stocks)
            if not success:
                self.log("❌ Excel 檔案初始化失敗，停止爬蟲")
                self.update_status("爬蟲失敗：Excel 初始化錯誤")
                return

            self.log("✅ Excel 檔案初始化完成")

            # 步驟 2：抓取 Summary 數據
            if not self.is_running:
                return

            self.update_status("抓取 Summary 數據")
            self.log("\n📊 步驟 2/7：正在抓取 Summary 數據...")

            await manager.process_summary(stocks)
            self.log("✅ Summary 數據處理完成")

            # 步驟 3：抓取 Financial 數據
            if not self.is_running:
                return

            self.update_status("抓取 Financial 數據")
            self.log("\n💰 步驟 3/7：正在抓取 Financial 數據...")

            await manager.process_financial(stocks)
            self.log("✅ Financial 數據處理完成")

            # 步驟 4：抓取 Ratios 數據
            if not self.is_running:
                return

            self.update_status("抓取 Ratios 數據")
            self.log("\n📈 步驟 4/7：正在抓取 Ratios 數據...")

            await manager.process_ratios(stocks)
            self.log("✅ Ratios 數據處理完成")

            # 步驟 5：抓取 EPS/PE/MarketCap 數據
            if not self.is_running:
                return

            self.update_status("抓取 EPS/PE/MarketCap 數據")
            self.log("\n📊 步驟 5/7：正在抓取 EPS/PE/MarketCap 數據...")

            await manager.process_EPS_PE_MarketCap(stocks)
            self.log("✅ EPS/PE/MarketCap 數據處理完成")

            # 步驟 6：抓取其他數據
            if not self.is_running:
                return

            self.update_status("抓取其他數據")
            self.log("\n🔍 步驟 6/7：正在抓取其他數據...")

            await manager.process_others_data(stocks)
            self.log("✅ 其他數據處理完成")

            # 步驟 7：處理 EPS 成長率
            if not self.is_running:
                return

            self.update_status("處理 EPS 成長率")
            self.log("\n📈 步驟 7/7：正在處理 EPS 成長率...")

            await manager.process_EPS_Growth_Rate(stocks)
            self.log("✅ EPS 成長率處理完成")

            # 保存檔案
            if not self.is_running:
                return

            self.update_status("保存 Excel 檔案")
            self.log("\n💾 正在保存 Excel 檔案...")

            output_folder = self.output_folder_var.get()
            saved_files = manager.save_all_excel_files(stocks, output_folder)

            # 計算執行時間
            end_time = time.time()
            execution_time = end_time - start_time

            # 顯示完成摘要
            self.log("\n" + "🎉" + "=" * 80)
            self.log("🎊 股票爬蟲完成！")
            self.log(f"⏱️ 總執行時間：{execution_time:.2f} 秒")
            self.log(f"📊 成功爬蟲股票：{len(stocks)} 支")
            self.log(f"💾 保存檔案數量：{len(saved_files)} 個")
            self.log(f"📁 保存位置：{output_folder}")

            if saved_files:
                self.log("\n📋 已保存的檔案：")
                for file_path in saved_files:
                    filename = os.path.basename(file_path)
                    self.log(f"   ✅ {filename}")

            self.log("🎉" + "=" * 80)

            self.update_status("爬蟲完成！")

            # 顯示完成對話框
            messagebox.showinfo(
                "🎉 爬蟲完成",
                f"股票爬蟲已成功完成！\n\n"
                f"📊 爬蟲股票：{len(stocks)} 支\n"
                f"⏱️ 執行時間：{execution_time:.1f} 秒\n"
                f"💾 保存檔案：{len(saved_files)} 個\n"
                f"📁 保存位置：{output_folder}"
            )

        except Exception as e:
            error_msg = f"系統錯誤：{str(e)}"
            self.log(f"❌ {error_msg}")
            self.update_status("爬蟲失敗")
            messagebox.showerror("❌ 錯誤", f"爬蟲過程中發生錯誤：\n{str(e)}")
            raise e

    def run(self):
        """啟動GUI"""
        self.root.mainloop()


# ===== 主程式 =====
def main():
    """主程式入口"""
    app = StockAnalyzerGUI()
    app.run()


if __name__ == "__main__":
    main()