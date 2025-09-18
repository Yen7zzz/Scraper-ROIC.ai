import asyncio
import base64
import io
from playwright.async_api import async_playwright
import random
import re
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from excel_template.excel_template import EXCEL_TEMPLATE_BASE64

class StockScraper:
    def __init__(self, stocks, headless=True, max_concurrent=3):
        """
        初始化爬蟲類別。
        stocks: 股票代碼的列表
        headless: 是否使用無頭模式
        max_concurrent: 同時執行的股票數量（控制併發數）
        """
        self.stocks = stocks
        self.headless = headless
        self.max_concurrent = max_concurrent
        self.browser = None
        self.playwright = None

    async def setup_browser(self):
        """設定瀏覽器環境。"""
        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.launch(
            headless=self.headless,
            args=[
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",
                "--disable-accelerated-2d-canvas",
                "--disable-gpu",
                "--disable-features=IsolateOrigins,site-per-process",
                "--disable-background-timer-throttling",
                "--disable-web-security",
                "--disable-features=VizDisplayCompositor",
            ],
        )

    async def fetch_wacc_data(self, stock, semaphore):
        async with semaphore:
            try:
                context = await self.browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
                    viewport={"width": 1920, "height": 1080},
                    java_script_enabled=True,
                )
                try:
                    page_summary = await context.new_page()
                    wacc_value = await self.get_wacc_html(stock, page_summary)
                    return {stock: wacc_value}
                finally:
                    await context.close()
            except Exception as e:
                return {stock: None}  # 如果出錯返回None

    async def get_wacc_html(self, stock, page, retries=3):
        """抓取特定股票的WACC資料並回傳int數值。"""
        URL = f'https://www.gurufocus.com/term/wacc/{stock}'
        attempt = 0

        while attempt < retries:
            try:
                print(f"正在嘗試抓取 {stock} 的WACC資料 (第 {attempt + 1} 次)...")

                # 隨機等待時間
                await asyncio.sleep(random.uniform(2, 5))

                # 前往頁面
                await page.goto(URL, wait_until='networkidle', timeout=60000)

                # 等待頁面載入完成
                await page.wait_for_load_state('networkidle')

                # 等待關鍵內容載入
                try:
                    await page.wait_for_selector('h1', timeout=30000)
                    await asyncio.sleep(3)
                except Exception as e:
                    print(f"等待頁面載入時發生錯誤: {e}")

                # 獲取頁面內容
                content = await page.content()

                # 使用BeautifulSoup解析WACC數值
                soup = BeautifulSoup(content, 'html.parser')

                # 尋找包含WACC數值的特定元素
                wacc_value = None

                # 方法1: 尋找包含":X.X% (As of"模式的font標籤
                font_elements = soup.find_all('font', style=True)
                for font in font_elements:
                    text = font.get_text(strip=True)
                    if '% (As of' in text and text.startswith(':'):
                        # 提取百分比數值
                        match = re.search(r':(\d+\.?\d*)%', text)
                        if match:
                            wacc_value = float(match.group(1))
                            print(f"找到WACC值: {wacc_value}%")
                            break

                if wacc_value is not None:
                    return wacc_value
                else:
                    print(f"未能找到 {stock} 的WACC數值")
                    return None

            except Exception as e:
                print(f"第 {attempt + 1} 次嘗試失敗: {e}")
                attempt += 1
                if attempt < retries:
                    await asyncio.sleep(random.uniform(5, 10))

        print(f"Failed to retrieve WACC data for {stock} after {retries} attempts")
        return None

    async def get_multiple_wacc_data(self, stocks):
        """
        批量獲取多個股票的WACC數據
        返回格式: {'O': 8, 'AAPL': 10, 'MSFT': 9}
        """
        semaphore = asyncio.Semaphore(self.max_concurrent)
        tasks = [self.fetch_wacc_data(stock, semaphore) for stock in stocks]
        results = await asyncio.gather(*tasks)

        # 合併結果為單一字典
        final_result = {}
        for result in results:
            final_result.update(result)

        return final_result

    async def cleanup(self):
        """清理資源。"""
        if self.browser:
            await self.browser.close()
        if self.playwright:
            await self.playwright.stop()

class StockProcess:
    def __init__(self, max_concurrent=2, request_delay=2.0):
        # 將 semaphore 移到類別層級，確保全域限制
        self.semaphore = asyncio.Semaphore(max_concurrent)
        self.request_delay = request_delay  # 請求之間的延遲（秒）
        self.last_request_time = {}  # 記錄每個API的上次請求時間

    def create_excel_from_base64(self, stock):
        """從base64模板創建Excel文件的base64"""
        try:
            if EXCEL_TEMPLATE_BASE64.strip() == "" or "請將您從轉換工具得到的" in EXCEL_TEMPLATE_BASE64:
                return "", "❌ 錯誤：請先設定 EXCEL_TEMPLATE_BASE64 變數"

            excel_binary = base64.b64decode(EXCEL_TEMPLATE_BASE64.strip())
            excel_buffer = io.BytesIO(excel_binary)
            workbook = load_workbook(excel_buffer)

            # 儲存修改後的檔案到記憶體
            output_buffer = io.BytesIO()
            workbook.save(output_buffer)
            output_buffer.seek(0)
            excel_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            return excel_base64, f"成功為 {stock} 創建Excel檔案"

        except Exception as e:
            return "", f"創建Excel檔案時發生錯誤: {e}"

    def write_wacc_to_excel_base64(self, stock_code: str, wacc_value: str, excel_base64: str) -> tuple:
        """將WACC值寫入Excel的base64"""
        try:
            excel_binary = base64.b64decode(excel_base64)
            excel_buffer = io.BytesIO(excel_binary)
            workbook = load_workbook(excel_buffer)

            if "現金流折現法" in workbook.sheetnames:
                worksheet = workbook["現金流折現法"]
            else:
                worksheet = workbook.active

            # 寫入 WACC 值到 C6 儲存格
            if wacc_value != "未知":
                if wacc_value.endswith('%'):
                    numeric_value = float(wacc_value.rstrip('%')) / 100
                    worksheet['C6'] = numeric_value
                    worksheet['C6'].number_format = '0.00%'
                else:
                    worksheet['C6'] = wacc_value
            else:
                worksheet['C6'] = "無法取得"

            # 儲存修改後的檔案到記憶體
            output_buffer = io.BytesIO()
            workbook.save(output_buffer)
            output_buffer.seek(0)
            modified_base64 = base64.b64encode(output_buffer.read()).decode('utf-8')

            return modified_base64, f"已將 {stock_code} 的 WACC 值 ({wacc_value}) 寫入 C6 儲存格"

        except Exception as e:
            return excel_base64, f"操作 Excel 檔案時發生錯誤: {e}"


async def main():
    # 可以輸入多個股票代碼
    stocks = ['O', 'AAPL', 'MSFT']  # 範例：多個股票
    scraper = StockScraper(stocks, headless=True)

    try:
        await scraper.setup_browser()

        # 批量獲取WACC數據
        result = await scraper.get_multiple_wacc_data(stocks)

        print("=== 最終結果 ===")
        print(f"結果格式: {result}")
        print(f"結果型別: {type(result)}")

        # 顯示每個股票的WACC值和型別
        for stock, wacc in result.items():
            if wacc is not None:
                print(f"{stock}: {wacc} (型別: {type(wacc)})")
            else:
                print(f"{stock}: 獲取失敗")

    except Exception as e:
        print(f"執行錯誤: {e}")
    finally:
        await scraper.cleanup()


if __name__ == "__main__":
    asyncio.run(main())